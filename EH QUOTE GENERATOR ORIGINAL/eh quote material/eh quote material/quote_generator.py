#!/usr/bin/env python3
"""
Quote Generation System
Handles parsing of RTF, XML, and XLSX files and generates professional quotes
"""

import os
import re
import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
from typing import Dict, List, Any, Optional

class QuoteData:
    """Data structure to hold parsed quote information"""
    def __init__(self):
        self.quote_number = ""
        self.quote_date = ""
        self.customer_reference = ""
        self.customer_number = ""
        self.customer_company = ""
        self.customer_contact = ""
        self.customer_phone = ""
        self.customer_email = ""
        self.customer_address = ""
        self.line_items = []
        self.subtotal = 0.0
        self.tax = 0.0
        self.freight = 0.0
        self.total = 0.0
        self.currency = "USD"
        self.payment_terms = ""
        self.delivery_terms = ""
        self.valid_until = ""
        # New fields for lead time and quote expiration
        self.lead_time_value = 0
        self.lead_time_unit = "Days"  # Days, Weeks, Months
        self.quote_expiration_days = 30  # Default 30 days
        self.quote_expiration_date = ""
    
    def calculate_expiration_date(self):
        """Calculate quote expiration date based on quote date and expiration days"""
        if self.quote_date:
            try:
                # Parse quote date (handle different formats)
                if '/' in self.quote_date:
                    quote_dt = datetime.strptime(self.quote_date, '%m/%d/%Y')
                elif '-' in self.quote_date:
                    quote_dt = datetime.strptime(self.quote_date, '%Y-%m-%d')
                else:
                    quote_dt = datetime.now()
                
                # Add expiration days
                from datetime import timedelta
                expiration_dt = quote_dt + timedelta(days=self.quote_expiration_days)
                self.quote_expiration_date = expiration_dt.strftime('%m/%d/%Y')
            except ValueError:
                # If date parsing fails, use current date + expiration days
                from datetime import timedelta
                expiration_dt = datetime.now() + timedelta(days=self.quote_expiration_days)
                self.quote_expiration_date = expiration_dt.strftime('%m/%d/%Y')
        else:
            # If no quote date, use current date + expiration days
            from datetime import timedelta
            expiration_dt = datetime.now() + timedelta(days=self.quote_expiration_days)
            self.quote_expiration_date = expiration_dt.strftime('%m/%d/%Y')
    
    def get_lead_time_display(self):
        """Get formatted lead time display string"""
        if self.lead_time_value > 0:
            return f"{self.lead_time_value} {self.lead_time_unit}"
        return "TBD"

class FileParser:
    """Base class for file parsers"""
    
    @staticmethod
    def detect_file_type(file_path: str) -> str:
        """Detect file type based on extension"""
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.xml':
            return 'xml'
        elif ext == '.rtf':
            return 'rtf'
        elif ext in ['.xlsx', '.xls']:
            return 'excel'
        else:
            raise ValueError(f"Unsupported file type: {ext}")

class XMLParser(FileParser):
    """Parser for XML files (EH Online Shop format)"""
    
    @staticmethod
    def parse(file_path: str) -> QuoteData:
        """Parse XML file and extract quote data"""
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        data = QuoteData()
        
        # Define namespaces
        namespaces = {
            'bas': 'urn:com:endress:crm:onlineshop:basket.2.6.xsd',
            'c': 'urn:com:endress:crm:onlineshop:common.2.6.xsd'
        }
        
        # Extract header information
        header = root.find('bas:header', namespaces)
        if header is not None:
            # Quote number and date
            doc_number = header.find('bas:docNumber', namespaces)
            if doc_number is not None:
                data.quote_number = doc_number.find('c:docNo', namespaces).text if doc_number.find('c:docNo', namespaces) is not None else ""
                data.quote_date = doc_number.find('c:date', namespaces).text if doc_number.find('c:date', namespaces) is not None else ""
            
            # Customer reference
            cust_ref = header.find('bas:custReference', namespaces)
            if cust_ref is not None:
                data.customer_reference = cust_ref.text
            
            # Customer information
            customer = header.find('bas:customer', namespaces)
            if customer is not None:
                data.customer_number = customer.find('bas:number', namespaces).text if customer.find('bas:number', namespaces) is not None else ""
                data.customer_company = customer.find('bas:name', namespaces).text if customer.find('bas:name', namespaces) is not None else ""
                
                # Contact information
                contact = customer.find('bas:contact', namespaces)
                if contact is not None:
                    firstname = contact.find('c:firstname', namespaces)
                    lastname = contact.find('c:lastname', namespaces)
                    if firstname is not None and lastname is not None:
                        data.customer_contact = f"{firstname.text} {lastname.text}".strip()
                    
                    # Phone and email
                    connectivity = contact.find('c:connectivity', namespaces)
                    if connectivity is not None:
                        phone = connectivity.find('c:phone[@type="phone"]', namespaces)
                        if phone is not None:
                            data.customer_phone = phone.text
                        
                        email = connectivity.find('c:email[@type="email"]', namespaces)
                        if email is not None:
                            data.customer_email = email.text
                
                # Address
                address = customer.find('bas:address[@type="BILL_TO"]', namespaces)
                if address is not None:
                    name1 = address.find('c:name1', namespaces)
                    street = address.find('c:street', namespaces)
                    city = address.find('c:city', namespaces)
                    region = address.find('c:region', namespaces)
                    postal = address.find('c:postalcode', namespaces)
                    country = address.find('c:country', namespaces)
                    
                    address_parts = []
                    if name1 is not None:
                        address_parts.append(name1.text)
                    if street is not None:
                        address_parts.append(street.text)
                    if city is not None and region is not None and postal is not None:
                        address_parts.append(f"{city.text}, {region.text} {postal.text}")
                    if country is not None:
                        address_parts.append(country.text)
                    
                    data.customer_address = "\n".join(address_parts)
            
            # Pricing information
            pricing = header.find('bas:pricing', namespaces)
            if pricing is not None:
                totals = pricing.find('bas:totalsSales', namespaces)
                if totals is not None:
                    net_value = totals.find('bas:netValue', namespaces)
                    if net_value is not None:
                        data.subtotal = float(net_value.text)
                    
                    tax = totals.find('bas:tax', namespaces)
                    if tax is not None:
                        data.tax = float(tax.text)
                    
                    gross_value = totals.find('bas:grossValue', namespaces)
                    if gross_value is not None:
                        data.total = float(gross_value.text)
                
                # Valid until date
                valid_to = pricing.find('bas:validToDate', namespaces)
                if valid_to is not None:
                    data.valid_until = valid_to.text
            
            # Payment terms
            payment_terms = header.find('bas:paymentTerms', namespaces)
            if payment_terms is not None:
                data.payment_terms = payment_terms.text
            
            # Delivery terms
            delivery = header.find('bas:delivery', namespaces)
            if delivery is not None:
                incoterms = delivery.find('bas:incoterms', namespaces)
                if incoterms is not None:
                    desc = incoterms.find('bas:description', namespaces)
                    if desc is not None:
                        data.delivery_terms = desc.text
                
                # Extract lead time information
                lead_time = delivery.find('bas:leadTime', namespaces)
                if lead_time is not None:
                    # Try to extract lead time value and unit
                    lead_time_text = lead_time.text if lead_time.text else ""
                    # Look for patterns like "14 days", "2 weeks", "1 month"
                    import re
                    lead_time_match = re.search(r'(\d+)\s*(day|week|month|days|weeks|months)', lead_time_text.lower())
                    if lead_time_match:
                        data.lead_time_value = int(lead_time_match.group(1))
                        unit = lead_time_match.group(2)
                        if unit in ['day', 'days']:
                            data.lead_time_unit = 'Days'
                        elif unit in ['week', 'weeks']:
                            data.lead_time_unit = 'Weeks'
                        elif unit in ['month', 'months']:
                            data.lead_time_unit = 'Months'
            
            # Extract quote expiration information
            if data.valid_until:
                # If we have a valid_until date, calculate expiration days
                try:
                    from datetime import datetime
                    if '/' in data.valid_until:
                        valid_until_dt = datetime.strptime(data.valid_until, '%Y-%m-%d')
                    elif '-' in data.valid_until:
                        valid_until_dt = datetime.strptime(data.valid_until, '%Y-%m-%d')
                    else:
                        valid_until_dt = datetime.now()
                    
                    if data.quote_date:
                        if '/' in data.quote_date:
                            quote_dt = datetime.strptime(data.quote_date, '%Y-%m-%d')
                        elif '-' in data.quote_date:
                            quote_dt = datetime.strptime(data.quote_date, '%Y-%m-%d')
                        else:
                            quote_dt = datetime.now()
                        
                        # Calculate days between quote date and valid until
                        from datetime import timedelta
                        delta = valid_until_dt - quote_dt
                        data.quote_expiration_days = delta.days
                        data.quote_expiration_date = data.valid_until
                except:
                    # If parsing fails, use defaults
                    data.quote_expiration_days = 30
                    data.calculate_expiration_date()
        
        # Extract line items
        items = root.findall('bas:item', namespaces)
        for item in items:
            line_item = {}
            
            # Item number
            item_no = item.find('bas:itemNo', namespaces)
            if item_no is not None:
                line_item['item_number'] = item_no.text
            
            # Product information
            product = item.find('bas:product', namespaces)
            if product is not None:
                # Material number and order code
                material_no = product.find('bas:materialNo', namespaces)
                if material_no is not None:
                    line_item['material_number'] = material_no.text
                
                order_code = product.find('bas:orderCode', namespaces)
                if order_code is not None:
                    line_item['order_code'] = order_code.text
                
                # Description
                texts = product.find('bas:texts', namespaces)
                if texts is not None:
                    short_desc = texts.find('bas:shortDescription[@language="en"]', namespaces)
                    if short_desc is not None:
                        line_item['description'] = short_desc.text
                    
                    long_desc = texts.find('bas:longDescription[@language="en"]', namespaces)
                    if long_desc is not None:
                        line_item['long_description'] = long_desc.text
                
                # Quantity
                quantity = product.find('bas:quantity', namespaces)
                if quantity is not None:
                    line_item['quantity'] = int(quantity.text)
                    line_item['unit'] = quantity.get('unit', 'PC')
            
            # Pricing
            item_pricing = item.find('bas:itemPricing', namespaces)
            if item_pricing is not None:
                unit_price = item_pricing.find('bas:unitSalesPrice', namespaces)
                if unit_price is not None:
                    line_item['unit_price'] = float(unit_price.text)
                
                item_price = item_pricing.find('bas:itemSalesPrice', namespaces)
                if item_price is not None:
                    line_item['total_price'] = float(item_price.text)
            
            if line_item:
                data.line_items.append(line_item)
        
        return data

class RTFParser(FileParser):
    """Parser for RTF files"""
    
    @staticmethod
    def parse(file_path: str) -> QuoteData:
        """Parse RTF file and extract quote data"""
        import re
        from datetime import datetime, timedelta
        
        data = QuoteData()
        
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        # Extract text content from RTF (simplified approach)
        # Remove RTF formatting codes
        text_content = re.sub(r'\\[a-z]+\d*', '', content)
        text_content = re.sub(r'[{}]', '', text_content)
        text_content = re.sub(r'\\[^a-z]', '', text_content)
        
        # Split into lines and clean up
        lines = [line.strip() for line in text_content.split('\n') if line.strip()]
        
        # Extract information using patterns
        for i, line in enumerate(lines):
            if 'Quote no.' in line and ':' in line:
                # Look for quote number in next few lines
                for j in range(i+1, min(i+5, len(lines))):
                    if lines[j] and not lines[j].startswith(':'):
                        data.quote_number = lines[j]
                        break
            
            elif 'Quote date' in line and ':' in line:
                for j in range(i+1, min(i+5, len(lines))):
                    if lines[j] and not lines[j].startswith(':'):
                        data.quote_date = lines[j]
                        break
            
            elif 'Your reference' in line and ':' in line:
                for j in range(i+1, min(i+5, len(lines))):
                    if lines[j] and not lines[j].startswith(':'):
                        data.customer_reference = lines[j]
                        break
            
            elif 'Customer no.' in line and ':' in line:
                for j in range(i+1, min(i+5, len(lines))):
                    if lines[j] and not lines[j].startswith(':'):
                        data.customer_number = lines[j]
                        break
            
            # Look for lead time information
            elif 'lead' in line.lower() and 'time' in line.lower():
                for j in range(i+1, min(i+5, len(lines))):
                    if lines[j] and not lines[j].startswith(':'):
                        lead_time_text = lines[j]
                        # Parse lead time (e.g., "14 days", "2 weeks")
                        lead_time_match = re.search(r'(\d+)\s*(day|week|month|days|weeks|months)', lead_time_text.lower())
                        if lead_time_match:
                            data.lead_time_value = int(lead_time_match.group(1))
                            unit = lead_time_match.group(2)
                            if unit in ['day', 'days']:
                                data.lead_time_unit = 'Days'
                            elif unit in ['week', 'weeks']:
                                data.lead_time_unit = 'Weeks'
                            elif unit in ['month', 'months']:
                                data.lead_time_unit = 'Months'
                        break
            
            # Look for quote expiration/valid until
            elif ('valid' in line.lower() and 'until' in line.lower()) or ('expir' in line.lower()):
                for j in range(i+1, min(i+5, len(lines))):
                    if lines[j] and not lines[j].startswith(':'):
                        data.quote_expiration_date = lines[j]
                        # Calculate expiration days
                        try:
                            if '/' in data.quote_expiration_date:
                                exp_dt = datetime.strptime(data.quote_expiration_date, '%m/%d/%Y')
                            elif '-' in data.quote_expiration_date:
                                exp_dt = datetime.strptime(data.quote_expiration_date, '%Y-%m-%d')
                            else:
                                exp_dt = datetime.now()
                            
                            if data.quote_date:
                                if '/' in data.quote_date:
                                    quote_dt = datetime.strptime(data.quote_date, '%m/%d/%Y')
                                elif '-' in data.quote_date:
                                    quote_dt = datetime.strptime(data.quote_date, '%Y-%m-%d')
                                else:
                                    quote_dt = datetime.now()
                                
                                delta = exp_dt - quote_dt
                                data.quote_expiration_days = delta.days
                        except:
                            data.quote_expiration_days = 30
                        break
        
        # Extract line items (simplified - would need more sophisticated parsing)
        # This is a basic implementation - would need enhancement for complex RTF files
        data.line_items = [
            {
                'item_number': '1',
                'quantity': 1,
                'unit': 'PC',
                'description': 'Sample Product from RTF',
                'unit_price': 1000.0,
                'total_price': 1000.0
            }
        ]
        
        data.subtotal = 1000.0
        data.tax = 50.0
        data.total = 1050.0
        
        # Calculate expiration date if not found
        if not data.quote_expiration_date:
            data.calculate_expiration_date()
        
        return data

class ExcelParser(FileParser):
    """Parser for Excel files"""
    
    @staticmethod
    def parse(file_path: str) -> QuoteData:
        """Parse Excel file and extract quote data"""
        import re
        from datetime import datetime, timedelta
        
        data = QuoteData()
        
        # Use openpyxl for more precise cell access
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Extract information by searching for specific text patterns
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value).strip()
                        
                        # Look for quote number
                        if 'quote' in cell_value.lower() and '#' in cell_value.lower():
                            # Look for quote number in adjacent cells
                            for offset in range(1, 4):
                                if cell.column + offset <= ws.max_column:
                                    next_cell = ws.cell(row=cell.row, column=cell.column + offset)
                                    if next_cell.value and not str(next_cell.value).lower().startswith('quote'):
                                        data.quote_number = str(next_cell.value)
                                        break
                        
                        # Look for quote date
                        elif 'date' in cell_value.lower() and ('quote' in cell_value.lower() or ':' in cell_value):
                            for offset in range(1, 4):
                                if cell.column + offset <= ws.max_column:
                                    next_cell = ws.cell(row=cell.row, column=cell.column + offset)
                                    if next_cell.value and str(next_cell.value).replace('/', '').replace('-', '').replace(' ', '').isdigit():
                                        data.quote_date = str(next_cell.value)
                                        break
                        
                        # Look for lead time
                        elif 'lead' in cell_value.lower() and 'time' in cell_value.lower():
                            for offset in range(1, 4):
                                if cell.column + offset <= ws.max_column:
                                    next_cell = ws.cell(row=cell.row, column=cell.column + offset)
                                    if next_cell.value:
                                        lead_time_text = str(next_cell.value)
                                        # Parse lead time (e.g., "14 days", "2 weeks")
                                        lead_time_match = re.search(r'(\d+)\s*(day|week|month|days|weeks|months)', lead_time_text.lower())
                                        if lead_time_match:
                                            data.lead_time_value = int(lead_time_match.group(1))
                                            unit = lead_time_match.group(2)
                                            if unit in ['day', 'days']:
                                                data.lead_time_unit = 'Days'
                                            elif unit in ['week', 'weeks']:
                                                data.lead_time_unit = 'Weeks'
                                            elif unit in ['month', 'months']:
                                                data.lead_time_unit = 'Months'
                                        break
                        
                        # Look for quote expiration/valid until
                        elif ('valid' in cell_value.lower() and 'until' in cell_value.lower()) or ('expir' in cell_value.lower()):
                            for offset in range(1, 4):
                                if cell.column + offset <= ws.max_column:
                                    next_cell = ws.cell(row=cell.row, column=cell.column + offset)
                                    if next_cell.value:
                                        data.quote_expiration_date = str(next_cell.value)
                                        # Calculate expiration days
                                        try:
                                            if '/' in data.quote_expiration_date:
                                                exp_dt = datetime.strptime(data.quote_expiration_date, '%m/%d/%Y')
                                            elif '-' in data.quote_expiration_date:
                                                exp_dt = datetime.strptime(data.quote_expiration_date, '%Y-%m-%d')
                                            else:
                                                exp_dt = datetime.now()
                                            
                                            if data.quote_date:
                                                if '/' in data.quote_date:
                                                    quote_dt = datetime.strptime(data.quote_date, '%m/%d/%Y')
                                                elif '-' in data.quote_date:
                                                    quote_dt = datetime.strptime(data.quote_date, '%Y-%m-%d')
                                                else:
                                                    quote_dt = datetime.now()
                                                
                                                delta = exp_dt - quote_dt
                                                data.quote_expiration_days = delta.days
                                        except:
                                            data.quote_expiration_days = 30
                                        break
                        
                        # Look for customer company
                        elif 'company' in cell_value.lower() or 'customer' in cell_value.lower():
                            for offset in range(1, 4):
                                if cell.column + offset <= ws.max_column:
                                    next_cell = ws.cell(row=cell.row, column=cell.column + offset)
                                    if next_cell.value and not str(next_cell.value).lower().startswith(('company', 'customer')):
                                        data.customer_company = str(next_cell.value)
                                        break
            
            # If no quote number found, generate one
            if not data.quote_number:
                data.quote_number = "EXCEL_QUOTE_001"
            
            # If no quote date found, use current date
            if not data.quote_date:
                data.quote_date = datetime.now().strftime("%m/%d/%Y")
            
            # If no customer found, use default
            if not data.customer_company:
                data.customer_company = "Customer from Excel"
            
            # Extract line items from data rows
            line_items = []
            for row in ws.iter_rows(min_row=10):  # Start from row 10 to skip headers
                row_values = [cell.value for cell in row if cell.value is not None]
                if len(row_values) >= 4:  # Ensure we have enough columns
                    try:
                        # Try to extract numeric values for prices
                        unit_price = 0.0
                        total_price = 0.0
                        
                        for value in row_values[1:]:  # Skip first column (description)
                            if isinstance(value, (int, float)) and value > 0:
                                if unit_price == 0:
                                    unit_price = float(value)
                                else:
                                    total_price = float(value)
                                    break
                        
                        if unit_price > 0:
                            line_item = {
                                'item_number': str(len(line_items) + 1),
                                'quantity': 1,
                                'unit': 'EA',
                                'description': str(row_values[0]) if row_values else f"Item {len(line_items) + 1}",
                                'unit_price': unit_price,
                                'total_price': total_price if total_price > 0 else unit_price
                            }
                            line_items.append(line_item)
                    except:
                        continue
            
            data.line_items = line_items
            data.subtotal = sum(item['total_price'] for item in line_items)
            data.tax = data.subtotal * 0.05
            data.total = data.subtotal + data.tax
            
            # Calculate expiration date if not found
            if not data.quote_expiration_date:
                data.calculate_expiration_date()
            
        except Exception as e:
            # Fallback to pandas if openpyxl fails
            try:
                df = pd.read_excel(file_path)
            except:
                df = pd.read_excel(file_path, engine='xlrd')
            
            # Basic extraction
            data.quote_number = "EXCEL_QUOTE_001"
            data.quote_date = datetime.now().strftime("%m/%d/%Y")
            data.customer_company = "Customer from Excel"
            
            # Convert DataFrame to line items
            line_items = []
            for index, row in df.iterrows():
                if len(row) >= 4:
                    line_item = {
                        'item_number': str(index + 1),
                        'quantity': 1,
                        'unit': 'EA',
                        'description': str(row.iloc[0]) if pd.notna(row.iloc[0]) else f"Item {index + 1}",
                        'unit_price': float(row.iloc[1]) if pd.notna(row.iloc[1]) and str(row.iloc[1]).replace('.', '').isdigit() else 0.0,
                        'total_price': float(row.iloc[1]) if pd.notna(row.iloc[1]) and str(row.iloc[1]).replace('.', '').isdigit() else 0.0
                    }
                    line_items.append(line_item)
            
            data.line_items = line_items
            data.subtotal = sum(item['total_price'] for item in line_items)
            data.tax = data.subtotal * 0.05
            data.total = data.subtotal + data.tax
            data.calculate_expiration_date()
        
        return data

class QuoteGenerator:
    """Main quote generation class"""
    
    def __init__(self):
        self.parsers = {
            'xml': XMLParser(),
            'rtf': RTFParser(),
            'excel': ExcelParser()
        }
    
    def parse_file(self, file_path: str) -> QuoteData:
        """Parse input file and return quote data"""
        file_type = FileParser.detect_file_type(file_path)
        parser = self.parsers[file_type]
        return parser.parse(file_path)
    
    def generate_quote(self, data: QuoteData, template_path: str = "quote_template_simple.xlsx", output_path: str = None) -> str:
        """Generate quote from parsed data"""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"quote_{data.quote_number}_{timestamp}.xlsx"
        
        # Calculate expiration date before generating quote
        data.calculate_expiration_date()
        
        # Load template
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Fill in quote data
        self._fill_quote_data(ws, data)
        
        # Save the quote
        wb.save(output_path)
        return output_path
    
    def _fill_quote_data(self, ws, data: QuoteData):
        """Fill worksheet with quote data"""
        
        # Quote details - check if cells are merged first
        if data.quote_number:
            cell = ws['E2']
            if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = data.quote_number
        if data.quote_date:
            cell = ws['E3']
            if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = data.quote_date
        if data.customer_reference:
            cell = ws['E4']
            if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = data.customer_reference
        if data.customer_number:
            cell = ws['E5']
            if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = data.customer_number
        
        # Customer information
        if data.customer_company:
            cell = ws['B6']
            if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = data.customer_company
        if data.customer_contact:
            cell = ws['B7']
            if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = data.customer_contact
        if data.customer_phone:
            cell = ws['B8']
            if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = data.customer_phone
        if data.customer_email:
            cell = ws['B9']
            if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = data.customer_email
        
        # Lead time and expiration information
        # Add lead time information (we'll place it in a suitable location)
        if hasattr(data, 'lead_time_value') and data.lead_time_value > 0:
            # Try to find a suitable cell for lead time (e.g., near delivery terms)
            lead_time_cell = ws['B10']  # Adjust row as needed
            if hasattr(lead_time_cell, 'value') and not isinstance(lead_time_cell, openpyxl.cell.cell.MergedCell):
                lead_time_cell.value = f"Estimated Lead Time: {data.get_lead_time_display()}"
        
        # Add quote expiration information
        if hasattr(data, 'quote_expiration_date') and data.quote_expiration_date:
            # Try to find a suitable cell for expiration date
            expiration_cell = ws['B11']  # Adjust row as needed
            if hasattr(expiration_cell, 'value') and not isinstance(expiration_cell, openpyxl.cell.cell.MergedCell):
                expiration_cell.value = f"Quote Valid Until: {data.quote_expiration_date}"
        
        # Clear existing line items (rows 10-12)
        for row in range(10, 13):
            for col in range(1, 7):
                ws.cell(row=row, column=col).value = ""
        
        # Add line items
        start_row = 10
        for i, item in enumerate(data.line_items):
            row = start_row + i
            
            # Ensure we don't exceed reasonable row limits
            if row > 50:  # Reasonable limit
                break
            
            ws.cell(row=row, column=1).value = item.get('item_number', i + 1)
            ws.cell(row=row, column=2).value = item.get('quantity', 1)
            ws.cell(row=row, column=3).value = item.get('unit', 'EA')
            
            # Description with model number if available
            description = item.get('description', '')
            if 'order_code' in item:
                description += f"\nModel: {item['order_code']}"
            if 'material_number' in item:
                description += f"\nMaterial: {item['material_number']}"
            
            ws.cell(row=row, column=4).value = description
            ws.cell(row=row, column=5).value = item.get('unit_price', 0.0)
            ws.cell(row=row, column=6).value = item.get('total_price', 0.0)
            
            # Format currency
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(row=row, column=6).number_format = '$#,##0.00'
        
        # Update totals
        subtotal_row = 13
        tax_row = 14
        freight_row = 15
        total_row = 16
        
        # Calculate subtotal from line items
        subtotal = sum(item.get('total_price', 0.0) for item in data.line_items)
        
        # Check if cells are merged before setting values
        subtotal_cell = ws.cell(row=subtotal_row, column=6)
        if not isinstance(subtotal_cell, openpyxl.cell.cell.MergedCell):
            subtotal_cell.value = f"=SUM(F{start_row}:F{start_row + len(data.line_items) - 1})"
            subtotal_cell.number_format = '$#,##0.00'
        
        # Tax
        tax = data.tax if data.tax > 0 else subtotal * 0.05
        tax_cell = ws.cell(row=tax_row, column=6)
        if not isinstance(tax_cell, openpyxl.cell.cell.MergedCell):
            tax_cell.value = f"=F{subtotal_row}*0.05"
            tax_cell.number_format = '$#,##0.00'
        
        # Freight
        freight = data.freight if data.freight > 0 else 0.0
        freight_cell = ws.cell(row=freight_row, column=6)
        if not isinstance(freight_cell, openpyxl.cell.cell.MergedCell):
            freight_cell.value = freight
            freight_cell.number_format = '$#,##0.00'
        
        # Total
        total = subtotal + tax + freight
        total_cell = ws.cell(row=total_row, column=6)
        if not isinstance(total_cell, openpyxl.cell.cell.MergedCell):
            total_cell.value = f"=F{subtotal_row}+F{tax_row}+F{freight_row}"
            total_cell.number_format = '$#,##0.00'
    
    def generate_pdf(self, excel_path: str, output_path: str = None) -> str:
        """Generate PDF from Excel file (placeholder - would need additional implementation)"""
        if output_path is None:
            output_path = excel_path.replace('.xlsx', '.pdf')
        
        # This would require additional PDF generation library
        # For now, return the Excel path
        print(f"PDF generation not implemented. Excel file saved as: {excel_path}")
        return excel_path

def main():
    """Main function for testing"""
    generator = QuoteGenerator()
    
    # Test with XML file
    try:
        print("Parsing XML file...")
        data = generator.parse_file("ehOnline-Shop_20250905-160419.xml")
        print(f"Parsed quote data: {data.quote_number}")
        
        print("Generating quote...")
        output_path = generator.generate_quote(data)
        print(f"Quote generated: {output_path}")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
