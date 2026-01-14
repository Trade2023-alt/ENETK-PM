"""
Export modules for the Desktop Quote Generator
Handles Excel and PDF quote generation
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime
import os
import base64

class ExcelExporter:
    """Excel quote exporter with ENETK/EH branding"""
    
    def __init__(self):
        self.header_color = '8B0000'  # ENETK maroon
        self.light_color = 'F8F9FA'  # Very light gray
        self.accent_color = '6C757D'  # Medium gray
        self.text_color = '495057'   # Dark gray text
        self.border_color = 'DEE2E6'  # Light border
        self.highlight_color = 'E9ECEF'  # Highlight background
    
    def _soft_breaks(self, s: str) -> str:
        """Add soft break hints for better text wrapping"""
        if not s:
            return s
        # Allow wrapping after common separators
        for ch in ['/', '-', '_', ':', '(', ')', '+', '=']:
            s = s.replace(ch, ch + '\u200b')  # Zero-width space
        return s
    
    def export_quote(self, quote_data, output_path):
        """Export quote to Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quote"
        
        # Set up styles
        self.setup_styles()
        
        # Create the quote
        self.create_header(ws, quote_data)
        self.create_company_info(ws, quote_data)
        self.create_quote_details(ws, quote_data)
        self.create_customer_info(ws, quote_data)
        self.create_line_items(ws, quote_data)
        self.create_totals(ws, quote_data)
        self.create_terms(ws, quote_data)
        
        # Format the worksheet
        self.format_worksheet(ws)
        
        # Save the file
        wb.save(output_path)
        return output_path
    
    def setup_styles(self):
        """Setup Excel styles"""
        self.header_font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
        self.subheader_font = Font(name='Arial', size=12, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        self.bold_font = Font(name='Arial', size=10, bold=True)
        
        self.header_fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
        self.light_fill = PatternFill(start_color=self.light_color, end_color=self.light_color, fill_type='solid')
        self.accent_fill = PatternFill(start_color=self.accent_color, end_color=self.accent_color, fill_type='solid')
        
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.thick_border = Border(
            left=Side(style='thick'), right=Side(style='thick'),
            top=Side(style='thick'), bottom=Side(style='thick')
        )
    
    def create_header(self, ws, quote_data):
        """Create quote header"""
        # ENETK LLC branding with enhanced styling
        ws['A1'] = 'ENETK LLC'
        ws['A1'].font = Font(name='Arial', size=18, bold=True, color=self.header_color)
        ws['A1'].fill = PatternFill(start_color=self.light_color, end_color=self.light_color, fill_type='solid')
        ws['A1'].border = Border(
            left=Side(style='thick', color=self.header_color),
            right=Side(style='thick', color=self.header_color),
            top=Side(style='thick', color=self.header_color),
            bottom=Side(style='thick', color=self.header_color)
        )
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 35
        
        ws['A2'] = 'PLC AUTOMATION & INTEGRATION'
        ws['A2'].font = Font(name='Arial', size=11, bold=True, color=self.accent_color)
        ws['A2'].fill = PatternFill(start_color=self.light_color, end_color=self.light_color, fill_type='solid')
        ws.merge_cells('A2:F2')
        ws.row_dimensions[2].height = 25
        
        ws['A3'] = '11085 32E ST SW'
        ws['A3'].font = Font(name='Arial', size=10, color=self.text_color)
        ws['A3'].fill = PatternFill(start_color=self.light_color, end_color=self.light_color, fill_type='solid')
        ws.merge_cells('A3:F3')
        ws.row_dimensions[3].height = 20
        
        ws['A4'] = 'DICKINSON ND 58601-7810'
        ws['A4'].font = Font(name='Arial', size=10, color=self.text_color)
        ws['A4'].fill = PatternFill(start_color=self.light_color, end_color=self.light_color, fill_type='solid')
        ws.merge_cells('A4:F4')
        ws.row_dimensions[4].height = 20
        
        # Spacer
        ws.row_dimensions[5].height = 10
        
        # Main title with enhanced styling
        ws['A6'] = 'QUOTE'
        ws['A6'].font = Font(name='Arial', size=24, bold=True, color='FFFFFF')
        ws['A6'].fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
        ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A6'].border = Border(
            left=Side(style='thick', color=self.header_color),
            right=Side(style='thick', color=self.header_color),
            top=Side(style='thick', color=self.header_color),
            bottom=Side(style='thick', color=self.header_color)
        )
        ws.merge_cells('A6:F6')
        ws.row_dimensions[6].height = 45
    
    def create_company_info(self, ws, quote_data):
        """Create company information section"""
        # This is now handled in create_header
        pass
    
    def create_quote_details(self, ws, quote_data):
        """Create quote details section"""
        # Quote information with enhanced styling
        quote_details = [
            ('Quote Number:', quote_data.get('quote_number', '')),
            ('Quote Date:', quote_data.get('quote_date', '')),
            ('Project:', quote_data.get('project_name', '')),
            ('Customer Ref:', quote_data.get('customer_ref', ''))
        ]
        
        for i, (label, value) in enumerate(quote_details, 8):
            # Label styling
            ws[f'D{i}'] = label
            ws[f'D{i}'].font = Font(name='Arial', size=10, bold=True, color=self.text_color)
            ws[f'D{i}'].fill = PatternFill(start_color=self.highlight_color, end_color=self.highlight_color, fill_type='solid')
            ws[f'D{i}'].border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
            
            # Value styling
            ws[f'E{i}'] = value
            ws[f'E{i}'].font = Font(name='Arial', size=10, color=self.text_color)
            ws[f'E{i}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            ws[f'E{i}'].border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
    
    def create_customer_info(self, ws, quote_data):
        """Create customer information section"""
        start_row = 13
        
        # Customer Information header with enhanced styling
        ws[f'A{start_row}'] = 'CUSTOMER INFORMATION'
        ws[f'A{start_row}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{start_row}'].border = Border(
            left=Side(style='thick', color=self.header_color),
            right=Side(style='thick', color=self.header_color),
            top=Side(style='thick', color=self.header_color),
            bottom=Side(style='thick', color=self.header_color)
        )
        ws.merge_cells(f'A{start_row}:F{start_row}')
        ws.row_dimensions[start_row].height = 30
        
        # Customer details with enhanced styling
        customer_fields = [
            ('Company:', quote_data.get('customer_company', '')),
            ('Contact:', quote_data.get('contact_person', '')),
            ('Phone:', quote_data.get('phone', '')),
            ('Email:', quote_data.get('email', ''))
        ]
        
        for i, (label, value) in enumerate(customer_fields, start_row + 1):
            # Label styling
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(name='Arial', size=10, bold=True, color=self.text_color)
            ws[f'A{i}'].fill = PatternFill(start_color=self.highlight_color, end_color=self.highlight_color, fill_type='solid')
            ws[f'A{i}'].border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
            
            # Value styling
            ws[f'B{i}'] = value
            ws[f'B{i}'].font = Font(name='Arial', size=10, color=self.text_color)
            ws[f'B{i}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            ws[f'B{i}'].border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
        
        # Address information with enhanced styling
        address_row = start_row + 5
        ws[f'A{address_row}'] = 'Bill To:'
        ws[f'A{address_row}'].font = Font(name='Arial', size=10, bold=True, color=self.text_color)
        ws[f'A{address_row}'].fill = PatternFill(start_color=self.highlight_color, end_color=self.highlight_color, fill_type='solid')
        ws[f'A{address_row}'].border = Border(
            left=Side(style='thin', color=self.border_color),
            right=Side(style='thin', color=self.border_color),
            top=Side(style='thin', color=self.border_color),
            bottom=Side(style='thin', color=self.border_color)
        )
        ws[f'B{address_row}'] = quote_data.get('bill_to', '')
        ws[f'B{address_row}'].font = Font(name='Arial', size=10, color=self.text_color)
        ws[f'B{address_row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'B{address_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws[f'B{address_row}'].border = Border(
            left=Side(style='thin', color=self.border_color),
            right=Side(style='thin', color=self.border_color),
            top=Side(style='thin', color=self.border_color),
            bottom=Side(style='thin', color=self.border_color)
        )
        
        ws[f'D{address_row}'] = 'Ship To:'
        ws[f'D{address_row}'].font = Font(name='Arial', size=10, bold=True, color=self.text_color)
        ws[f'D{address_row}'].fill = PatternFill(start_color=self.highlight_color, end_color=self.highlight_color, fill_type='solid')
        ws[f'D{address_row}'].border = Border(
            left=Side(style='thin', color=self.border_color),
            right=Side(style='thin', color=self.border_color),
            top=Side(style='thin', color=self.border_color),
            bottom=Side(style='thin', color=self.border_color)
        )
        ws[f'E{address_row}'] = quote_data.get('ship_to', '')
        ws[f'E{address_row}'].font = Font(name='Arial', size=10, color=self.text_color)
        ws[f'E{address_row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'E{address_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws[f'E{address_row}'].border = Border(
            left=Side(style='thin', color=self.border_color),
            right=Side(style='thin', color=self.border_color),
            top=Side(style='thin', color=self.border_color),
            bottom=Side(style='thin', color=self.border_color)
        )
    
    def create_line_items(self, ws, quote_data):
        """Create line items table"""
        start_row = 20
        
        # Table headers with enhanced styling
        headers = ['Item #', 'Description', 'Qty', 'Unit', 'Unit Price', 'Total Price']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=i)
            cell.value = header
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thick', color=self.header_color),
                right=Side(style='thick', color=self.header_color),
                top=Side(style='thick', color=self.header_color),
                bottom=Side(style='thick', color=self.header_color)
            )
        ws.row_dimensions[start_row].height = 35
        
        # Line items
        markup = quote_data.get('markup_percentage', 20.0) / 100
        current_row = start_row + 1
        
        for i, item in enumerate(quote_data.get('line_items', []), 1):
            # Alternating row colors
            is_even_row = (i % 2 == 0)
            row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid') if not is_even_row else PatternFill(start_color=self.light_color, end_color=self.light_color, fill_type='solid')
            
            # Item number
            item_cell = ws.cell(row=current_row, column=1, value=i)
            item_cell.font = Font(name='Arial', size=10, bold=True, color=self.text_color)
            item_cell.fill = row_fill
            item_cell.alignment = Alignment(horizontal='center', vertical='center')
            item_cell.border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
            
            # Create clean, professional description formatting
            description_parts = []
            
            # Main product name and model
            product_name = item.get('description', '').split('\n')[0] if item.get('description') else ''
            if product_name:
                description_parts.append(product_name)
            
            # Add model number if available
            if item.get('model'):
                description_parts.append(f"Model: {item['model']}")
            
            # Sales text (product description)
            if item.get('sales_text'):
                description_parts.append(f"Description: {item['sales_text']}")
            
            # Delivery time
            if item.get('delivery_time'):
                description_parts.append(f"Delivery time: {item['delivery_time']}")
            
            # Order code description
            if item.get('order_code'):
                description_parts.append("Order code description:")
                description_parts.append(item['order_code'])
            
            # Product configuration (consolidated with smaller font)
            if item.get('config'):
                description_parts.append("Product Configuration:")
                # Consolidate configuration lines into single lines
                config_lines = item['config'].split('  ')
                consolidated_lines = []
                current_line = ""
                
                for line in config_lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # If line starts with a number (like "030:"), start a new consolidated line
                    if line and line[0].isdigit():
                        if current_line:
                            consolidated_lines.append(current_line)
                        current_line = line
                    else:
                        # Append to current line with a space
                        if current_line:
                            current_line += f" {line}"
                        else:
                            current_line = line
                
                # Add the last line
                if current_line:
                    consolidated_lines.append(current_line)
                
                # Add consolidated lines (smaller font handled by CSS)
                for line in consolidated_lines:
                    description_parts.append(line)
            
            # Country information
            country_info = []
            if item.get('country_origin'):
                country_info.append(f"Origin: {item['country_origin']}")
            if item.get('country_dispatch'):
                country_info.append(f"Dispatch: {item['country_dispatch']}")
            if country_info:
                description_parts.append("Country Information:")
                description_parts.extend(country_info)
            
            description = "\n\n".join(description_parts)
            
            # Clean up the description for better formatting
            description = description.replace('\r\n', '\n').replace('\r', '\n')
            
            # Add soft breaks for better text wrapping
            description = self._soft_breaks(description)
            
            desc_cell = ws.cell(row=current_row, column=2, value=description)
            desc_cell.font = Font(name='Arial', size=9, color=self.text_color)
            desc_cell.fill = row_fill
            desc_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', shrink_to_fit=False)
            desc_cell.border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
            
            # Quantity and unit
            qty_cell = ws.cell(row=current_row, column=3, value=item.get('quantity', 1))
            qty_cell.font = Font(name='Arial', size=10, color=self.text_color)
            qty_cell.fill = row_fill
            qty_cell.alignment = Alignment(horizontal='center', vertical='center')
            qty_cell.border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
            
            unit_cell = ws.cell(row=current_row, column=4, value=item.get('unit', 'EA'))
            unit_cell.font = Font(name='Arial', size=10, color=self.text_color)
            unit_cell.fill = row_fill
            unit_cell.alignment = Alignment(horizontal='center', vertical='center')
            unit_cell.border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
            
            # Prices
            TAX_RATE = 0.08
            unit_price = float(item.get('unit_price', 0))
            price_with_tax = unit_price * (1 + TAX_RATE)
            quoted_price = price_with_tax * (1 + markup)
            total_price = quoted_price * float(item.get('quantity', 1))
            
            unit_price_cell = ws.cell(row=current_row, column=5, value=quoted_price)
            unit_price_cell.font = Font(name='Arial', size=10, bold=True, color=self.text_color)
            unit_price_cell.fill = row_fill
            unit_price_cell.alignment = Alignment(horizontal='right', vertical='center')
            unit_price_cell.number_format = '$#,##0.00'
            unit_price_cell.border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
            
            total_price_cell = ws.cell(row=current_row, column=6, value=total_price)
            total_price_cell.font = Font(name='Arial', size=10, bold=True, color=self.text_color)
            total_price_cell.fill = row_fill
            total_price_cell.alignment = Alignment(horizontal='right', vertical='center')
            total_price_cell.number_format = '$#,##0.00'
            total_price_cell.border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
            
            # Calculate row height based on content length and line breaks
            description_length = len(description)
            line_breaks = description.count('\n') + description.count('\r')
            
            # Calculate estimated lines based on content and line breaks
            if line_breaks > 0:
                # If there are explicit line breaks, use them as a guide
                estimated_lines = max(3, line_breaks + 1)
            else:
                # Estimate based on content length
                estimated_lines = max(3, description_length // 60)  # More conservative estimate
            
            # Set row height with better calculation
            row_height = max(80, estimated_lines * 25)  # Increased base height and line height
            
            ws.row_dimensions[current_row].height = row_height
            current_row += 1
        
        return current_row
    
    def create_totals(self, ws, quote_data):
        """Create totals section"""
        line_items = quote_data.get('line_items', [])
        markup = quote_data.get('markup_percentage', 20.0) / 100
        
        # Calculate totals
        TAX_RATE = 0.08
        subtotal = 0
        for item in line_items:
            unit_price = float(item.get('unit_price', 0))
            quantity = float(item.get('quantity', 1))
            price_with_tax = unit_price * (1 + TAX_RATE)
            quoted_price = price_with_tax * (1 + markup)
            subtotal += quoted_price * quantity
        
        total = subtotal
        
        # Find the last row with data
        last_row = 20 + len(line_items) + 2
        
        # Totals with enhanced styling
        ws[f'E{last_row}'] = 'Subtotal:'
        ws[f'E{last_row}'].font = Font(name='Arial', size=11, bold=True, color=self.text_color)
        ws[f'E{last_row}'].fill = PatternFill(start_color=self.highlight_color, end_color=self.highlight_color, fill_type='solid')
        ws[f'E{last_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'E{last_row}'].border = Border(
            left=Side(style='thick', color=self.header_color),
            right=Side(style='thick', color=self.header_color),
            top=Side(style='thick', color=self.header_color),
            bottom=Side(style='thick', color=self.header_color)
        )
        ws[f'F{last_row}'] = subtotal
        ws[f'F{last_row}'].font = Font(name='Arial', size=11, bold=True, color=self.text_color)
        ws[f'F{last_row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'F{last_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'F{last_row}'].number_format = '$#,##0.00'
        ws[f'F{last_row}'].border = Border(
            left=Side(style='thick', color=self.header_color),
            right=Side(style='thick', color=self.header_color),
            top=Side(style='thick', color=self.header_color),
            bottom=Side(style='thick', color=self.header_color)
        )
        
        ws[f'E{last_row + 1}'] = 'TOTAL:'
        ws[f'E{last_row + 1}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws[f'E{last_row + 1}'].fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
        ws[f'E{last_row + 1}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'E{last_row + 1}'].border = Border(
            left=Side(style='thick', color=self.header_color),
            right=Side(style='thick', color=self.header_color),
            top=Side(style='thick', color=self.header_color),
            bottom=Side(style='thick', color=self.header_color)
        )
        ws[f'F{last_row + 1}'] = total
        ws[f'F{last_row + 1}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws[f'F{last_row + 1}'].fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
        ws[f'F{last_row + 1}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'F{last_row + 1}'].number_format = '$#,##0.00'
        ws[f'F{last_row + 1}'].border = Border(
            left=Side(style='thick', color=self.header_color),
            right=Side(style='thick', color=self.header_color),
            top=Side(style='thick', color=self.header_color),
            bottom=Side(style='thick', color=self.header_color)
        )
        
        # Set row heights
        ws.row_dimensions[last_row].height = 25
        ws.row_dimensions[last_row + 1].height = 35
        
        # Add borders to totals
        for col in range(1, 7):
            for row in range(last_row, last_row + 3):
                ws.cell(row=row, column=col).border = self.thick_border
    
    def create_terms(self, ws, quote_data):
        """Create terms and conditions section"""
        last_row = 20 + len(quote_data.get('line_items', [])) + 6
        
        # Terms header with enhanced styling
        ws[f'A{last_row}'] = 'TERMS AND CONDITIONS'
        ws[f'A{last_row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws[f'A{last_row}'].fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
        ws[f'A{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{last_row}'].border = Border(
            left=Side(style='thick', color=self.header_color),
            right=Side(style='thick', color=self.header_color),
            top=Side(style='thick', color=self.header_color),
            bottom=Side(style='thick', color=self.header_color)
        )
        ws.merge_cells(f'A{last_row}:F{last_row}')
        ws.row_dimensions[last_row].height = 35
        
        # Lead time and quote validity information
        lead_time = quote_data.get('lead_time_value', 0)
        lead_time_unit = quote_data.get('lead_time_unit', 'Days')
        expiration_date = quote_data.get('quote_expiration_date', '')
        
        # Comprehensive terms and conditions
        terms = []
        
        # Add lead time and quote validity information
        if lead_time > 0 or expiration_date:
            terms.extend([
                "LEAD TIME & QUOTE VALIDITY:",
                f"Estimated Lead Time: {lead_time} {lead_time_unit}" if lead_time > 0 else "",
                f"Quote Valid Until: {expiration_date}" if expiration_date else "",
                "",
                "TERMS AND CONDITIONS:"
            ])
        
        # Add standard terms
        terms.extend([
            "1. Acceptance. The Buyer's purchase order (\"Order\") is an offer to buy Goods and/or Services under these",
            "Terms and Conditions. The Agreement is formed when Seller provides written acceptance of the Order.",
            "Any Seller quotation is not an offer and is valid for 30 days from issuance unless notated differently on",
            "proposal.",
            "",
            "2. Buyer's Assent. Shipment by Seller and Buyer's acceptance or payment for any part of the Goods and/or",
            "Services constitutes Buyer's agreement to these Terms. The seller may withdraw or modify its conditional",
            "acceptance of the Order before Buyer accepts the Goods and Services.",
            "",
            "3. Modification. No modifications or changes are binding on Seller unless agreed in writing by Seller. Seller",
            "is not bound by any additional or different terms in Buyer's communications unless specifically accepted in",
            "writing. Prior dealings, trade practices, or verbal agreements not formally documented and signed by Seller",
            "are not binding. Any attempted modification or repudiation by Buyer, without Seller's consent, may be",
            "treated as a breach of the Agreement.",
            "",
            "4. Payment. The price of Goods is as stated in the Order Confirmation and excludes packaging, insurance,",
            "and transport. Service charges are based on time and materials per Seller's standard rates, available upon",
            "request. Prices for Goods and Services (collectively, \"Prices\") are subject to Seller's standard annual",
            "increases, with prior notice to Buyer. Payment is due in full within 30 days of shipment unless otherwise",
            "specified on Order, without set-off or withholding except as legally required. Unpaid balances incur a 1.5%",
            "monthly service charge. If Seller identifies a significant negative change in Buyer's financial condition or if",
            "payments are overdue, Seller may revoke credit, require C.O.D. terms, or suspend further performance.",
            "Seller retains a security interest in the Goods until full payment is received, and Buyer authorizes Seller to",
            "file any documents necessary to protect this security interest.",
            "",
            "5. Termination. If the Agreement is terminated by mutual consent and no other arrangement is specified, the",
            "Buyer shall pay termination charges equal to the greater of: : 10% of the net sales price, or : the sum of",
            ": the Order price for completed Goods and Services, : Seller's actual costs and liabilities for",
            "uncompleted portions, and : reasonable estimated profit on uncompleted portions",
            "",
            "6. Returns. Custom products are made specifically to the Buyer's specifications and may not be eligible for",
            "return. If a return is permitted at the Seller's discretion, a restocking fee of 15% of the purchase price will",
            "apply. Restocking Fees will vary from different manufacturers and vendors. All return requests for",
            "custom products must be pre-approved in writing by the Seller, and returned items must be in their original",
            "condition.",
            "",
            "7. Delivery / Shipment. Seller will deliver Goods to the location specified in the Order Confirmation. Buyer",
            "is responsible for shipping, packing, and handling costs, and assumes risk of loss once Goods are",
            "transferred to the carrier. Buyer has five (5) days after delivery to report any discrepancies with the",
            "itemized packing list.",
            "",
            "8. Storage Fees. If the Buyer delays shipment or fails to provide required information, causing Goods to be",
            "stored, storage fees will apply starting from the scheduled shipment date unless otherwise extended in",
            "writing. Buyers will be invoiced separately for storage at Seller's actual cost, with payment due within 30",
            "days of the invoice date.",
            "",
            "9. Warranties. All warranties on products will defer to the manufacturer's warranty. The Seller provides no",
            "additional warranties, and any warranty claims should coordinate with the product's manufacturer.",
            "",
            "10. Sales or Use Taxes. Seller will apply sales tax to taxable orders unless Buyer provides valid and legally",
            "acceptable documentation, such as a tax-exempt certificate or direct pay permit, as required by the",
            "jurisdiction where delivery or service occurs. Buyer is responsible for any additional sales tax if a",
            "subsequent audit finds that the provided documentation was incomplete or invalid, unless Buyer corrects",
            "the deficiency within 30 days of notification."
        ])
        
        for i, term in enumerate(terms, last_row + 1):
            ws[f'A{i}'] = term
            ws[f'A{i}'].font = Font(name='Arial', size=9, color=self.text_color)
            ws[f'A{i}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            ws[f'A{i}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            ws[f'A{i}'].border = Border(
                left=Side(style='thin', color=self.border_color),
                right=Side(style='thin', color=self.border_color),
                top=Side(style='thin', color=self.border_color),
                bottom=Side(style='thin', color=self.border_color)
            )
            # Merge across all columns for better readability
            ws.merge_cells(f'A{i}:F{i}')
            ws.row_dimensions[i].height = 20
    
    def format_worksheet(self, ws):
        """Format the entire worksheet"""
        # Set column widths for better layout
        column_widths = {
            'A': 8,   # Item #
            'B': 80,  # Description (even wider for full product descriptions)
            'C': 8,   # Qty
            'D': 8,   # Unit
            'E': 15,  # Unit Price
            'F': 15   # Total Price
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Ensure text wrapping is enabled for the entire worksheet
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.alignment is None:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    # Update existing alignment to ensure wrap_text is True
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical=cell.alignment.vertical or 'top',
                        horizontal=cell.alignment.horizontal or 'left'
                    )
        
        # Only set row heights if none have been set already
        # This preserves the dynamic heights calculated in create_line_items
        for row in range(1, ws.max_row + 1):
            if ws.row_dimensions[row].height is None:
                if row == 1:  # Header row
                    ws.row_dimensions[row].height = 40
                else:
                    ws.row_dimensions[row].height = 20  # default

class PDFExporter:
    """PDF quote exporter with ENETK/EH branding"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
    
    def setup_custom_styles(self):
        """Setup custom PDF styles"""
        # Title style
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=28,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#8B0000')
        )
        
        # Header style
        self.header_style = ParagraphStyle(
            'CustomHeader',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=15,
            textColor=colors.HexColor('#8B0000')
        )
        
        # Subheader style
        self.subheader_style = ParagraphStyle(
            'CustomSubHeader',
            parent=self.styles['Heading3'],
            fontSize=12,
            spaceAfter=10,
            textColor=colors.HexColor('#6C757D')
        )
        
        # Normal style
        self.normal_style = ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            textColor=colors.HexColor('#495057')
        )
        
        # Description style for better text wrapping
        self.desc_style = ParagraphStyle(
            'Desc',
            parent=self.styles['Normal'],
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
            wordWrap='CJK',  # allows breaking inside long tokens
            spaceAfter=6,
            spaceBefore=6
        )
    
    def _sanitize_for_pdf(self, s: str) -> str:
        """Sanitize text for PDF output by removing problematic characters."""
        if not s:
            return s
        
        import re
        
        # Replace problematic characters that cause black squares
        repl = {
            '\u200b': '',      # zero-width space
            '\ufeff': '',      # BOM
            '\u00a0': ' ',     # NBSP -> space
            '\u25a0': '',      # ■ black square -> remove
            '\u2022': '* ',    # • bullet -> ASCII bullet
            '\u2018': "'", '\u2019': "'",  # curly quotes -> straight
            '\u201c': '"', '\u201d': '"',
            '\u2013': '-', '\u2014': '-', '\u2212': '-',  # dashes -> hyphen
            '\u2026': '...',    # ellipsis
            '\u2122': ' TM', '\u00ae': ' (R)',
            '°': ' degrees',   # degree symbol
        }
        
        for k, v in repl.items():
            s = s.replace(k, v)
        
        # Drop anything still non-ASCII (safe fallback)
        s = re.sub(r'[^\x20-\x7E\n\r\t]', '', s)
        
        return s
    
    def export_quote(self, quote_data, output_path):
        """Export quote to PDF file"""
        doc = SimpleDocTemplate(output_path, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        story = []
        
        # Create the quote content
        self.create_header(story, quote_data)
        self.create_company_info(story, quote_data)
        self.create_quote_details(story, quote_data)
        self.create_customer_info(story, quote_data)
        self.create_line_items(story, quote_data, doc.width)  # Pass available width
        self.create_totals(story, quote_data)
        self.create_terms(story, quote_data)
        
        # Build PDF
        doc.build(story)
        return output_path
    
    def create_header(self, story, quote_data):
        """Create PDF header"""
        # ENETK LLC branding
        story.append(Paragraph("ENETK LLC", self.title_style))
        story.append(Paragraph("PLC AUTOMATION & INTEGRATION", 
                              ParagraphStyle('CompanyTagline', parent=self.styles['Normal'], 
                                           fontSize=12, textColor=colors.HexColor('#808080'))))
        story.append(Paragraph("11085 32E ST SW<br/>DICKINSON ND 58601-7810", 
                              ParagraphStyle('Address', parent=self.styles['Normal'], 
                                           fontSize=10, textColor=colors.HexColor('#808080'))))
        story.append(Spacer(1, 20))
        story.append(Paragraph("QUOTE", self.title_style))
        story.append(Spacer(1, 20))
    
    def create_company_info(self, story, quote_data):
        """Create company information"""
        # Company contact information removed as requested
        story.append(Spacer(1, 20))
    
    def create_quote_details(self, story, quote_data):
        """Create quote details"""
        details_data = [
            ['Quote Number:', quote_data.get('quote_number', '')],
            ['Quote Date:', quote_data.get('quote_date', '')],
            ['Project:', quote_data.get('project_name', '')],
            ['Customer Ref:', quote_data.get('customer_ref', '')]
        ]
        
        details_table = Table(details_data, colWidths=[2*inch, 4*inch])
        details_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(details_table)
        story.append(Spacer(1, 20))
    
    def create_customer_info(self, story, quote_data):
        """Create customer information"""
        story.append(Paragraph("CUSTOMER INFORMATION", self.header_style))
        
        customer_data = [
            ['Company:', quote_data.get('customer_company', '')],
            ['Contact:', quote_data.get('contact_person', '')],
            ['Phone:', quote_data.get('phone', '')],
            ['Email:', quote_data.get('email', '')]
        ]
        
        customer_table = Table(customer_data, colWidths=[1.5*inch, 4.5*inch])
        customer_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(customer_table)
        story.append(Spacer(1, 20))
    
    def create_line_items(self, story, quote_data, available_width):
        """Create line items table with proper width calculation and text wrapping"""
        story.append(Paragraph("LINE ITEMS", self.header_style))
        
        # Table headers
        headers = ['Item #', 'Description', 'Qty', 'Unit', 'Unit Price', 'Total Price']
        
        # Calculate column widths to fit available page width
        from reportlab.lib.units import inch
        item_w = 0.5 * inch
        qty_w = 0.5 * inch
        unit_w = 0.5 * inch
        uprice_w = 0.9 * inch
        tprice_w = 1.0 * inch
        
        other_total = item_w + qty_w + unit_w + uprice_w + tprice_w
        desc_w = max(2.5 * inch, available_width - other_total)  # remainder fits the page
        
        col_widths = [item_w, desc_w, qty_w, unit_w, uprice_w, tprice_w]
        
        # Prepare data
        table_data = [headers]
        markup = quote_data.get('markup_percentage', 20.0) / 100
        
        for i, item in enumerate(quote_data.get('line_items', []), 1):
            # Create clean, professional description formatting
            description_parts = []
            
            # Main product name and model
            product_name = item.get('description', '').split('\n')[0] if item.get('description') else ''
            if product_name:
                description_parts.append(product_name)
            
            # Add model number if available
            if item.get('model'):
                description_parts.append(f"Model: {item['model']}")
            
            # Sales text (product description)
            if item.get('sales_text'):
                description_parts.append(f"Description: {item['sales_text']}")
            
            # Delivery time
            if item.get('delivery_time'):
                description_parts.append(f"Delivery time: {item['delivery_time']}")
            
            # Order code description
            if item.get('order_code'):
                description_parts.append("Order code description:")
                description_parts.append(item['order_code'])
            
            # Product configuration (consolidated with smaller font)
            if item.get('config'):
                description_parts.append("Product Configuration:")
                # Consolidate configuration lines into single lines
                config_lines = item['config'].split('  ')
                consolidated_lines = []
                current_line = ""
                
                for line in config_lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # If line starts with a number (like "030:"), start a new consolidated line
                    if line and line[0].isdigit():
                        if current_line:
                            consolidated_lines.append(current_line)
                        current_line = line
                    else:
                        # Append to current line with a space
                        if current_line:
                            current_line += f" {line}"
                        else:
                            current_line = line
                
                # Add the last line
                if current_line:
                    consolidated_lines.append(current_line)
                
                # Add consolidated lines (smaller font handled by CSS)
                for line in consolidated_lines:
                    description_parts.append(line)
            
            # Country information
            country_info = []
            if item.get('country_origin'):
                country_info.append(f"Origin: {item['country_origin']}")
            if item.get('country_dispatch'):
                country_info.append(f"Dispatch: {item['country_dispatch']}")
            if country_info:
                description_parts.append("Country Information:")
                description_parts.extend(country_info)
            
            # Sanitize text and create HTML
            from xml.sax.saxutils import escape
            parts = [self._sanitize_for_pdf(p) for p in description_parts if p]
            desc_html = "<br/><br/>".join(escape(p) for p in parts)
            
            # Create Paragraph object for proper text wrapping
            desc_para = Paragraph(desc_html, self.desc_style)
            
            # Prices
            TAX_RATE = 0.08
            unit_price = float(item.get('unit_price', 0))
            price_with_tax = unit_price * (1 + TAX_RATE)
            quoted_price = price_with_tax * (1 + markup)
            total_price = quoted_price * float(item.get('quantity', 1))
            
            row_data = [
                str(i),
                desc_para,  # Use Paragraph object instead of plain string
                str(item.get('quantity', 1)),
                item.get('unit', 'EA'),
                f"${quoted_price:.2f}",
                f"${total_price:.2f}"
            ]
            table_data.append(row_data)
        
        # Create table with calculated column widths and better row splitting
        items_table = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=1, splitInRow=1)
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),  # ENETK maroon
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),  # Smaller font for better fit
            
            # Alignment per column
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),   # Item #
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),     # Description
            ('ALIGN', (2, 1), (3, -1), 'CENTER'),   # Qty, Unit
            ('ALIGN', (4, 1), (5, -1), 'RIGHT'),    # Prices
            
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(items_table)
        story.append(Spacer(1, 20))
    
    def create_totals(self, story, quote_data):
        """Create totals section"""
        line_items = quote_data.get('line_items', [])
        markup = quote_data.get('markup_percentage', 20.0) / 100
        tax_rate = quote_data.get('tax_percentage', 8.0) / 100
        
        # Calculate totals
        TAX_RATE = 0.08
        subtotal = 0
        for item in line_items:
            unit_price = float(item.get('unit_price', 0))
            quantity = float(item.get('quantity', 1))
            price_with_tax = unit_price * (1 + TAX_RATE)
            quoted_price = price_with_tax * (1 + markup)
            subtotal += quoted_price * quantity
        
        total = subtotal
        
        # Totals table
        totals_data = [
            ['Subtotal:', f"${subtotal:.2f}"],
            ['TOTAL:', f"${total:.2f}"]
        ]
        
        totals_table = Table(totals_data, colWidths=[4*inch, 2*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('FONTSIZE', (0, 1), (-1, 1), 14),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F2F2F2')),
        ]))
        
        story.append(totals_table)
        story.append(Spacer(1, 20))
    
    def create_terms(self, story, quote_data):
        """Create terms and conditions"""
        # Add lead time and quote validity information
        lead_time = quote_data.get('lead_time_value', 0)
        lead_time_unit = quote_data.get('lead_time_unit', 'Days')
        expiration_date = quote_data.get('quote_expiration_date', '')
        
        if lead_time > 0 or expiration_date:
            story.append(Paragraph("LEAD TIME & QUOTE VALIDITY", self.header_style))
            
            validity_text = ""
            if lead_time > 0:
                validity_text += f"<b>Estimated Lead Time:</b> {lead_time} {lead_time_unit}<br/>"
            if expiration_date:
                validity_text += f"<b>Quote Valid Until:</b> {expiration_date}<br/>"
            
            if validity_text:
                story.append(Paragraph(validity_text, self.normal_style))
                story.append(Spacer(1, 10))
        
        story.append(Paragraph("TERMS AND CONDITIONS", self.header_style))
        
        terms_text = """
        <b>1. Acceptance.</b> The Buyer's purchase order ("Order") is an offer to buy Goods and/or Services under these
        Terms and Conditions. The Agreement is formed when Seller provides written acceptance of the Order.
        Any Seller quotation is not an offer and is valid for 30 days from issuance unless notated differently on
        proposal.<br/><br/>
        
        <b>2. Buyer's Assent.</b> Shipment by Seller and Buyer's acceptance or payment for any part of the Goods and/or
        Services constitutes Buyer's agreement to these Terms. The seller may withdraw or modify its conditional
        acceptance of the Order before Buyer accepts the Goods and Services.<br/><br/>
        
        <b>3. Modification.</b> No modifications or changes are binding on Seller unless agreed in writing by Seller. Seller
        is not bound by any additional or different terms in Buyer's communications unless specifically accepted in
        writing. Prior dealings, trade practices, or verbal agreements not formally documented and signed by Seller
        are not binding. Any attempted modification or repudiation by Buyer, without Seller's consent, may be
        treated as a breach of the Agreement.<br/><br/>
        
        <b>4. Payment.</b> The price of Goods is as stated in the Order Confirmation and excludes packaging, insurance,
        and transport. Service charges are based on time and materials per Seller's standard rates, available upon
        request. Prices for Goods and Services (collectively, "Prices") are subject to Seller's standard annual
        increases, with prior notice to Buyer. Payment is due in full within 30 days of shipment unless otherwise
        specified on Order, without set-off or withholding except as legally required. Unpaid balances incur a 1.5%
        monthly service charge. If Seller identifies a significant negative change in Buyer's financial condition or if
        payments are overdue, Seller may revoke credit, require C.O.D. terms, or suspend further performance.
        Seller retains a security interest in the Goods until full payment is received, and Buyer authorizes Seller to
        file any documents necessary to protect this security interest.<br/><br/>
        
        <b>5. Termination.</b> If the Agreement is terminated by mutual consent and no other arrangement is specified, the
        Buyer shall pay termination charges equal to the greater of: : 10% of the net sales price, or : the sum of
        : the Order price for completed Goods and Services, : Seller's actual costs and liabilities for
        uncompleted portions, and : reasonable estimated profit on uncompleted portions<br/><br/>
        
        <b>6. Returns.</b> Custom products are made specifically to the Buyer's specifications and may not be eligible for
        return. If a return is permitted at the Seller's discretion, a restocking fee of 15% of the purchase price will
        apply. Restocking Fees will vary from different manufacturers and vendors. All return requests for
        custom products must be pre-approved in writing by the Seller, and returned items must be in their original
        condition.<br/><br/>
        
        <b>7. Delivery / Shipment.</b> Seller will deliver Goods to the location specified in the Order Confirmation. Buyer
        is responsible for shipping, packing, and handling costs, and assumes risk of loss once Goods are
        transferred to the carrier. Buyer has five (5) days after delivery to report any discrepancies with the
        itemized packing list.<br/><br/>
        
        <b>8. Storage Fees.</b> If the Buyer delays shipment or fails to provide required information, causing Goods to be
        stored, storage fees will apply starting from the scheduled shipment date unless otherwise extended in
        writing. Buyers will be invoiced separately for storage at Seller's actual cost, with payment due within 30
        days of the invoice date.<br/><br/>
        
        <b>9. Warranties.</b> All warranties on products will defer to the manufacturer's warranty. The Seller provides no
        additional warranties, and any warranty claims should coordinate with the product's manufacturer.<br/><br/>
        
        <b>10. Sales or Use Taxes.</b> Seller will apply sales tax to taxable orders unless Buyer provides valid and legally
        acceptable documentation, such as a tax-exempt certificate or direct pay permit, as required by the
        jurisdiction where delivery or service occurs. Buyer is responsible for any additional sales tax if a
        subsequent audit finds that the provided documentation was incomplete or invalid, unless Buyer corrects
        the deficiency within 30 days of notification.
        """
        
        story.append(Paragraph(terms_text, self.normal_style))
