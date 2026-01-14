#!/usr/bin/env python3
"""
Create Simple Excel Quote Template
Creates a professional quote template without merged cells to avoid issues
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

def create_simple_quote_template():
    """Create a simple quote template Excel file without merged cells"""
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote Template"
    
    # Define styles
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    small_font = Font(name='Arial', size=9)
    
    # Colors
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 8   # Item #
    ws.column_dimensions['B'].width = 6   # Qty
    ws.column_dimensions['C'].width = 8   # Unit
    ws.column_dimensions['D'].width = 50  # Description
    ws.column_dimensions['E'].width = 15  # Unit Price
    ws.column_dimensions['F'].width = 15  # Total Price
    
    # Header Section
    ws['A1'] = 'QUOTE'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thick_border
    ws.row_dimensions[1].height = 30
    
    # Company Info Section (Row 2-4)
    ws['A2'] = 'KINDER MORGAN'
    ws['A2'].font = normal_font
    ws['A3'] = '1234 Energy Drive'
    ws['A3'].font = normal_font
    ws['A4'] = 'Houston, TX 77002'
    ws['A4'].font = normal_font
    ws['A5'] = 'Phone: (713) 369-9000'
    ws['A5'].font = normal_font
    
    # Quote Details Section (Row 2-5, right side)
    quote_details = [
        ('Quote No.:', 'QUOTE_NUMBER'),
        ('Date:', 'QUOTE_DATE'),
        ('Customer Ref:', 'CUSTOMER_REF'),
        ('Customer No.:', 'CUSTOMER_NUMBER')
    ]
    
    for i, (label, placeholder) in enumerate(quote_details, start=2):
        ws[f'D{i}'] = label
        ws[f'D{i}'].font = subheader_font
        ws[f'E{i}'] = placeholder
        ws[f'E{i}'].font = normal_font
        ws[f'E{i}'].alignment = Alignment(horizontal='left')
    
    # Customer Information Section (Row 6-9)
    ws['A6'] = 'CUSTOMER INFORMATION'
    ws['A6'].font = subheader_font
    ws['A6'].fill = light_fill
    ws['A6'].alignment = Alignment(horizontal='center')
    ws['A6'].border = thin_border
    
    customer_info = [
        ('Company:', 'CUSTOMER_COMPANY'),
        ('Contact:', 'CUSTOMER_CONTACT'),
        ('Phone:', 'CUSTOMER_PHONE'),
        ('Email:', 'CUSTOMER_EMAIL')
    ]
    
    for i, (label, placeholder) in enumerate(customer_info, start=7):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = subheader_font
        ws[f'B{i}'] = placeholder
        ws[f'B{i}'].font = normal_font
        ws[f'B{i}'].alignment = Alignment(horizontal='left')
    
    # Line Items Header (Row 10)
    headers = ['Item #', 'Qty', 'Unit', 'Description', 'Unit Price', 'Total Price']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = subheader_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thick_border
    
    # Sample line items (Row 11-13)
    sample_items = [
        ['1', '1', 'EA', 'Sample Product 1\nModel: ABC-123\nDescription details...', '$1,000.00', '$1,000.00'],
        ['2', '2', 'EA', 'Sample Product 2\nModel: XYZ-456\nDescription details...', '$500.00', '$1,000.00'],
        ['', '', '', '', '', '']
    ]
    
    for row_idx, item in enumerate(sample_items, start=11):
        for col_idx, value in enumerate(item, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx in [4, 5, 6]:  # Description, Unit Price, Total Price
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Totals Section (Row 14-17)
    totals_start_row = 14
    totals = [
        ('Subtotal:', '=SUM(F11:F12)'),
        ('Tax (5%):', '=F14*0.05'),
        ('Freight:', '0.00'),
        ('TOTAL:', '=F14+F15+F16')
    ]
    
    for i, (label, formula) in enumerate(totals, start=totals_start_row):
        ws[f'E{i}'] = label
        ws[f'E{i}'].font = subheader_font
        ws[f'E{i}'].alignment = Alignment(horizontal='right')
        
        ws[f'F{i}'] = formula
        ws[f'F{i}'].font = normal_font
        ws[f'F{i}'].alignment = Alignment(horizontal='right')
        ws[f'F{i}'].number_format = '$#,##0.00'
        
        if i == totals_start_row + 3:  # Total row
            ws[f'E{i}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'F{i}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'E{i}'].fill = light_fill
            ws[f'F{i}'].fill = light_fill
        
        # Add borders
        for col in ['E', 'F']:
            ws[f'{col}{i}'].border = thin_border
    
    # Terms and Conditions Section (Row 19-25)
    ws['A19'] = 'TERMS AND CONDITIONS'
    ws['A19'].font = subheader_font
    ws['A19'].fill = light_fill
    ws['A19'].alignment = Alignment(horizontal='center')
    ws['A19'].border = thin_border
    
    terms = [
        'Payment Terms: Net 30 days',
        'Delivery: FOB Origin',
        'Validity: 30 days from quote date',
        'All prices subject to change without notice',
        'Freight charges not included unless specified',
        'Lead time: 2-4 weeks standard delivery'
    ]
    
    for i, term in enumerate(terms, start=20):
        ws[f'A{i}'] = f'• {term}'
        ws[f'A{i}'].font = small_font
        ws[f'A{i}'].alignment = Alignment(horizontal='left')
    
    # Add borders to terms section
    for row in range(19, 26):
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = thin_border
    
    # Footer
    ws['A27'] = 'Thank you for your business!'
    ws['A27'].font = subheader_font
    ws['A27'].alignment = Alignment(horizontal='center')
    
    # Save the template
    wb.save('quote_template_simple.xlsx')
    print("Simple quote template created successfully: quote_template_simple.xlsx")

if __name__ == "__main__":
    create_simple_quote_template()
