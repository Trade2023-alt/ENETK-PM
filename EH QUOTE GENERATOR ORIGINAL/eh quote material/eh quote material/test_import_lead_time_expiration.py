#!/usr/bin/env python3
"""
Test script for lead time and quote expiration features with imported files
"""

from quote_generator import QuoteGenerator, QuoteData
from datetime import datetime
import os

def test_import_with_lead_time_expiration():
    """Test importing files and extracting lead time/expiration information"""
    print("🧪 Testing Lead Time and Quote Expiration with File Import")
    print("=" * 70)
    
    generator = QuoteGenerator()
    
    # Test with existing sample files
    sample_files = [
        "ehOnline-Shop_20250905-160419.xml",
        "ehOnline-Shop_2061348427.rtf", 
        "ehOnline-Shop_2061348427.xlsx"
    ]
    
    for file_path in sample_files:
        if os.path.exists(file_path):
            print(f"\n📁 Testing file: {file_path}")
            print("-" * 50)
            
            try:
                # Parse the file
                data = generator.parse_file(file_path)
                
                print(f"   Quote Number: {data.quote_number}")
                print(f"   Quote Date: {data.quote_date}")
                print(f"   Customer: {data.customer_company}")
                print(f"   Lead Time: {data.get_lead_time_display()}")
                print(f"   Quote Valid Until: {data.quote_expiration_date}")
                print(f"   Expiration Days: {data.quote_expiration_days}")
                print(f"   Line Items: {len(data.line_items)}")
                
                # Generate a quote with the imported data
                output_path = generator.generate_quote(data, output_path=f"imported_quote_{os.path.splitext(file_path)[0]}.xlsx")
                print(f"   ✅ Generated quote: {output_path}")
                
            except Exception as e:
                print(f"   ❌ Error processing {file_path}: {str(e)}")
        else:
            print(f"   ⚠️  File not found: {file_path}")
    
    print(f"\n🎯 Testing with custom Excel file containing lead time and expiration...")
    
    # Create a test Excel file with lead time and expiration information
    import openpyxl
    from openpyxl.styles import Font
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add test data with lead time and expiration
    ws['A1'] = 'Quote #:'
    ws['B1'] = 'TEST-IMPORT-001'
    ws['A2'] = 'Quote Date:'
    ws['B2'] = '01/15/2025'
    ws['A3'] = 'Company:'
    ws['B3'] = 'Test Import Company'
    ws['A4'] = 'Lead Time:'
    ws['B4'] = '21 days'
    ws['A5'] = 'Valid Until:'
    ws['B5'] = '02/14/2025'
    
    # Add some line items
    ws['A7'] = 'Item #'
    ws['B7'] = 'Description'
    ws['C7'] = 'Qty'
    ws['D7'] = 'Unit Price'
    ws['E7'] = 'Total Price'
    
    ws['A8'] = '1'
    ws['B8'] = 'Test Product A'
    ws['C8'] = '2'
    ws['D8'] = '1000.00'
    ws['E8'] = '2000.00'
    
    ws['A9'] = '2'
    ws['B9'] = 'Test Product B'
    ws['C9'] = '1'
    ws['D9'] = '500.00'
    ws['E9'] = '500.00'
    
    test_file = "test_import_with_lead_time.xlsx"
    wb.save(test_file)
    print(f"   📄 Created test file: {test_file}")
    
    # Test importing this file
    try:
        data = generator.parse_file(test_file)
        print(f"   Quote Number: {data.quote_number}")
        print(f"   Quote Date: {data.quote_date}")
        print(f"   Customer: {data.customer_company}")
        print(f"   Lead Time: {data.get_lead_time_display()}")
        print(f"   Quote Valid Until: {data.quote_expiration_date}")
        print(f"   Expiration Days: {data.quote_expiration_days}")
        print(f"   Line Items: {len(data.line_items)}")
        
        # Generate quote
        output_path = generator.generate_quote(data, output_path="imported_with_lead_time_quote.xlsx")
        print(f"   ✅ Generated quote: {output_path}")
        
    except Exception as e:
        print(f"   ❌ Error processing test file: {str(e)}")
    
    print(f"\n✅ Import testing completed!")
    print(f"   - Check the generated Excel files to see lead time and expiration information")
    print(f"   - The information should appear in the 'Lead Time & Quote Validity' section")

if __name__ == "__main__":
    test_import_with_lead_time_expiration()
