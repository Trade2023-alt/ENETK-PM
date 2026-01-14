#!/usr/bin/env python3
"""
Create Excel Macro Workbook
Creates an Excel file with VBA macros for quote generation
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

def create_excel_macro_workbook():
    """Create an Excel workbook with VBA macros for quote generation"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote Generator"
    
    # Define styles
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    # Colors
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    
    # Header
    ws['A1'] = 'QUOTE GENERATOR'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    
    # Instructions
    ws['A3'] = 'INSTRUCTIONS:'
    ws['A3'].font = subheader_font
    
    instructions = [
        '1. Fill in the quote details below',
        '2. Add line items in the table',
        '3. Click "Generate Quote" button',
        '4. A new worksheet will be created with your quote'
    ]
    
    for i, instruction in enumerate(instructions, start=4):
        ws[f'A{i}'] = instruction
        ws[f'A{i}'].font = normal_font
    
    # Quote Details Section
    ws['A7'] = 'QUOTE DETAILS'
    ws['A7'].font = subheader_font
    ws['A7'].fill = light_fill
    ws.merge_cells('A7:D7')
    
    # Quote form fields
    quote_fields = [
        ('Quote Number:', 'B8', 'QUOTE-001'),
        ('Quote Date:', 'B9', '=TODAY()'),
        ('Customer Company:', 'B10', 'Customer Company Name'),
        ('Contact Person:', 'B11', 'Contact Name'),
        ('Phone:', 'B12', '(555) 123-4567'),
        ('Email:', 'B13', 'contact@customer.com'),
        ('Customer Reference:', 'B14', ''),
        ('Payment Terms:', 'B15', 'Net 30 Days'),
        ('Delivery Terms:', 'B16', 'FOB Origin')
    ]
    
    for i, (label, cell, default) in enumerate(quote_fields, start=8):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = subheader_font
        ws[f'B{i}'] = default
        ws[f'B{i}'].font = normal_font
        ws[f'B{i}'].border = thin_border
    
    # Line Items Section
    ws['A18'] = 'LINE ITEMS'
    ws['A18'].font = subheader_font
    ws['A18'].fill = light_fill
    ws.merge_cells('A18:D18')
    
    # Line items headers
    headers = ['Item #', 'Description', 'Qty', 'Unit Price']
    for col, header in enumerate(['A', 'B', 'C', 'D'], start=19):
        ws[f'{header}19'] = headers[col-19]
        ws[f'{header}19'].font = subheader_font
        ws[f'{header}19'].fill = header_fill
        ws[f'{header}19'].alignment = Alignment(horizontal='center')
        ws[f'{header}19'].border = thin_border
    
    # Sample line items
    sample_items = [
        ['1', 'Sample Product 1', '2', '100.00'],
        ['2', 'Sample Product 2', '1', '250.00'],
        ['3', '', '', ''],
        ['4', '', '', ''],
        ['5', '', '', ''],
        ['6', '', '', ''],
        ['7', '', '', ''],
        ['8', '', '', ''],
        ['9', '', '', ''],
        ['10', '', '', '']
    ]
    
    for row_idx, item in enumerate(sample_items, start=20):
        for col_idx, value in enumerate(['A', 'B', 'C', 'D']):
            cell = ws[f'{value}{row_idx}']
            cell.value = item[col_idx]
            cell.font = normal_font
            cell.border = thin_border
            if col_idx == 1:  # Description column
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')
    
    # Totals Section
    ws['A31'] = 'TOTALS'
    ws['A31'].font = subheader_font
    ws['A31'].fill = light_fill
    ws.merge_cells('A31:D31')
    
    # Total fields
    total_fields = [
        ('Subtotal:', 'C32', '=SUMPRODUCT(C20:C29,D20:D29)'),
        ('Tax Rate:', 'B33', '5%'),
        ('Tax Amount:', 'C33', '=C32*0.05'),
        ('Freight:', 'B34', '0.00'),
        ('Freight Amount:', 'C34', '0.00'),
        ('TOTAL:', 'C35', '=C32+C33+C34')
    ]
    
    for i, (label, cell, formula) in enumerate(total_fields, start=32):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = subheader_font
        ws[f'{cell}'] = formula
        ws[f'{cell}'].font = normal_font
        ws[f'{cell}'].border = thin_border
        if i == 35:  # Total row
            ws[f'A{i}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'{cell}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'A{i}'].fill = light_fill
            ws[f'{cell}'].fill = light_fill
    
    # Buttons area
    ws['A37'] = 'ACTIONS'
    ws['A37'].font = subheader_font
    ws['A37'].fill = light_fill
    ws.merge_cells('A37:D37')
    
    # Button placeholders (will be actual buttons in VBA)
    ws['A39'] = '[Generate Quote Button]'
    ws['A39'].font = normal_font
    ws['A39'].alignment = Alignment(horizontal='center')
    ws['A39'].border = thin_border
    ws.merge_cells('A39:B39')
    
    ws['C39'] = '[Clear Form Button]'
    ws['C39'].font = normal_font
    ws['C39'].alignment = Alignment(horizontal='center')
    ws['C39'].border = thin_border
    ws.merge_cells('C39:D39')
    
    # Save the workbook
    wb.save('Quote_Generator.xlsx')
    print("Excel macro workbook created: Quote_Generator.xlsx")
    print("Note: VBA macros need to be added manually in Excel")

def create_vba_code():
    """Create VBA code for the Excel workbook"""
    
    vba_code = '''
Sub GenerateQuote()
    Dim ws As Worksheet
    Dim newWs As Worksheet
    Dim i As Integer
    Dim lastRow As Integer
    Dim quoteNum As String
    Dim quoteDate As String
    Dim customerCompany As String
    Dim contactPerson As String
    Dim phone As String
    Dim email As String
    Dim customerRef As String
    Dim paymentTerms As String
    Dim deliveryTerms As String
    Dim subtotal As Double
    Dim taxRate As Double
    Dim taxAmount As Double
    Dim freight As Double
    Dim total As Double
    
    ' Get data from form
    quoteNum = Range("B8").Value
    quoteDate = Range("B9").Value
    customerCompany = Range("B10").Value
    contactPerson = Range("B11").Value
    phone = Range("B12").Value
    email = Range("B13").Value
    customerRef = Range("B14").Value
    paymentTerms = Range("B15").Value
    deliveryTerms = Range("B16").Value
    subtotal = Range("C32").Value
    taxRate = Range("B33").Value
    taxAmount = Range("C33").Value
    freight = Range("B34").Value
    total = Range("C35").Value
    
    ' Create new worksheet
    Set newWs = Worksheets.Add
    newWs.Name = "Quote_" & quoteNum
    
    ' Set up the quote layout
    With newWs
        ' Header
        .Range("A1").Value = "QUOTE"
        .Range("A1").Font.Bold = True
        .Range("A1").Font.Size = 16
        .Range("A1").Interior.Color = RGB(54, 96, 146)
        .Range("A1").Font.Color = RGB(255, 255, 255)
        .Range("A1").HorizontalAlignment = xlCenter
        .Range("A1:F1").Merge
        
        ' Company info
        .Range("A2").Value = "KINDER MORGAN"
        .Range("A3").Value = "1234 Energy Drive"
        .Range("A4").Value = "Houston, TX 77002"
        .Range("A5").Value = "Phone: (713) 369-9000"
        
        ' Quote details
        .Range("D2").Value = "Quote No.:"
        .Range("E2").Value = quoteNum
        .Range("D3").Value = "Date:"
        .Range("E3").Value = quoteDate
        .Range("D4").Value = "Customer Ref:"
        .Range("E4").Value = customerRef
        
        ' Customer info
        .Range("A7").Value = "CUSTOMER INFORMATION"
        .Range("A7").Font.Bold = True
        .Range("A7").Interior.Color = RGB(242, 242, 242)
        .Range("A7:F7").Merge
        
        .Range("A8").Value = "Company:"
        .Range("B8").Value = customerCompany
        .Range("A9").Value = "Contact:"
        .Range("B9").Value = contactPerson
        .Range("A10").Value = "Phone:"
        .Range("B10").Value = phone
        .Range("A11").Value = "Email:"
        .Range("B11").Value = email
        
        ' Line items header
        .Range("A13").Value = "Item #"
        .Range("B13").Value = "Qty"
        .Range("C13").Value = "Unit"
        .Range("D13").Value = "Description"
        .Range("E13").Value = "Unit Price"
        .Range("F13").Value = "Total Price"
        
        ' Format header
        .Range("A13:F13").Font.Bold = True
        .Range("A13:F13").Interior.Color = RGB(54, 96, 146)
        .Range("A13:F13").Font.Color = RGB(255, 255, 255)
        .Range("A13:F13").HorizontalAlignment = xlCenter
        
        ' Add line items
        lastRow = 13
        For i = 20 To 29
            If Range("B" & i).Value <> "" Then
                lastRow = lastRow + 1
                .Range("A" & lastRow).Value = Range("A" & i).Value
                .Range("B" & lastRow).Value = Range("C" & i).Value
                .Range("C" & lastRow).Value = "EA"
                .Range("D" & lastRow).Value = Range("B" & i).Value
                .Range("E" & lastRow).Value = Range("D" & i).Value
                .Range("F" & lastRow).Value = "=B" & lastRow & "*E" & lastRow
                .Range("E" & lastRow & ":F" & lastRow).NumberFormat = "$#,##0.00"
            End If
        Next i
        
        ' Totals
        .Range("A" & lastRow + 2).Value = "Subtotal:"
        .Range("F" & lastRow + 2).Value = "=SUM(F14:F" & lastRow & ")"
        .Range("A" & lastRow + 3).Value = "Tax (5%):"
        .Range("F" & lastRow + 3).Value = "=F" & lastRow + 2 & "*0.05"
        .Range("A" & lastRow + 4).Value = "Freight:"
        .Range("F" & lastRow + 4).Value = freight
        .Range("A" & lastRow + 5).Value = "TOTAL:"
        .Range("F" & lastRow + 5).Value = "=F" & lastRow + 2 & "+F" & lastRow + 3 & "+F" & lastRow + 4
        
        ' Format totals
        .Range("A" & lastRow + 2 & ":F" & lastRow + 5).Font.Bold = True
        .Range("F" & lastRow + 2 & ":F" & lastRow + 5).NumberFormat = "$#,##0.00"
        
        ' Terms
        .Range("A" & lastRow + 7).Value = "TERMS AND CONDITIONS"
        .Range("A" & lastRow + 7).Font.Bold = True
        .Range("A" & lastRow + 7).Interior.Color = RGB(242, 242, 242)
        .Range("A" & lastRow + 7 & ":F" & lastRow + 7).Merge
        
        .Range("A" & lastRow + 8).Value = "Payment Terms: " & paymentTerms
        .Range("A" & lastRow + 9).Value = "Delivery: " & deliveryTerms
        .Range("A" & lastRow + 10).Value = "Validity: 30 days from quote date"
        
        ' Auto-fit columns
        .Columns.AutoFit
    End With
    
    MsgBox "Quote generated successfully!", vbInformation
End Sub

Sub ClearForm()
    ' Clear all form fields
    Range("B8").Value = ""
    Range("B10").Value = ""
    Range("B11").Value = ""
    Range("B12").Value = ""
    Range("B13").Value = ""
    Range("B14").Value = ""
    
    ' Clear line items
    Range("B20:B29").Value = ""
    Range("C20:C29").Value = ""
    Range("D20:D29").Value = ""
    
    ' Reset quote number
    Range("B8").Value = "QUOTE-" & Format(Now, "YYYYMMDD") & "-001"
    Range("B9").Value = "=TODAY()"
    
    MsgBox "Form cleared!", vbInformation
End Sub
'''
    
    # Save VBA code to file
    with open('Quote_Generator_VBA.txt', 'w') as f:
        f.write(vba_code)
    
    print("VBA code saved to: Quote_Generator_VBA.txt")
    print("Copy this code into the VBA editor in Excel")

if __name__ == "__main__":
    create_excel_macro_workbook()
    create_vba_code()
