"""
File parsing modules for the Desktop Quote Generator
Handles RTF, XML, XLSX, and CSV file imports
"""

import re
import xml.etree.ElementTree as ET
import openpyxl
import pandas as pd
from datetime import datetime

class FileParser:
    """Base class for file parsers"""
    
    def parse(self, file_path):
        """Parse file and return list of items"""
        raise NotImplementedError("Subclasses must implement parse method")
    
    def clean_text(self, text):
        """Clean and normalize text"""
        if not text:
            return ""
        return str(text).strip().replace('\n', ' ').replace('\r', ' ')
    
    def extract_price(self, price_text):
        """Extract numeric price from text"""
        if not price_text:
            return 0.0
        
        # Remove common currency symbols and text
        price_text = str(price_text).replace('$', '').replace(',', '').replace('USD', '')
        
        # Extract first number found
        match = re.search(r'(\d+\.?\d*)', price_text)
        if match:
            return float(match.group(1))
        return 0.0
    
    def extract_quantity(self, qty_text):
        """Extract numeric quantity from text"""
        if not qty_text:
            return 1
        
        # Extract first number found
        match = re.search(r'(\d+\.?\d*)', str(qty_text))
        if match:
            return int(float(match.group(1)))
        return 1

class RTFParser(FileParser):
    """Parser for RTF files"""
    
    def parse(self, file_path):
        """Parse RTF file and extract quote items"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Remove RTF formatting
            plain_text = self.clean_rtf(content)
            return self.extract_items_from_text(plain_text)
            
        except Exception as e:
            raise Exception(f"Error parsing RTF file: {str(e)}")
    
    def clean_rtf(self, rtf_content):
        """Remove RTF formatting codes"""
        # Remove RTF control codes
        plain_text = re.sub(r'\\[a-z]+\d*\s?', '', rtf_content)
        plain_text = re.sub(r'[{}]', '', plain_text)
        plain_text = re.sub(r'\\[^a-z]', '', plain_text)
        
        # Clean up whitespace
        plain_text = re.sub(r'\s+', ' ', plain_text)
        plain_text = re.sub(r'\n\s*\n', '\n\n', plain_text)
        
        return plain_text
    
    def extract_items_from_text(self, text):
        """Extract items from cleaned text"""
        items = []
        lines = text.split('\n')
        
        current_item = {}
        item_count = 0
        
        for line in lines:
            line = line.strip()
            if not line:
                if current_item and 'description' in current_item:
                    items.append(current_item)
                    current_item = {}
                continue
            
            # Look for item patterns
            if self.is_item_line(line):
                if current_item and 'description' in current_item:
                    items.append(current_item)
                
                current_item = self.parse_item_line(line)
                item_count += 1
            elif current_item and 'description' in current_item:
                # Add to current item description
                current_item['description'] += f" {line}"
        
        # Add last item
        if current_item and 'description' in current_item:
            items.append(current_item)
        
        return items
    
    def is_item_line(self, line):
        """Check if line contains item information"""
        # Look for common item patterns
        patterns = [
            r'\d+\s+\d+\s+PC',  # Quantity PC pattern
            r'Model no\.:',      # Model number
            r'Level, radar,',    # Product description start
            r'Qty:',             # Quantity label
            r'Price:',           # Price label
        ]
        
        return any(re.search(pattern, line, re.IGNORECASE) for pattern in patterns)
    
    def parse_item_line(self, line):
        """Parse individual item line"""
        item = {
            'description': '',
            'quantity': 1,
            'unit': 'EA',
            'unit_price': 0.0,
            'model': '',
            'order_code': ''
        }
        
        # Extract quantity
        qty_match = re.search(r'(\d+)\s+(\d+)\s+PC', line)
        if qty_match:
            item['quantity'] = int(qty_match.group(2))
        
        # Extract model number
        model_match = re.search(r'Model no\.:\s*([^\n]+)', line)
        if model_match:
            item['model'] = model_match.group(1).strip()
        
        # Extract order code
        order_match = re.search(r'\(([^)]+)\)', line)
        if order_match:
            item['order_code'] = order_match.group(1).strip()
        
        # Extract description
        desc_match = re.search(r'Level, radar,([^:]+)', line)
        if desc_match:
            item['description'] = desc_match.group(1).strip()
        
        # Extract price
        price_match = re.search(r'(\d{1,3}(?:,\d{3})*\.\d{2})', line)
        if price_match:
            item['unit_price'] = self.extract_price(price_match.group(1))
        
        return item

class XMLParser(FileParser):
    """Parser for XML files"""
    
    def parse(self, file_path):
        """Parse XML file and extract quote items"""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Try different XML structures
            items = []
            
            # Look for common item elements
            item_elements = (
                root.findall('.//item') or 
                root.findall('.//product') or 
                root.findall('.//line') or
                root.findall('.//lineItem')
            )
            
            for element in item_elements:
                item = self.parse_xml_item(element)
                if item:
                    items.append(item)
            
            return items
            
        except Exception as e:
            raise Exception(f"Error parsing XML file: {str(e)}")
    
    def parse_xml_item(self, element):
        """Parse individual XML item element"""
        item = {
            'description': '',
            'quantity': 1,
            'unit': 'EA',
            'unit_price': 0.0,
            'model': '',
            'order_code': ''
        }
        
        # Extract description
        desc_elem = (
            element.find('description') or 
            element.find('name') or 
            element.find('title') or
            element.find('productName')
        )
        if desc_elem is not None and desc_elem.text:
            item['description'] = self.clean_text(desc_elem.text)
        
        # Extract quantity
        qty_elem = (
            element.find('quantity') or 
            element.find('qty') or
            element.find('amount')
        )
        if qty_elem is not None and qty_elem.text:
            item['quantity'] = self.extract_quantity(qty_elem.text)
        
        # Extract unit price
        price_elem = (
            element.find('price') or 
            element.find('unitPrice') or 
            element.find('unitprice') or
            element.find('cost')
        )
        if price_elem is not None and price_elem.text:
            item['unit_price'] = self.extract_price(price_elem.text)
        
        # Extract model
        model_elem = (
            element.find('model') or 
            element.find('modelNumber') or
            element.find('partNumber')
        )
        if model_elem is not None and model_elem.text:
            item['model'] = self.clean_text(model_elem.text)
        
        # Extract order code
        order_elem = (
            element.find('orderCode') or 
            element.find('sku') or
            element.find('itemCode')
        )
        if order_elem is not None and order_elem.text:
            item['order_code'] = self.clean_text(order_elem.text)
        
        # Only return item if it has a description
        return item if item['description'] else None

class ExcelParser(FileParser):
    """Parser for Excel files"""
    
    def parse(self, file_path):
        """Parse Excel file and extract quote items"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            items = []
            
            # Try to find data starting from row 2 (assuming row 1 is header)
            for row in range(2, ws.max_row + 1):
                item = self.parse_excel_row(ws, row)
                if item and item['description']:
                    items.append(item)
            
            return items
            
        except Exception as e:
            raise Exception(f"Error parsing Excel file: {str(e)}")
    
    def parse_excel_row(self, worksheet, row):
        """Parse individual Excel row"""
        item = {
            'description': '',
            'quantity': 1,
            'unit': 'EA',
            'unit_price': 0.0,
            'model': '',
            'order_code': '',
            'config': '',
            'sales_text': '',
            'delivery_time': '',
            'country_origin': '',
            'country_dispatch': '',
            'customer_ref': ''
        }
        
        # Get cell values
        cells = []
        for col in range(1, 20):  # Check up to column S
            cell_value = worksheet.cell(row=row, column=col).value
            cells.append(str(cell_value) if cell_value else '')
        
        # Skip empty rows
        if all(not cell.strip() for cell in cells):
            return None
        
        # Skip header rows - look for common header patterns
        first_cell = cells[0].lower().strip()
        if any(header in first_cell for header in [
            'sales document', 'endress+hauser', 'quote', 'product details', 
            'item', 'quantity', 'unit', 'product name', 'order code'
        ]):
            return None
        
        # For this specific format, the columns are:
        # Col 1: Item number (numeric), Col 2: Quantity, Col 3: Unit, Col 4: Product name, Col 5: Order code
        
        # Check if first cell is a numeric item number (like 20.0, 30.0)
        try:
            item_number = float(cells[0])
            if item_number > 0:  # Valid item number
                # Extract data based on known column structure
                if len(cells) >= 4:
                    # Quantity (column 2)
                    try:
                        item['quantity'] = int(float(cells[1])) if cells[1] else 1
                    except (ValueError, TypeError):
                        item['quantity'] = 1
                    
                    # Unit (column 3)
                    item['unit'] = cells[2].strip() if cells[2] else 'EA'
                    
                    # Description/Product name (column 4)
                    base_description = self.clean_text(cells[3]) if cells[3] else ''
                    
                    # Order code (column 5)
                    if len(cells) >= 5 and cells[4]:
                        item['order_code'] = self.clean_text(cells[4])
                    
                    # Model number (column 7 - Endress+Hauser material number)
                    if len(cells) >= 7 and cells[6]:
                        item['model'] = self.clean_text(cells[6])
                    
                    # Configuration details (column 8) - this contains the detailed product description
                    config_details = ''
                    if len(cells) >= 8 and cells[7]:
                        config_details = self.clean_text(cells[7])
                        item['config'] = config_details
                    
                    # Sales text (column 9)
                    if len(cells) >= 9 and cells[8]:
                        item['sales_text'] = self.clean_text(cells[8])
                    
                    # Delivery time (column 10)
                    if len(cells) >= 10 and cells[9]:
                        item['delivery_time'] = self.clean_text(cells[9])
                    
                    # Look for price in columns 11-12 (unit price and total price)
                    for i in range(10, min(13, len(cells))):  # Columns 11-12 (0-indexed 10-11)
                        if cells[i] and self.is_price_value(cells[i]):
                            item['unit_price'] = self.extract_price(cells[i])
                            break
                    
                    # Country of origin (column 13)
                    if len(cells) >= 13 and cells[12]:
                        item['country_origin'] = self.clean_text(cells[12])
                    
                    # Country of dispatch (column 14)
                    if len(cells) >= 14 and cells[13]:
                        item['country_dispatch'] = self.clean_text(cells[13])
                    
                    # Customer reference (column 6)
                    if len(cells) >= 6 and cells[5]:
                        item['customer_ref'] = self.clean_text(cells[5])
                    
                    # Create comprehensive description with all details
                    description_parts = []
                    
                    # Main product name
                    if base_description:
                        description_parts.append(base_description)
                    
                    # Sales text (product description)
                    if item['sales_text']:
                        description_parts.append(f"Description: {item['sales_text']}")
                    
                    # Delivery time
                    if item['delivery_time']:
                        description_parts.append(f"Delivery time: {item['delivery_time']}")
                    
                    # Order code description
                    if item['order_code']:
                        description_parts.append(f"Order code description:")
                        description_parts.append(item['order_code'])
                    
                    # Product configuration
                    if config_details:
                        description_parts.append("Product Configuration:")
                        description_parts.append(config_details)
                    
                    # Country information
                    country_info = []
                    if item['country_origin']:
                        country_info.append(f"Origin: {item['country_origin']}")
                    if item['country_dispatch']:
                        country_info.append(f"Dispatch: {item['country_dispatch']}")
                    if country_info:
                        description_parts.append("Country Information:")
                        description_parts.extend(country_info)
                    
                    item['description'] = "\n\n".join(description_parts)
                    
                    return item if item['description'] else None
        except (ValueError, TypeError):
            pass
        
        # Fallback to original parsing logic for other formats
        description_found = False
        price_found = False
        qty_found = False
        
        for i, cell in enumerate(cells):
            if not cell.strip():
                continue
            
            # Check if it's a description (longer text, no numbers)
            if len(cell) > 10 and not re.search(r'^\d+\.?\d*$', cell) and not description_found:
                item['description'] = self.clean_text(cell)
                description_found = True
                continue
            
            # Check if it's a quantity (small integer)
            if re.search(r'^\d+$', cell) and int(cell) < 1000 and not qty_found:
                item['quantity'] = int(cell)
                qty_found = True
                continue
            
            # Check if it's a price (decimal number)
            if self.is_price_value(cell) and not price_found:
                item['unit_price'] = self.extract_price(cell)
                price_found = True
                continue
            
            # Check if it's a model/part number (alphanumeric, shorter)
            if re.search(r'^[A-Za-z0-9\-_]+$', cell) and len(cell) < 20 and not item['model']:
                item['model'] = cell
                continue
        
        # If no description found, use first non-empty cell
        if not description_found and cells[0].strip():
            item['description'] = self.clean_text(cells[0])
        
        # Try to extract additional fields from other cells
        for i, cell in enumerate(cells):
            if not cell.strip():
                continue
            
            # Look for sales text (longer descriptive text)
            if len(cell) > 50 and 'sales_text' not in item and not self.is_price_value(cell):
                item['sales_text'] = self.clean_text(cell)
            
            # Look for delivery time patterns
            if 'wrk.day' in cell.lower() or 'day' in cell.lower():
                item['delivery_time'] = self.clean_text(cell)
            
            # Look for country codes (2-3 letter codes)
            if len(cell) <= 3 and cell.isalpha() and len(cell) >= 2:
                if 'country_origin' not in item:
                    item['country_origin'] = cell.upper()
                elif 'country_dispatch' not in item:
                    item['country_dispatch'] = cell.upper()
        
        return item if item['description'] else None
    
    def is_price_value(self, value):
        """Check if a value looks like a price"""
        if not value:
            return False
        
        # Remove common currency symbols and text
        clean_value = str(value).replace('$', '').replace(',', '').replace('USD', '').strip()
        
        # Check if it's a valid price (decimal number)
        try:
            price = float(clean_value)
            return 0.01 <= price <= 1000000  # Reasonable price range
        except (ValueError, TypeError):
            return False

class CSVParser(FileParser):
    """Parser for CSV files"""
    
    def parse(self, file_path):
        """Parse CSV file and extract quote items"""
        try:
            df = pd.read_csv(file_path)
            items = []
            
            for _, row in df.iterrows():
                item = self.parse_csv_row(row)
                if item and item['description']:
                    items.append(item)
            
            return items
            
        except Exception as e:
            raise Exception(f"Error parsing CSV file: {str(e)}")
    
    def parse_csv_row(self, row):
        """Parse individual CSV row"""
        item = {
            'description': '',
            'quantity': 1,
            'unit': 'EA',
            'unit_price': 0.0,
            'model': '',
            'order_code': ''
        }
        
        # Try to map columns by name
        columns = row.index.tolist()
        
        for col in columns:
            col_lower = col.lower()
            value = str(row[col]) if pd.notna(row[col]) else ''
            
            if not value.strip():
                continue
            
            if any(keyword in col_lower for keyword in ['description', 'desc', 'name', 'product']):
                item['description'] = self.clean_text(value)
            elif any(keyword in col_lower for keyword in ['quantity', 'qty', 'amount']):
                item['quantity'] = self.extract_quantity(value)
            elif any(keyword in col_lower for keyword in ['price', 'cost', 'unit_price']):
                item['unit_price'] = self.extract_price(value)
            elif any(keyword in col_lower for keyword in ['model', 'part', 'sku']):
                item['model'] = self.clean_text(value)
            elif any(keyword in col_lower for keyword in ['unit', 'uom']):
                item['unit'] = self.clean_text(value)
        
        # If no specific mapping worked, try positional mapping
        if not item['description'] and len(columns) > 0:
            item['description'] = self.clean_text(str(row.iloc[0]))
        
        if item['quantity'] == 1 and len(columns) > 1:
            item['quantity'] = self.extract_quantity(str(row.iloc[1]))
        
        if item['unit_price'] == 0.0 and len(columns) > 2:
            item['unit_price'] = self.extract_price(str(row.iloc[2]))
        
        return item if item['description'] else None

def get_parser(file_path):
    """Get appropriate parser for file type"""
    file_ext = file_path.lower().split('.')[-1]
    
    if file_ext == 'rtf':
        return RTFParser()
    elif file_ext in ['xml']:
        return XMLParser()
    elif file_ext in ['xlsx', 'xls']:
        return ExcelParser()
    elif file_ext == 'csv':
        return CSVParser()
    else:
        raise Exception(f"Unsupported file type: {file_ext}")
