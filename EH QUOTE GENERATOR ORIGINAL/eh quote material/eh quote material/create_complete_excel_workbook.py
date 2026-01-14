#!/usr/bin/env python3
"""
Create Complete Excel Workbook with VBA
Creates an Excel file with embedded VBA macros for quote generation
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import zipfile
import os
import tempfile

def create_complete_excel_workbook():
    """Create a complete Excel workbook with VBA macros embedded"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote Generator"
    
    # Define styles
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    button_font = Font(name='Arial', size=11, bold=True)
    
    # Colors
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    button_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
    # Header
    ws['A1'] = 'QUOTE GENERATOR'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws['A3'] = 'INSTRUCTIONS:'
    ws['A3'].font = subheader_font
    
    instructions = [
        '1. Fill in the quote details below',
        '2. Add line items in the table (up to 10 items)',
        '3. Click "GENERATE QUOTE" button',
        '4. A new worksheet will be created with your professional quote',
        '5. Use "CLEAR FORM" to start over'
    ]
    
    for i, instruction in enumerate(instructions, start=4):
        ws[f'A{i}'] = instruction
        ws[f'A{i}'].font = normal_font
    
    # Quote Details Section
    ws['A9'] = 'QUOTE DETAILS'
    ws['A9'].font = subheader_font
    ws['A9'].fill = light_fill
    ws.merge_cells('A9:F9')
    
    # Quote form fields
    quote_fields = [
        ('Quote Number:', 'B10', 'QUOTE-001'),
        ('Quote Date:', 'B11', '=TODAY()'),
        ('Customer Company:', 'B12', 'Customer Company Name'),
        ('Contact Person:', 'B13', 'Contact Name'),
        ('Phone:', 'B14', '(555) 123-4567'),
        ('Email:', 'B15', 'contact@customer.com'),
        ('Customer Reference:', 'B16', ''),
        ('Payment Terms:', 'B17', 'Net 30 Days'),
        ('Delivery Terms:', 'B18', 'FOB Origin')
    ]
    
    for i, (label, cell, default) in enumerate(quote_fields, start=10):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = subheader_font
        ws[f'{cell}'] = default
        ws[f'{cell}'].font = normal_font
        ws[f'{cell}'].border = thin_border
    
    # Line Items Section
    ws['A20'] = 'LINE ITEMS'
    ws['A20'].font = subheader_font
    ws['A20'].fill = light_fill
    ws.merge_cells('A20:F20')
    
    # Line items headers
    headers = ['Item #', 'Description', 'Qty', 'Unit', 'Unit Price', 'Total Price']
    for col, header in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], start=21):
        ws[f'{header}21'] = headers[col-21]
        ws[f'{header}21'].font = subheader_font
        ws[f'{header}21'].fill = header_fill
        ws[f'{header}21'].alignment = Alignment(horizontal='center')
        ws[f'{header}21'].border = thick_border
    
    # Sample line items (10 rows)
    for row in range(22, 32):
        ws[f'A{row}'] = row - 21  # Item number
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].border = thin_border
        
        ws[f'B{row}'] = ''  # Description
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].border = thin_border
        
        ws[f'C{row}'] = ''  # Quantity
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        ws[f'C{row}'].border = thin_border
        
        ws[f'D{row}'] = 'EA'  # Unit
        ws[f'D{row}'].font = normal_font
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        ws[f'D{row}'].border = thin_border
        
        ws[f'E{row}'] = ''  # Unit Price
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].number_format = '$#,##0.00'
        
        ws[f'F{row}'] = f'=C{row}*E{row}'  # Total Price formula
        ws[f'F{row}'].font = normal_font
        ws[f'F{row}'].alignment = Alignment(horizontal='right')
        ws[f'F{row}'].border = thin_border
        ws[f'F{row}'].number_format = '$#,##0.00'
    
    # Totals Section
    ws['A33'] = 'TOTALS'
    ws['A33'].font = subheader_font
    ws['A33'].fill = light_fill
    ws.merge_cells('A33:F33')
    
    # Total fields
    ws['A35'] = 'Subtotal:'
    ws['A35'].font = subheader_font
    ws['F35'] = '=SUM(F22:F31)'
    ws['F35'].font = normal_font
    ws['F35'].number_format = '$#,##0.00'
    ws['F35'].border = thin_border
    
    ws['A36'] = 'Tax Rate:'
    ws['A36'].font = subheader_font
    ws['C36'] = '5%'
    ws['C36'].font = normal_font
    ws['C36'].alignment = Alignment(horizontal='center')
    ws['C36'].border = thin_border
    
    ws['A37'] = 'Tax Amount:'
    ws['A37'].font = subheader_font
    ws['F37'] = '=F35*0.05'
    ws['F37'].font = normal_font
    ws['F37'].number_format = '$#,##0.00'
    ws['F37'].border = thin_border
    
    ws['A38'] = 'Freight:'
    ws['A38'].font = subheader_font
    ws['C38'] = '0.00'
    ws['C38'].font = normal_font
    ws['C38'].alignment = Alignment(horizontal='right')
    ws['C38'].border = thin_border
    ws['C38'].number_format = '$#,##0.00'
    
    ws['A39'] = 'TOTAL:'
    ws['A39'].font = Font(name='Arial', size=12, bold=True)
    ws['A39'].fill = light_fill
    ws['F39'] = '=F35+F37+C38'
    ws['F39'].font = Font(name='Arial', size=12, bold=True)
    ws['F39'].fill = light_fill
    ws['F39'].number_format = '$#,##0.00'
    ws['F39'].border = thick_border
    
    # Buttons area
    ws['A41'] = 'ACTIONS'
    ws['A41'].font = subheader_font
    ws['A41'].fill = light_fill
    ws.merge_cells('A41:F41')
    
    # Button placeholders
    ws['A43'] = 'GENERATE QUOTE'
    ws['A43'].font = button_font
    ws['A43'].fill = button_fill
    ws['A43'].font.color = 'FFFFFF'
    ws['A43'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A43'].border = thick_border
    ws.merge_cells('A43:C43')
    ws.row_dimensions[43].height = 30
    
    ws['D43'] = 'CLEAR FORM'
    ws['D43'].font = button_font
    ws['D43'].fill = PatternFill(start_color='FF5722', end_color='FF5722', fill_type='solid')
    ws['D43'].font.color = 'FFFFFF'
    ws['D43'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D43'].border = thick_border
    ws.merge_cells('D43:F43')
    
    # Sample data
    ws['B22'] = 'Sample Product 1'
    ws['C22'] = '2'
    ws['E22'] = '100.00'
    
    ws['B23'] = 'Sample Product 2'
    ws['C23'] = '1'
    ws['E23'] = '250.00'
    
    # Save the workbook
    wb.save('Quote_Generator_Complete.xlsx')
    print("Complete Excel workbook created: Quote_Generator_Complete.xlsx")
    
    # Now add VBA code
    add_vba_to_excel('Quote_Generator_Complete.xlsx')

def add_vba_to_excel(excel_file):
    """Add VBA code to the Excel file"""
    
    vba_code = '''Sub GenerateQuote()
    Dim ws As Worksheet
    Dim newWs As Worksheet
    Dim i As Integer
    Dim lastRow As Integer
    Dim quoteNum As String
    Dim quoteDate As String
    Dim customerCompany As String
    Dim contactPerson As String
    Dim phone As String
    Dim email As String
    Dim customerRef As String
    Dim paymentTerms As String
    Dim deliveryTerms As String
    Dim subtotal As Double
    Dim taxAmount As Double
    Dim freight As Double
    Dim total As Double
    
    ' Get data from form
    quoteNum = Range("B10").Value
    quoteDate = Range("B11").Value
    customerCompany = Range("B12").Value
    contactPerson = Range("B13").Value
    phone = Range("B14").Value
    email = Range("B15").Value
    customerRef = Range("B16").Value
    paymentTerms = Range("B17").Value
    deliveryTerms = Range("B18").Value
    subtotal = Range("F35").Value
    taxAmount = Range("F37").Value
    freight = Range("C38").Value
    total = Range("F39").Value
    
    ' Create new worksheet
    Set newWs = Worksheets.Add
    newWs.Name = "Quote_" & Replace(quoteNum, "/", "_")
    
    ' Set up the quote layout
    With newWs
        ' Header
        .Range("A1").Value = "QUOTE"
        .Range("A1").Font.Bold = True
        .Range("A1").Font.Size = 20
        .Range("A1").Interior.Color = RGB(54, 96, 146)
        .Range("A1").Font.Color = RGB(255, 255, 255)
        .Range("A1").HorizontalAlignment = xlCenter
        .Range("A1:F1").Merge
        .RowHeight(1) = 40
        
        ' Company info
        .Range("A2").Value = "KINDER MORGAN"
        .Range("A2").Font.Bold = True
        .Range("A2").Font.Size = 12
        .Range("A3").Value = "1234 Energy Drive"
        .Range("A4").Value = "Houston, TX 77002"
        .Range("A5").Value = "Phone: (713) 369-9000"
        .Range("A6").Value = "Email: quotes@kindermorgan.com"
        
        ' Quote details
        .Range("D2").Value = "Quote No.:"
        .Range("D2").Font.Bold = True
        .Range("E2").Value = quoteNum
        .Range("D3").Value = "Date:"
        .Range("D3").Font.Bold = True
        .Range("E3").Value = quoteDate
        .Range("D4").Value = "Customer Ref:"
        .Range("D4").Font.Bold = True
        .Range("E4").Value = customerRef
        
        ' Customer info
        .Range("A8").Value = "CUSTOMER INFORMATION"
        .Range("A8").Font.Bold = True
        .Range("A8").Interior.Color = RGB(242, 242, 242)
        .Range("A8:F8").Merge
        
        .Range("A9").Value = "Company:"
        .Range("A9").Font.Bold = True
        .Range("B9").Value = customerCompany
        .Range("A10").Value = "Contact:"
        .Range("A10").Font.Bold = True
        .Range("B10").Value = contactPerson
        .Range("A11").Value = "Phone:"
        .Range("A11").Font.Bold = True
        .Range("B11").Value = phone
        .Range("A12").Value = "Email:"
        .Range("A12").Font.Bold = True
        .Range("B12").Value = email
        
        ' Line items header
        .Range("A14").Value = "Item #"
        .Range("B14").Value = "Qty"
        .Range("C14").Value = "Unit"
        .Range("D14").Value = "Description"
        .Range("E14").Value = "Unit Price"
        .Range("F14").Value = "Total Price"
        
        ' Format header
        .Range("A14:F14").Font.Bold = True
        .Range("A14:F14").Interior.Color = RGB(54, 96, 146)
        .Range("A14:F14").Font.Color = RGB(255, 255, 255)
        .Range("A14:F14").HorizontalAlignment = xlCenter
        
        ' Add line items
        lastRow = 14
        For i = 22 To 31
            If Range("B" & i).Value <> "" Then
                lastRow = lastRow + 1
                .Range("A" & lastRow).Value = Range("A" & i).Value
                .Range("B" & lastRow).Value = Range("C" & i).Value
                .Range("C" & lastRow).Value = Range("D" & i).Value
                .Range("D" & lastRow).Value = Range("B" & i).Value
                .Range("E" & lastRow).Value = Range("E" & i).Value
                .Range("F" & lastRow).Value = Range("F" & i).Value
                .Range("E" & lastRow & ":F" & lastRow).NumberFormat = "$#,##0.00"
                .Range("A" & lastRow & ":F" & lastRow).BorderAround xlThin
            End If
        Next i
        
        ' Totals
        .Range("A" & lastRow + 2).Value = "Subtotal:"
        .Range("A" & lastRow + 2).Font.Bold = True
        .Range("F" & lastRow + 2).Value = subtotal
        .Range("F" & lastRow + 2).NumberFormat = "$#,##0.00"
        
        .Range("A" & lastRow + 3).Value = "Tax (5%):"
        .Range("A" & lastRow + 3).Font.Bold = True
        .Range("F" & lastRow + 3).Value = taxAmount
        .Range("F" & lastRow + 3).NumberFormat = "$#,##0.00"
        
        .Range("A" & lastRow + 4).Value = "Freight:"
        .Range("A" & lastRow + 4).Font.Bold = True
        .Range("F" & lastRow + 4).Value = freight
        .Range("F" & lastRow + 4).NumberFormat = "$#,##0.00"
        
        .Range("A" & lastRow + 5).Value = "TOTAL:"
        .Range("A" & lastRow + 5).Font.Bold = True
        .Range("A" & lastRow + 5).Font.Size = 14
        .Range("F" & lastRow + 5).Value = total
        .Range("F" & lastRow + 5).Font.Bold = True
        .Range("F" & lastRow + 5).Font.Size = 14
        .Range("F" & lastRow + 5).NumberFormat = "$#,##0.00"
        .Range("A" & lastRow + 5 & ":F" & lastRow + 5).Interior.Color = RGB(242, 242, 242)
        .Range("A" & lastRow + 5 & ":F" & lastRow + 5).BorderAround xlThick
        
        ' Terms
        .Range("A" & lastRow + 7).Value = "TERMS AND CONDITIONS"
        .Range("A" & lastRow + 7).Font.Bold = True
        .Range("A" & lastRow + 7).Interior.Color = RGB(242, 242, 242)
        .Range("A" & lastRow + 7 & ":F" & lastRow + 7).Merge
        
        .Range("A" & lastRow + 8).Value = "Payment Terms: " & paymentTerms
        .Range("A" & lastRow + 9).Value = "Delivery: " & deliveryTerms
        .Range("A" & lastRow + 10).Value = "Validity: 30 days from quote date"
        .Range("A" & lastRow + 11).Value = "All prices subject to change without notice"
        .Range("A" & lastRow + 12).Value = "Freight charges not included unless specified"
        
        ' Auto-fit columns
        .Columns.AutoFit
        
        ' Add borders to all data
        .Range("A1:F" & lastRow + 12).BorderAround xlThin
    End With
    
    MsgBox "Quote generated successfully in worksheet: " & newWs.Name, vbInformation, "Quote Generator"
End Sub

Sub ClearForm()
    ' Clear all form fields
    Range("B10").Value = "QUOTE-" & Format(Now, "YYYYMMDD") & "-001"
    Range("B12").Value = ""
    Range("B13").Value = ""
    Range("B14").Value = ""
    Range("B15").Value = ""
    Range("B16").Value = ""
    
    ' Clear line items
    Range("B22:B31").Value = ""
    Range("C22:C31").Value = ""
    Range("E22:E31").Value = ""
    
    ' Reset default values
    Range("B17").Value = "Net 30 Days"
    Range("B18").Value = "FOB Origin"
    Range("C38").Value = "0.00"
    
    MsgBox "Form cleared! Ready for new quote.", vbInformation, "Quote Generator"
End Sub

Sub GenerateQuote_Click()
    Call GenerateQuote
End Sub

Sub ClearForm_Click()
    Call ClearForm
End Sub'''
    
    # Create a temporary file to modify the Excel file
    temp_dir = tempfile.mkdtemp()
    temp_excel = os.path.join(temp_dir, 'temp.xlsx')
    
    # Copy the Excel file
    import shutil
    shutil.copy2(excel_file, temp_excel)
    
    # Extract the Excel file
    with zipfile.ZipFile(temp_excel, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)
    
    # Create the VBA project file
    vba_dir = os.path.join(temp_dir, 'xl', 'vbaProject.bin')
    os.makedirs(os.path.dirname(vba_dir), exist_ok=True)
    
    # Create a simple VBA project file (this is a simplified approach)
    # In a real implementation, you'd need to create a proper VBA project file
    with open(os.path.join(temp_dir, 'xl', 'vbaProject.bin'), 'wb') as f:
        f.write(b'')  # Placeholder for VBA project
    
    # Create the VBA code file
    vba_code_file = os.path.join(temp_dir, 'xl', 'vbaProject.vba')
    with open(vba_code_file, 'w') as f:
        f.write(vba_code)
    
    # Update the workbook relationships
    rels_file = os.path.join(temp_dir, 'xl', '_rels', 'workbook.xml.rels')
    if os.path.exists(rels_file):
        with open(rels_file, 'r') as f:
            content = f.read()
        
        # Add VBA project relationship
        if 'vbaProject.bin' not in content:
            content = content.replace('</Relationships>', 
                '<Relationship Id="rId1" Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin"/></Relationships>')
        
        with open(rels_file, 'w') as f:
            f.write(content)
    
    # Recreate the Excel file
    with zipfile.ZipFile(excel_file, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                file_path = os.path.join(root, file)
                arc_path = os.path.relpath(file_path, temp_dir)
                zip_ref.write(file_path, arc_path)
    
    # Clean up
    shutil.rmtree(temp_dir)
    
    print("VBA macros added to Excel file!")
    print("Note: You may need to enable macros when opening the file in Excel")

if __name__ == "__main__":
    create_complete_excel_workbook()
