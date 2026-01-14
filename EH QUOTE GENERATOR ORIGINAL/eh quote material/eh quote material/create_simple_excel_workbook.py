#!/usr/bin/env python3
"""
Create Simple Excel Workbook
Creates an Excel file that can be used for quote generation
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_simple_excel_workbook():
    """Create a simple Excel workbook for quote generation"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote Generator"
    
    # Define styles
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    button_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    
    # Colors
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    button_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    clear_fill = PatternFill(start_color='FF5722', end_color='FF5722', fill_type='solid')
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
    # Header
    ws['A1'] = 'QUOTE GENERATOR'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws['A3'] = 'INSTRUCTIONS:'
    ws['A3'].font = subheader_font
    
    instructions = [
        '1. Fill in the quote details below',
        '2. Add line items in the table (up to 10 items)',
        '3. Copy the data to create your quote manually',
        '4. Or use the Python script to generate quotes automatically',
        '5. Use "CLEAR FORM" to start over'
    ]
    
    for i, instruction in enumerate(instructions, start=4):
        ws[f'A{i}'] = instruction
        ws[f'A{i}'].font = normal_font
    
    # Quote Details Section
    ws['A9'] = 'QUOTE DETAILS'
    ws['A9'].font = subheader_font
    ws['A9'].fill = light_fill
    ws.merge_cells('A9:F9')
    
    # Quote form fields
    quote_fields = [
        ('Quote Number:', 'B10', 'QUOTE-001'),
        ('Quote Date:', 'B11', datetime.now().strftime('%m/%d/%Y')),
        ('Customer Company:', 'B12', 'Customer Company Name'),
        ('Contact Person:', 'B13', 'Contact Name'),
        ('Phone:', 'B14', '(555) 123-4567'),
        ('Email:', 'B15', 'contact@customer.com'),
        ('Customer Reference:', 'B16', ''),
        ('Payment Terms:', 'B17', 'Net 30 Days'),
        ('Delivery Terms:', 'B18', 'FOB Origin')
    ]
    
    for i, (label, cell, default) in enumerate(quote_fields, start=10):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = subheader_font
        ws[f'{cell}'] = default
        ws[f'{cell}'].font = normal_font
        ws[f'{cell}'].border = thin_border
    
    # Line Items Section
    ws['A20'] = 'LINE ITEMS'
    ws['A20'].font = subheader_font
    ws['A20'].fill = light_fill
    ws.merge_cells('A20:F20')
    
    # Line items headers
    headers = ['Item #', 'Description', 'Qty', 'Unit', 'Unit Price', 'Total Price']
    for col, header in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], start=21):
        ws[f'{header}21'] = headers[col-21]
        ws[f'{header}21'].font = subheader_font
        ws[f'{header}21'].fill = header_fill
        ws[f'{header}21'].alignment = Alignment(horizontal='center')
        ws[f'{header}21'].border = thick_border
    
    # Sample line items (10 rows)
    for row in range(22, 32):
        ws[f'A{row}'] = row - 21  # Item number
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].border = thin_border
        
        ws[f'B{row}'] = ''  # Description
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].border = thin_border
        
        ws[f'C{row}'] = ''  # Quantity
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        ws[f'C{row}'].border = thin_border
        
        ws[f'D{row}'] = 'EA'  # Unit
        ws[f'D{row}'].font = normal_font
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        ws[f'D{row}'].border = thin_border
        
        ws[f'E{row}'] = ''  # Unit Price
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].number_format = '$#,##0.00'
        
        ws[f'F{row}'] = f'=C{row}*E{row}'  # Total Price formula
        ws[f'F{row}'].font = normal_font
        ws[f'F{row}'].alignment = Alignment(horizontal='right')
        ws[f'F{row}'].border = thin_border
        ws[f'F{row}'].number_format = '$#,##0.00'
    
    # Totals Section
    ws['A33'] = 'TOTALS'
    ws['A33'].font = subheader_font
    ws['A33'].fill = light_fill
    ws.merge_cells('A33:F33')
    
    # Total fields
    ws['A35'] = 'Subtotal:'
    ws['A35'].font = subheader_font
    ws['F35'] = '=SUM(F22:F31)'
    ws['F35'].font = normal_font
    ws['F35'].number_format = '$#,##0.00'
    ws['F35'].border = thin_border
    
    ws['A36'] = 'Tax Rate:'
    ws['A36'].font = subheader_font
    ws['C36'] = '5%'
    ws['C36'].font = normal_font
    ws['C36'].alignment = Alignment(horizontal='center')
    ws['C36'].border = thin_border
    
    ws['A37'] = 'Tax Amount:'
    ws['A37'].font = subheader_font
    ws['F37'] = '=F35*0.05'
    ws['F37'].font = normal_font
    ws['F37'].number_format = '$#,##0.00'
    ws['F37'].border = thin_border
    
    ws['A38'] = 'Freight:'
    ws['A38'].font = subheader_font
    ws['C38'] = '0.00'
    ws['C38'].font = normal_font
    ws['C38'].alignment = Alignment(horizontal='right')
    ws['C38'].border = thin_border
    ws['C38'].number_format = '$#,##0.00'
    
    ws['A39'] = 'TOTAL:'
    ws['A39'].font = Font(name='Arial', size=12, bold=True)
    ws['A39'].fill = light_fill
    ws['F39'] = '=F35+F37+C38'
    ws['F39'].font = Font(name='Arial', size=12, bold=True)
    ws['F39'].fill = light_fill
    ws['F39'].number_format = '$#,##0.00'
    ws['F39'].border = thick_border
    
    # Buttons area
    ws['A41'] = 'ACTIONS'
    ws['A41'].font = subheader_font
    ws['A41'].fill = light_fill
    ws.merge_cells('A41:F41')
    
    # Button placeholders
    ws['A43'] = 'GENERATE QUOTE'
    ws['A43'].font = button_font
    ws['A43'].fill = button_fill
    ws['A43'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A43'].border = thick_border
    ws.merge_cells('A43:C43')
    ws.row_dimensions[43].height = 30
    
    ws['D43'] = 'CLEAR FORM'
    ws['D43'].font = button_font
    ws['D43'].fill = clear_fill
    ws['D43'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D43'].border = thick_border
    ws.merge_cells('D43:F43')
    
    # Sample data
    ws['B22'] = 'Sample Product 1'
    ws['C22'] = '2'
    ws['E22'] = '100.00'
    
    ws['B23'] = 'Sample Product 2'
    ws['C23'] = '1'
    ws['E23'] = '250.00'
    
    # Add a second worksheet for the quote template
    quote_ws = wb.create_sheet("Quote Template")
    
    # Create quote template
    create_quote_template(quote_ws)
    
    # Save the workbook
    wb.save('Quote_Generator_Excel.xlsx')
    print("Excel workbook created: Quote_Generator_Excel.xlsx")
    print("This file contains:")
    print("1. Quote Generator sheet - Fill in your data")
    print("2. Quote Template sheet - Copy your data here to create quotes")
    print("3. Use the Python script for automated quote generation")

def create_quote_template(ws):
    """Create a quote template worksheet"""
    
    # Define styles
    header_font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    # Colors
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Header
    ws['A1'] = 'QUOTE'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 40
    
    # Company info
    ws['A2'] = 'KINDER MORGAN'
    ws['A2'].font = Font(name='Arial', size=12, bold=True)
    ws['A3'] = '1234 Energy Drive'
    ws['A4'] = 'Houston, TX 77002'
    ws['A5'] = 'Phone: (713) 369-9000'
    ws['A6'] = 'Email: quotes@kindermorgan.com'
    
    # Quote details
    ws['D2'] = 'Quote No.:'
    ws['D2'].font = subheader_font
    ws['E2'] = 'QUOTE-001'
    ws['D3'] = 'Date:'
    ws['D3'].font = subheader_font
    ws['E3'] = datetime.now().strftime('%m/%d/%Y')
    ws['D4'] = 'Customer Ref:'
    ws['D4'].font = subheader_font
    ws['E4'] = 'REF-001'
    
    # Customer info
    ws['A8'] = 'CUSTOMER INFORMATION'
    ws['A8'].font = subheader_font
    ws['A8'].fill = light_fill
    ws.merge_cells('A8:F8')
    
    ws['A9'] = 'Company:'
    ws['A9'].font = subheader_font
    ws['B9'] = 'Customer Company Name'
    ws['A10'] = 'Contact:'
    ws['A10'].font = subheader_font
    ws['B10'] = 'Contact Person'
    ws['A11'] = 'Phone:'
    ws['A11'].font = subheader_font
    ws['B11'] = '(555) 123-4567'
    ws['A12'] = 'Email:'
    ws['A12'].font = subheader_font
    ws['B12'] = 'contact@customer.com'
    
    # Line items header
    ws['A14'] = 'Item #'
    ws['B14'] = 'Qty'
    ws['C14'] = 'Unit'
    ws['D14'] = 'Description'
    ws['E14'] = 'Unit Price'
    ws['F14'] = 'Total Price'
    
    # Format header
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws[f'{col}14']
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thick_border
    
    # Sample line items
    sample_items = [
        ['1', '2', 'EA', 'Sample Product 1\nModel: ABC-123', '100.00', '200.00'],
        ['2', '1', 'EA', 'Sample Product 2\nModel: XYZ-456', '250.00', '250.00']
    ]
    
    for row_idx, item in enumerate(sample_items, start=15):
        for col_idx, value in enumerate(['A', 'B', 'C', 'D', 'E', 'F']):
            cell = ws[f'{value}{row_idx}']
            cell.value = item[col_idx]
            cell.font = normal_font
            cell.border = thin_border
            if col_idx == 3:  # Description column
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            elif col_idx in [4, 5]:  # Price columns
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '$#,##0.00'
            else:
                cell.alignment = Alignment(horizontal='center')
    
    # Totals
    ws['A18'] = 'Subtotal:'
    ws['A18'].font = subheader_font
    ws['F18'] = '=SUM(F15:F16)'
    ws['F18'].font = normal_font
    ws['F18'].number_format = '$#,##0.00'
    ws['F18'].border = thin_border
    
    ws['A19'] = 'Tax (5%):'
    ws['A19'].font = subheader_font
    ws['F19'] = '=F18*0.05'
    ws['F19'].font = normal_font
    ws['F19'].number_format = '$#,##0.00'
    ws['F19'].border = thin_border
    
    ws['A20'] = 'Freight:'
    ws['A20'].font = subheader_font
    ws['F20'] = '0.00'
    ws['F20'].font = normal_font
    ws['F20'].number_format = '$#,##0.00'
    ws['F20'].border = thin_border
    
    ws['A21'] = 'TOTAL:'
    ws['A21'].font = Font(name='Arial', size=12, bold=True)
    ws['A21'].fill = light_fill
    ws['F21'] = '=F18+F19+F20'
    ws['F21'].font = Font(name='Arial', size=12, bold=True)
    ws['F21'].fill = light_fill
    ws['F21'].number_format = '$#,##0.00'
    ws['F21'].border = thick_border
    
    # Terms
    ws['A23'] = 'TERMS AND CONDITIONS'
    ws['A23'].font = subheader_font
    ws['A23'].fill = light_fill
    ws.merge_cells('A23:F23')
    
    terms = [
        'Payment Terms: Net 30 days',
        'Delivery: FOB Origin',
        'Validity: 30 days from quote date',
        'All prices subject to change without notice',
        'Freight charges not included unless specified'
    ]
    
    for i, term in enumerate(terms, start=24):
        ws[f'A{i}'] = f'• {term}'
        ws[f'A{i}'].font = normal_font

if __name__ == "__main__":
    create_simple_excel_workbook()
