#!/usr/bin/env python3
"""
Simple Quote Generator - Terminal Only
A simple command-line quote generator that works reliably
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import os
import re

class SimpleQuoteGenerator:
    def __init__(self):
        # Quote data
        self.quote_data = {
            'quote_number': f"QUOTE-{datetime.now().strftime('%Y%m%d')}-001",
            'quote_date': datetime.now().strftime('%m/%d/%Y'),
            'customer_company': '',
            'contact_person': '',
            'phone': '',
            'email': '',
            'customer_ref': '',
            'payment_terms': 'Net 30 Days',
            'delivery_terms': 'FOB Origin',
            'line_items': []
        }
    
    def run(self):
        """Run the application"""
        print("\n" + "="*60)
        print("🚀 SIMPLE QUOTE GENERATOR")
        print("="*60)
        
        while True:
            print("\n📋 MENU:")
            print("1. Import file (.xlsx, .rtf, .xml)")
            print("2. Add line item manually")
            print("3. View current quote data")
            print("4. Generate quote")
            print("5. Clear all data")
            print("6. Exit")
            
            choice = input("\nEnter your choice (1-6): ").strip()
            
            if choice == "1":
                self.import_file()
            elif choice == "2":
                self.add_item()
            elif choice == "3":
                self.view_data()
            elif choice == "4":
                self.generate_quote()
            elif choice == "5":
                self.clear_all()
            elif choice == "6":
                print("👋 Goodbye!")
                break
            else:
                print("❌ Invalid choice. Please enter 1-6.")
    
    def import_file(self):
        """Import file"""
        print("\n📁 Available files:")
        files = [f for f in os.listdir('.') if f.endswith(('.xlsx', '.rtf', '.xml'))]
        for i, file in enumerate(files, 1):
            print(f"   {i}. {file}")
        
        if not files:
            print("❌ No .xlsx, .rtf, or .xml files found in current directory")
            return
        
        try:
            choice = int(input(f"\nSelect file (1-{len(files)}): ")) - 1
            if 0 <= choice < len(files):
                file_path = files[choice]
                print(f"📥 Importing {file_path}...")
                
                # Import the file
                if file_path.endswith('.xlsx'):
                    self.import_excel_file(file_path)
                elif file_path.endswith('.rtf'):
                    self.import_rtf_file(file_path)
                elif file_path.endswith('.xml'):
                    self.import_xml_file(file_path)
                
                print(f"✅ File imported successfully!")
            else:
                print("❌ Invalid file selection")
        except ValueError:
            print("❌ Please enter a valid number")
        except Exception as e:
            print(f"❌ Error importing file: {e}")
    
    def import_excel_file(self, file_path):
        """Import Excel file"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            print(f"   📊 Excel file: {ws.title}")
            print(f"   📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Clear existing items
            self.quote_data['line_items'] = []
            
            # Find data
            item_count = 0
            for row in range(2, ws.max_row + 1):  # Skip header
                if item_count >= 10:  # Limit to 10 items
                    break
                    
                # Try to find data in common columns
                if ws.cell(row=row, column=1).value or ws.cell(row=row, column=2).value:
                    description = ws.cell(row=row, column=2).value or ws.cell(row=row, column=3).value or ""
                    quantity = ws.cell(row=row, column=3).value or ws.cell(row=row, column=4).value or "1"
                    unit_price = ws.cell(row=row, column=4).value or ws.cell(row=row, column=5).value or ws.cell(row=row, column=6).value or "0.00"
                    
                    # Clean up the data
                    description = str(description).strip()
                    quantity = str(quantity).strip()
                    unit_price = str(unit_price).strip()
                    
                    # Skip if description contains header-like text
                    if any(header in description.lower() for header in ['customer:', 'quantity:', 'unit:', 'price:', 'product:']):
                        continue
                    
                    # Clean quantity - extract only numbers
                    try:
                        qty_match = re.search(r'(\d+\.?\d*)', quantity)
                        if qty_match:
                            quantity = qty_match.group(1)
                        else:
                            quantity = "1"
                    except:
                        quantity = "1"
                    
                    # Clean unit price - extract only numbers
                    try:
                        price_match = re.search(r'(\d+\.?\d*)', unit_price)
                        if price_match:
                            unit_price = price_match.group(1)
                        else:
                            unit_price = "0.00"
                    except:
                        unit_price = "0.00"
                    
                    if description and len(description) > 3:  # Only add meaningful descriptions
                        self.quote_data['line_items'].append({
                            'description': description,
                            'quantity': quantity,
                            'unit': 'EA',
                            'unit_price': unit_price
                        })
                        item_count += 1
                        print(f"   ✅ Added: {description} - Qty: {quantity} - Price: ${unit_price}")
            
            print(f"   📦 Total items imported: {item_count}")
            
        except Exception as e:
            print(f"   ❌ Error reading Excel file: {e}")
    
    def import_rtf_file(self, file_path):
        """Import RTF file"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            print(f"   📄 RTF file size: {len(content)} characters")
            
            # Clear existing items
            self.quote_data['line_items'] = []
            
            # Basic parsing for RTF
            lines = content.split('\n')
            item_count = 0
            
            for line in lines:
                line = line.strip()
                if len(line) > 10 and 'product' in line.lower() and item_count < 10:
                    self.quote_data['line_items'].append({
                        'description': line,
                        'quantity': '1',
                        'unit': 'EA',
                        'unit_price': '0.00'
                    })
                    item_count += 1
                    print(f"   ✅ Added: {line}")
            
            print(f"   📦 Total items imported: {item_count}")
            print("   ⚠️  Note: RTF parsing is basic - you may need to adjust quantities and prices manually")
            
        except Exception as e:
            print(f"   ❌ Error reading RTF file: {e}")
    
    def import_xml_file(self, file_path):
        """Import XML file"""
        try:
            import xml.etree.ElementTree as ET
            
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            print(f"   📄 XML root tag: {root.tag}")
            
            # Clear existing items
            self.quote_data['line_items'] = []
            
            # Look for product nodes
            products = root.findall('.//product') or root.findall('.//item') or root.findall('.//line')
            print(f"   📦 Found {len(products)} product/item nodes")
            
            item_count = 0
            for product in products:
                if item_count >= 10:
                    break
                    
                description = ""
                quantity = "1"
                unit_price = "0.00"
                
                # Try to get description
                desc_elem = product.find('description') or product.find('name') or product.find('title')
                if desc_elem is not None:
                    description = desc_elem.text or ""
                
                # Try to get quantity
                qty_elem = product.find('quantity') or product.find('qty')
                if qty_elem is not None:
                    quantity = qty_elem.text or "1"
                
                # Try to get price
                price_elem = product.find('price') or product.find('unitprice') or product.find('cost')
                if price_elem is not None:
                    unit_price = price_elem.text or "0.00"
                
                if description:
                    self.quote_data['line_items'].append({
                        'description': description,
                        'quantity': quantity,
                        'unit': 'EA',
                        'unit_price': unit_price
                    })
                    item_count += 1
                    print(f"   ✅ Added: {description} - Qty: {quantity} - Price: ${unit_price}")
            
            print(f"   📦 Total items imported: {item_count}")
            
        except Exception as e:
            print(f"   ❌ Error reading XML file: {e}")
    
    def add_item(self):
        """Add line item"""
        print("\n➕ ADD LINE ITEM:")
        description = input("Description: ").strip()
        if not description:
            print("❌ Description cannot be empty")
            return
        
        quantity = input("Quantity (default: 1): ").strip() or "1"
        unit = input("Unit (default: EA): ").strip() or "EA"
        unit_price = input("Unit Price (default: 0.00): ").strip() or "0.00"
        
        # Add to data
        self.quote_data['line_items'].append({
            'description': description,
            'quantity': quantity,
            'unit': unit,
            'unit_price': unit_price
        })
        
        print(f"✅ Added: {description} - Qty: {quantity} - Price: ${unit_price}")
    
    def view_data(self):
        """View current data"""
        print("\n📊 CURRENT QUOTE DATA:")
        print(f"Quote Number: {self.quote_data['quote_number']}")
        print(f"Quote Date: {self.quote_data['quote_date']}")
        print(f"Customer Company: {self.quote_data['customer_company']}")
        print(f"Contact Person: {self.quote_data['contact_person']}")
        print(f"Phone: {self.quote_data['phone']}")
        print(f"Email: {self.quote_data['email']}")
        print(f"Customer Reference: {self.quote_data['customer_ref']}")
        print(f"Payment Terms: {self.quote_data['payment_terms']}")
        print(f"Delivery Terms: {self.quote_data['delivery_terms']}")
        
        print(f"\n📦 LINE ITEMS ({len(self.quote_data['line_items'])}):")
        if self.quote_data['line_items']:
            for i, item in enumerate(self.quote_data['line_items'], 1):
                print(f"   {i}. {item['description']} - Qty: {item['quantity']} {item['unit']} - ${item['unit_price']}")
        else:
            print("   No items added yet")
    
    def generate_quote(self):
        """Generate quote"""
        if not self.quote_data['line_items']:
            print("❌ No line items to generate quote. Add some items first.")
            return
        
        print("\n🔄 Generating quote...")
        try:
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Quote"
            
            # Define styles
            header_font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
            subheader_font = Font(name='Arial', size=12, bold=True)
            normal_font = Font(name='Arial', size=10)
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
            thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), 
                                top=Side(style='thick'), bottom=Side(style='thick'))
            
            # Set column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 6
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            
            # Header
            ws['A1'] = 'QUOTE'
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:F1')
            ws.row_dimensions[1].height = 40
            
            # Company info
            ws['A2'] = 'KINDER MORGAN'
            ws['A2'].font = Font(name='Arial', size=12, bold=True)
            ws['A3'] = '1234 Energy Drive'
            ws['A4'] = 'Houston, TX 77002'
            ws['A5'] = 'Phone: (713) 369-9000'
            ws['A6'] = 'Email: quotes@kindermorgan.com'
            
            # Quote details
            ws['D2'] = 'Quote No.:'
            ws['D2'].font = subheader_font
            ws['E2'] = self.quote_data['quote_number']
            ws['D3'] = 'Date:'
            ws['D3'].font = subheader_font
            ws['E3'] = self.quote_data['quote_date']
            ws['D4'] = 'Customer Ref:'
            ws['D4'].font = subheader_font
            ws['E4'] = self.quote_data['customer_ref']
            
            # Customer info
            ws['A8'] = 'CUSTOMER INFORMATION'
            ws['A8'].font = subheader_font
            ws['A8'].fill = light_fill
            ws.merge_cells('A8:F8')
            
            ws['A9'] = 'Company:'
            ws['A9'].font = subheader_font
            ws['B9'] = self.quote_data['customer_company']
            ws['A10'] = 'Contact:'
            ws['A10'].font = subheader_font
            ws['B10'] = self.quote_data['contact_person']
            ws['A11'] = 'Phone:'
            ws['A11'].font = subheader_font
            ws['B11'] = self.quote_data['phone']
            ws['A12'] = 'Email:'
            ws['A12'].font = subheader_font
            ws['B12'] = self.quote_data['email']
            
            # Line items header
            ws['A14'] = 'Item #'
            ws['B14'] = 'Qty'
            ws['C14'] = 'Unit'
            ws['D14'] = 'Description'
            ws['E14'] = 'Unit Price'
            ws['F14'] = 'Total Price'
            
            # Format header
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}14'].font = subheader_font
                ws[f'{col}14'].fill = header_fill
                ws[f'{col}14'].alignment = Alignment(horizontal='center')
                ws[f'{col}14'].border = thick_border
            
            # Add line items
            last_row = 14
            subtotal = 0
            
            for i, item in enumerate(self.quote_data['line_items'], 1):
                last_row += 1
                
                # Safely convert quantity and unit price to float
                try:
                    quantity = float(str(item['quantity']).replace(',', '')) if item['quantity'] else 0
                except (ValueError, TypeError):
                    quantity = 0
                
                try:
                    unit_price = float(str(item['unit_price']).replace(',', '')) if item['unit_price'] else 0
                except (ValueError, TypeError):
                    unit_price = 0
                
                total_price = quantity * unit_price
                subtotal += total_price
                
                ws[f'A{last_row}'] = i
                ws[f'B{last_row}'] = item['quantity']
                ws[f'C{last_row}'] = item['unit']
                ws[f'D{last_row}'] = item['description']
                ws[f'E{last_row}'] = unit_price
                ws[f'F{last_row}'] = total_price
                
                # Format price columns
                ws[f'E{last_row}'].number_format = '$#,##0.00'
                ws[f'F{last_row}'].number_format = '$#,##0.00'
                
                # Add borders
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{last_row}'].border = thin_border
            
            # Totals
            tax_rate = 0.05
            tax_amount = subtotal * tax_rate
            freight = 0.00
            total = subtotal + tax_amount + freight
            
            ws[f'A{last_row + 2}'] = 'Subtotal:'
            ws[f'A{last_row + 2}'].font = subheader_font
            ws[f'F{last_row + 2}'] = subtotal
            ws[f'F{last_row + 2}'].number_format = '$#,##0.00'
            
            ws[f'A{last_row + 3}'] = 'Tax (5%):'
            ws[f'A{last_row + 3}'].font = subheader_font
            ws[f'F{last_row + 3}'] = tax_amount
            ws[f'F{last_row + 3}'].number_format = '$#,##0.00'
            
            ws[f'A{last_row + 4}'] = 'Freight:'
            ws[f'A{last_row + 4}'].font = subheader_font
            ws[f'F{last_row + 4}'] = freight
            ws[f'F{last_row + 4}'].number_format = '$#,##0.00'
            
            ws[f'A{last_row + 5}'] = 'TOTAL:'
            ws[f'A{last_row + 5}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'A{last_row + 5}'].fill = light_fill
            ws[f'F{last_row + 5}'] = total
            ws[f'F{last_row + 5}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'F{last_row + 5}'].fill = light_fill
            ws[f'F{last_row + 5}'].number_format = '$#,##0.00'
            
            # Add borders to totals
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{last_row + 5}'].border = thick_border
            
            # Terms
            ws[f'A{last_row + 7}'] = 'TERMS AND CONDITIONS'
            ws[f'A{last_row + 7}'].font = subheader_font
            ws[f'A{last_row + 7}'].fill = light_fill
            ws.merge_cells(f'A{last_row + 7}:F{last_row + 7}')
            
            ws[f'A{last_row + 8}'] = f"Payment Terms: {self.quote_data['payment_terms']}"
            ws[f'A{last_row + 9}'] = f"Delivery: {self.quote_data['delivery_terms']}"
            ws[f'A{last_row + 10}'] = "Validity: 30 days from quote date"
            ws[f'A{last_row + 11}'] = "All prices subject to change without notice"
            ws[f'A{last_row + 12}'] = "Freight charges not included unless specified"
            
            # Auto-fit columns
            for col in range(1, 7):  # Columns A through F
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col)
                for row in range(1, last_row + 15):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            filename = f"quote_{self.quote_data['quote_number'].replace('/', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            print(f"✅ Quote generated successfully!")
            print(f"📁 Saved as: {filename}")
            print(f"💰 Subtotal: ${subtotal:,.2f}")
            print(f"💰 Tax (5%): ${tax_amount:,.2f}")
            print(f"💰 Total: ${total:,.2f}")
            print(f"📦 Items: {len(self.quote_data['line_items'])}")
            
        except Exception as e:
            print(f"❌ Error generating quote: {e}")
    
    def clear_all(self):
        """Clear all data"""
        self.quote_data['line_items'] = []
        self.quote_data['quote_number'] = f"QUOTE-{datetime.now().strftime('%Y%m%d')}-001"
        self.quote_data['quote_date'] = datetime.now().strftime('%m/%d/%Y')
        self.quote_data['customer_company'] = ''
        self.quote_data['contact_person'] = ''
        self.quote_data['phone'] = ''
        self.quote_data['email'] = ''
        self.quote_data['customer_ref'] = ''
        self.quote_data['payment_terms'] = 'Net 30 Days'
        self.quote_data['delivery_terms'] = 'FOB Origin'
        
        print("✅ All data cleared!")

if __name__ == "__main__":
    app = SimpleQuoteGenerator()
    app.run()
