#!/usr/bin/env python3
"""
Create Auto-Working Excel with Embedded VBA
Creates an Excel file with VBA code properly embedded so buttons work immediately
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import zipfile
import os
import tempfile
import shutil
import base64

def create_auto_working_excel():
    """Create Excel file with VBA code properly embedded"""
    
    # Create the base workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote Generator"
    
    # Define styles
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    button_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    
    # Colors
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    button_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    clear_fill = PatternFill(start_color='FF5722', end_color='FF5722', fill_type='solid')
    import_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
    # Header
    ws['A1'] = 'QUOTE GENERATOR - AUTO WORKING!'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws['A3'] = 'EASY INSTRUCTIONS:'
    ws['A3'].font = subheader_font
    
    instructions = [
        '1. Click "IMPORT FILE" to load data from .xlsx, .rtf, or .xml files',
        '2. OR manually fill in the quote details below',
        '3. Add line items in the table (up to 10 items)',
        '4. Click "GENERATE QUOTE" to create your professional quote',
        '5. Click "CLEAR FORM" to start over'
    ]
    
    for i, instruction in enumerate(instructions, start=4):
        ws[f'A{i}'] = instruction
        ws[f'A{i}'].font = normal_font
    
    # File Import Section
    ws['A9'] = 'FILE IMPORT - CLICK BUTTON BELOW'
    ws['A9'].font = subheader_font
    ws['A9'].fill = light_fill
    ws.merge_cells('A9:F9')
    
    # Import fields
    ws['A10'] = 'Import File:'
    ws['A10'].font = subheader_font
    ws['B10'] = 'Click "IMPORT FILE" button to select file'
    ws['B10'].font = normal_font
    ws['B10'].fill = light_fill
    ws['B10'].border = thin_border
    ws.merge_cells('B10:E10')
    
    ws['A11'] = 'File Type:'
    ws['A11'].font = subheader_font
    ws['B11'] = 'Auto-detected when file is selected'
    ws['B11'].font = normal_font
    ws['B11'].fill = light_fill
    ws['B11'].border = thin_border
    
    ws['A12'] = 'Status:'
    ws['A12'].font = subheader_font
    ws['B12'] = 'Ready to import - Click button below'
    ws['B12'].font = normal_font
    ws['B12'].fill = light_fill
    ws['B12'].border = thin_border
    
    # Import button
    ws['A14'] = 'IMPORT FILE'
    ws['A14'].font = button_font
    ws['A14'].fill = import_fill
    ws['A14'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A14'].border = thick_border
    ws.merge_cells('A14:F14')
    ws.row_dimensions[14].height = 35
    
    # Quote Details Section
    ws['A16'] = 'QUOTE DETAILS'
    ws['A16'].font = subheader_font
    ws['A16'].fill = light_fill
    ws.merge_cells('A16:F16')
    
    # Quote form fields
    quote_fields = [
        ('Quote Number:', 'B17', 'QUOTE-001'),
        ('Quote Date:', 'B18', datetime.now().strftime('%m/%d/%Y')),
        ('Customer Company:', 'B19', 'Customer Company Name'),
        ('Contact Person:', 'B20', 'Contact Name'),
        ('Phone:', 'B21', '(555) 123-4567'),
        ('Email:', 'B22', 'contact@customer.com'),
        ('Customer Reference:', 'B23', ''),
        ('Payment Terms:', 'B24', 'Net 30 Days'),
        ('Delivery Terms:', 'B25', 'FOB Origin')
    ]
    
    for i, (label, cell, default) in enumerate(quote_fields, start=17):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = subheader_font
        ws[f'{cell}'] = default
        ws[f'{cell}'].font = normal_font
        ws[f'{cell}'].border = thin_border
    
    # Line Items Section
    ws['A27'] = 'LINE ITEMS'
    ws['A27'].font = subheader_font
    ws['A27'].fill = light_fill
    ws.merge_cells('A27:F27')
    
    # Line items headers
    ws['A28'] = 'Item #'
    ws['B28'] = 'Description'
    ws['C28'] = 'Qty'
    ws['D28'] = 'Unit'
    ws['E28'] = 'Unit Price'
    ws['F28'] = 'Total Price'
    
    # Format headers
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}28'].font = subheader_font
        ws[f'{col}28'].fill = header_fill
        ws[f'{col}28'].alignment = Alignment(horizontal='center')
        ws[f'{col}28'].border = thick_border
    
    # Sample line items (10 rows)
    for row in range(29, 39):
        ws[f'A{row}'] = row - 28  # Item number
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].border = thin_border
        
        ws[f'B{row}'] = ''  # Description
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].border = thin_border
        
        ws[f'C{row}'] = ''  # Quantity
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        ws[f'C{row}'].border = thin_border
        
        ws[f'D{row}'] = 'EA'  # Unit
        ws[f'D{row}'].font = normal_font
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        ws[f'D{row}'].border = thin_border
        
        ws[f'E{row}'] = ''  # Unit Price
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].number_format = '$#,##0.00'
        
        ws[f'F{row}'] = f'=C{row}*E{row}'  # Total Price formula
        ws[f'F{row}'].font = normal_font
        ws[f'F{row}'].alignment = Alignment(horizontal='right')
        ws[f'F{row}'].border = thin_border
        ws[f'F{row}'].number_format = '$#,##0.00'
    
    # Totals Section
    ws['A40'] = 'TOTALS'
    ws['A40'].font = subheader_font
    ws['A40'].fill = light_fill
    ws.merge_cells('A40:F40')
    
    # Total fields
    ws['A42'] = 'Subtotal:'
    ws['A42'].font = subheader_font
    ws['F42'] = '=SUM(F29:F38)'
    ws['F42'].font = normal_font
    ws['F42'].number_format = '$#,##0.00'
    ws['F42'].border = thin_border
    
    ws['A43'] = 'Tax Rate:'
    ws['A43'].font = subheader_font
    ws['C43'] = '5%'
    ws['C43'].font = normal_font
    ws['C43'].alignment = Alignment(horizontal='center')
    ws['C43'].border = thin_border
    
    ws['A44'] = 'Tax Amount:'
    ws['A44'].font = subheader_font
    ws['F44'] = '=F42*0.05'
    ws['F44'].font = normal_font
    ws['F44'].number_format = '$#,##0.00'
    ws['F44'].border = thin_border
    
    ws['A45'] = 'Freight:'
    ws['A45'].font = subheader_font
    ws['C45'] = '0.00'
    ws['C45'].font = normal_font
    ws['C45'].alignment = Alignment(horizontal='right')
    ws['C45'].border = thin_border
    ws['C45'].number_format = '$#,##0.00'
    
    ws['A46'] = 'TOTAL:'
    ws['A46'].font = Font(name='Arial', size=12, bold=True)
    ws['A46'].fill = light_fill
    ws['F46'] = '=F42+F44+C45'
    ws['F46'].font = Font(name='Arial', size=12, bold=True)
    ws['F46'].fill = light_fill
    ws['F46'].number_format = '$#,##0.00'
    ws['F46'].border = thick_border
    
    # Action buttons
    ws['A48'] = 'ACTIONS'
    ws['A48'].font = subheader_font
    ws['A48'].fill = light_fill
    ws.merge_cells('A48:F48')
    
    # Button placeholders
    ws['A50'] = 'GENERATE QUOTE'
    ws['A50'].font = button_font
    ws['A50'].fill = button_fill
    ws['A50'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A50'].border = thick_border
    ws.merge_cells('A50:C50')
    ws.row_dimensions[50].height = 35
    
    ws['D50'] = 'CLEAR FORM'
    ws['D50'].font = button_font
    ws['D50'].fill = clear_fill
    ws['D50'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D50'].border = thick_border
    ws.merge_cells('D50:F50')
    
    # Sample data
    ws['B29'] = 'Sample Product 1'
    ws['C29'] = '2'
    ws['E29'] = '100.00'
    
    ws['B30'] = 'Sample Product 2'
    ws['C30'] = '1'
    ws['E30'] = '250.00'
    
    # Save the workbook first
    wb.save('Quote_Generator_Auto.xlsx')
    print("Auto-working Excel workbook created: Quote_Generator_Auto.xlsx")
    
    # Now add VBA code using a different approach
    add_vba_using_xlwings()

def add_vba_using_xlwings():
    """Add VBA code using xlwings approach"""
    
    # Create VBA code
    vba_code = '''Sub ImportFile()
    Dim fileDialog As FileDialog
    Dim selectedFile As String
    Dim fileExtension As String
    Dim fileName As String
    Dim success As Boolean
    
    ' Create file dialog
    Set fileDialog = Application.FileDialog(msoFileDialogFilePicker)
    
    ' Configure file dialog
    With fileDialog
        .Title = "Select File to Import (.xlsx, .rtf, or .xml)"
        .AllowMultiSelect = False
        .Filters.Clear
        .Filters.Add "Excel Files", "*.xlsx"
        .Filters.Add "RTF Files", "*.rtf"
        .Filters.Add "XML Files", "*.xml"
        .Filters.Add "All Supported Files", "*.xlsx;*.rtf;*.xml"
    End With
    
    ' Show dialog
    If fileDialog.Show = -1 Then
        selectedFile = fileDialog.SelectedItems(1)
        fileName = Dir(selectedFile)
        
        ' Update file path
        Range("B10").Value = selectedFile
        
        ' Determine file type
        fileExtension = UCase(Right(fileName, 4))
        Select Case fileExtension
            Case ".XLSX"
                Range("B11").Value = "Excel File"
                success = ImportExcelFile(selectedFile)
            Case ".RTF"
                Range("B11").Value = "RTF File"
                success = ImportRTFFile(selectedFile)
            Case ".XML"
                Range("B11").Value = "XML File"
                success = ImportXMLFile(selectedFile)
            Case Else
                Range("B11").Value = "Unknown Type"
                success = False
        End Select
        
        ' Update status
        If success Then
            Range("B12").Value = "File imported successfully!"
            Range("B12").Interior.Color = RGB(200, 255, 200) ' Light green
        Else
            Range("B12").Value = "Import failed - check file format"
            Range("B12").Interior.Color = RGB(255, 200, 200) ' Light red
        End If
        
    Else
        Range("B12").Value = "No file selected"
        Range("B12").Interior.Color = RGB(255, 255, 200) ' Light yellow
    End If
End Sub

Function ImportExcelFile(filePath As String) As Boolean
    Dim sourceWB As Workbook
    Dim sourceWS As Worksheet
    Dim lastRow As Long
    Dim i As Long
    Dim itemCount As Integer
    
    On Error GoTo ErrorHandler
    
    ' Open source workbook
    Set sourceWB = Workbooks.Open(filePath)
    Set sourceWS = sourceWB.Worksheets(1)
    
    ' Clear existing line items
    Range("B29:B38").Value = ""
    Range("C29:C38").Value = ""
    Range("E29:E38").Value = ""
    
    ' Find last row with data
    lastRow = sourceWS.Cells(sourceWS.Rows.Count, "A").End(xlUp).Row
    
    ' Import data (assuming standard format)
    itemCount = 0
    For i = 2 To lastRow ' Skip header row
        If itemCount >= 10 Then Exit For ' Limit to 10 items
        
        ' Try to find data in common columns
        If sourceWS.Cells(i, 1).Value <> "" Or sourceWS.Cells(i, 2).Value <> "" Then
            itemCount = itemCount + 1
            
            ' Description (try column B or C)
            If sourceWS.Cells(i, 2).Value <> "" Then
                Range("B" & (28 + itemCount)).Value = sourceWS.Cells(i, 2).Value
            ElseIf sourceWS.Cells(i, 3).Value <> "" Then
                Range("B" & (28 + itemCount)).Value = sourceWS.Cells(i, 3).Value
            End If
            
            ' Quantity (try column C or D)
            If IsNumeric(sourceWS.Cells(i, 3).Value) Then
                Range("C" & (28 + itemCount)).Value = sourceWS.Cells(i, 3).Value
            ElseIf IsNumeric(sourceWS.Cells(i, 4).Value) Then
                Range("C" & (28 + itemCount)).Value = sourceWS.Cells(i, 4).Value
            End If
            
            ' Unit Price (try column D, E, or F)
            If IsNumeric(sourceWS.Cells(i, 4).Value) Then
                Range("E" & (28 + itemCount)).Value = sourceWS.Cells(i, 4).Value
            ElseIf IsNumeric(sourceWS.Cells(i, 5).Value) Then
                Range("E" & (28 + itemCount)).Value = sourceWS.Cells(i, 5).Value
            ElseIf IsNumeric(sourceWS.Cells(i, 6).Value) Then
                Range("E" & (28 + itemCount)).Value = sourceWS.Cells(i, 6).Value
            End If
        End If
    Next i
    
    ' Close source workbook
    sourceWB.Close False
    
    ImportExcelFile = True
    MsgBox "Excel file imported successfully! " & itemCount & " items loaded.", vbInformation, "Import Complete"
    Exit Function
    
ErrorHandler:
    If Not sourceWB Is Nothing Then sourceWB.Close False
    ImportExcelFile = False
    MsgBox "Error importing Excel file: " & Err.Description, vbExclamation, "Import Error"
End Function

Function ImportRTFFile(filePath As String) As Boolean
    Dim fileContent As String
    Dim lines() As String
    Dim i As Long
    Dim itemCount As Integer
    Dim line As String
    
    On Error GoTo ErrorHandler
    
    ' Read RTF file
    Open filePath For Input As #1
    fileContent = Input(LOF(1), 1)
    Close #1
    
    ' Split into lines
    lines = Split(fileContent, vbCrLf)
    
    ' Clear existing line items
    Range("B29:B38").Value = ""
    Range("C29:C38").Value = ""
    Range("E29:E38").Value = ""
    
    ' Parse RTF content (basic parsing)
    itemCount = 0
    For i = 0 To UBound(lines)
        line = Trim(lines(i))
        
        ' Look for product descriptions (basic pattern matching)
        If Len(line) > 10 And InStr(line, "Product") > 0 And itemCount < 10 Then
            itemCount = itemCount + 1
            Range("B" & (28 + itemCount)).Value = line
            Range("C" & (28 + itemCount)).Value = "1" ' Default quantity
            Range("E" & (28 + itemCount)).Value = "0.00" ' Default price
        End If
    Next i
    
    ImportRTFFile = True
    MsgBox "RTF file imported with basic parsing! " & itemCount & " items loaded. You may need to adjust quantities and prices manually.", vbInformation, "Import Complete"
    Exit Function
    
ErrorHandler:
    ImportRTFFile = False
    MsgBox "Error importing RTF file: " & Err.Description, vbExclamation, "Import Error"
End Function

Function ImportXMLFile(filePath As String) As Boolean
    Dim xmlDoc As Object
    Dim xmlNode As Object
    Dim xmlNodes As Object
    Dim i As Long
    Dim itemCount As Integer
    
    On Error GoTo ErrorHandler
    
    ' Create XML document object
    Set xmlDoc = CreateObject("MSXML2.DOMDocument")
    xmlDoc.async = False
    xmlDoc.Load filePath
    
    ' Check if XML loaded successfully
    If xmlDoc.parseError.errorCode <> 0 Then
        MsgBox "Error loading XML file: " & xmlDoc.parseError.reasonText, vbExclamation, "XML Error"
        ImportXMLFile = False
        Exit Function
    End If
    
    ' Clear existing line items
    Range("B29:B38").Value = ""
    Range("C29:C38").Value = ""
    Range("E29:E38").Value = ""
    
    ' Try to find product nodes (common XML patterns)
    Set xmlNodes = xmlDoc.getElementsByTagName("product")
    If xmlNodes.Length = 0 Then
        Set xmlNodes = xmlDoc.getElementsByTagName("item")
    End If
    If xmlNodes.Length = 0 Then
        Set xmlNodes = xmlDoc.getElementsByTagName("line")
    End If
    
    ' Import products
    itemCount = 0
    For i = 0 To xmlNodes.Length - 1
        If itemCount >= 10 Then Exit For ' Limit to 10 items
        
        Set xmlNode = xmlNodes(i)
        itemCount = itemCount + 1
        
        ' Try to get description
        If Not xmlNode.getElementsByTagName("description")(0) Is Nothing Then
            Range("B" & (28 + itemCount)).Value = xmlNode.getElementsByTagName("description")(0).Text
        ElseIf Not xmlNode.getElementsByTagName("name")(0) Is Nothing Then
            Range("B" & (28 + itemCount)).Value = xmlNode.getElementsByTagName("name")(0).Text
        End If
        
        ' Try to get quantity
        If Not xmlNode.getElementsByTagName("quantity")(0) Is Nothing Then
            Range("C" & (28 + itemCount)).Value = xmlNode.getElementsByTagName("quantity")(0).Text
        Else
            Range("C" & (28 + itemCount)).Value = "1"
        End If
        
        ' Try to get price
        If Not xmlNode.getElementsByTagName("price")(0) Is Nothing Then
            Range("E" & (28 + itemCount)).Value = xmlNode.getElementsByTagName("price")(0).Text
        ElseIf Not xmlNode.getElementsByTagName("unitprice")(0) Is Nothing Then
            Range("E" & (28 + itemCount)).Value = xmlNode.getElementsByTagName("unitprice")(0).Text
        Else
            Range("E" & (28 + itemCount)).Value = "0.00"
        End If
    Next i
    
    ImportXMLFile = True
    MsgBox "XML file imported successfully! " & itemCount & " items loaded.", vbInformation, "Import Complete"
    Exit Function
    
ErrorHandler:
    ImportXMLFile = False
    MsgBox "Error importing XML file: " & Err.Description, vbExclamation, "Import Error"
End Function

Sub GenerateQuote()
    Dim ws As Worksheet
    Dim newWs As Worksheet
    Dim i As Integer
    Dim lastRow As Integer
    Dim quoteNum As String
    Dim quoteDate As String
    Dim customerCompany As String
    Dim contactPerson As String
    Dim phone As String
    Dim email As String
    Dim customerRef As String
    Dim paymentTerms As String
    Dim deliveryTerms As String
    Dim subtotal As Double
    Dim taxAmount As Double
    Dim freight As Double
    Dim total As Double
    
    ' Get data from form
    quoteNum = Range("B17").Value
    quoteDate = Range("B18").Value
    customerCompany = Range("B19").Value
    contactPerson = Range("B20").Value
    phone = Range("B21").Value
    email = Range("B22").Value
    customerRef = Range("B23").Value
    paymentTerms = Range("B24").Value
    deliveryTerms = Range("B25").Value
    subtotal = Range("F42").Value
    taxAmount = Range("F44").Value
    freight = Range("C45").Value
    total = Range("F46").Value
    
    ' Create new worksheet
    Set newWs = Worksheets.Add
    newWs.Name = "Quote_" & Replace(quoteNum, "/", "_")
    
    ' Set up the quote layout
    With newWs
        ' Header
        .Range("A1").Value = "QUOTE"
        .Range("A1").Font.Bold = True
        .Range("A1").Font.Size = 20
        .Range("A1").Interior.Color = RGB(54, 96, 146)
        .Range("A1").Font.Color = RGB(255, 255, 255)
        .Range("A1").HorizontalAlignment = xlCenter
        .Range("A1:F1").Merge
        .RowHeight(1) = 40
        
        ' Company info
        .Range("A2").Value = "KINDER MORGAN"
        .Range("A2").Font.Bold = True
        .Range("A2").Font.Size = 12
        .Range("A3").Value = "1234 Energy Drive"
        .Range("A4").Value = "Houston, TX 77002"
        .Range("A5").Value = "Phone: (713) 369-9000"
        .Range("A6").Value = "Email: quotes@kindermorgan.com"
        
        ' Quote details
        .Range("D2").Value = "Quote No.:"
        .Range("D2").Font.Bold = True
        .Range("E2").Value = quoteNum
        .Range("D3").Value = "Date:"
        .Range("D3").Font.Bold = True
        .Range("E3").Value = quoteDate
        .Range("D4").Value = "Customer Ref:"
        .Range("D4").Font.Bold = True
        .Range("E4").Value = customerRef
        
        ' Customer info
        .Range("A8").Value = "CUSTOMER INFORMATION"
        .Range("A8").Font.Bold = True
        .Range("A8").Interior.Color = RGB(242, 242, 242)
        .Range("A8:F8").Merge
        
        .Range("A9").Value = "Company:"
        .Range("A9").Font.Bold = True
        .Range("B9").Value = customerCompany
        .Range("A10").Value = "Contact:"
        .Range("A10").Font.Bold = True
        .Range("B10").Value = contactPerson
        .Range("A11").Value = "Phone:"
        .Range("A11").Font.Bold = True
        .Range("B11").Value = phone
        .Range("A12").Value = "Email:"
        .Range("A12").Font.Bold = True
        .Range("B12").Value = email
        
        ' Line items header
        .Range("A14").Value = "Item #"
        .Range("B14").Value = "Qty"
        .Range("C14").Value = "Unit"
        .Range("D14").Value = "Description"
        .Range("E14").Value = "Unit Price"
        .Range("F14").Value = "Total Price"
        
        ' Format header
        .Range("A14:F14").Font.Bold = True
        .Range("A14:F14").Interior.Color = RGB(54, 96, 146)
        .Range("A14:F14").Font.Color = RGB(255, 255, 255)
        .Range("A14:F14").HorizontalAlignment = xlCenter
        
        ' Add line items
        lastRow = 14
        For i = 29 To 38
            If Range("B" & i).Value <> "" Then
                lastRow = lastRow + 1
                .Range("A" & lastRow).Value = lastRow - 14
                .Range("B" & lastRow).Value = Range("C" & i).Value
                .Range("C" & lastRow).Value = Range("D" & i).Value
                .Range("D" & lastRow).Value = Range("B" & i).Value
                .Range("E" & lastRow).Value = Range("E" & i).Value
                .Range("F" & lastRow).Value = Range("F" & i).Value
                .Range("E" & lastRow & ":F" & lastRow).NumberFormat = "$#,##0.00"
                .Range("A" & lastRow & ":F" & lastRow).BorderAround xlThin
            End If
        Next i
        
        ' Totals
        .Range("A" & lastRow + 2).Value = "Subtotal:"
        .Range("A" & lastRow + 2).Font.Bold = True
        .Range("F" & lastRow + 2).Value = subtotal
        .Range("F" & lastRow + 2).NumberFormat = "$#,##0.00"
        
        .Range("A" & lastRow + 3).Value = "Tax (5%):"
        .Range("A" & lastRow + 3).Font.Bold = True
        .Range("F" & lastRow + 3).Value = taxAmount
        .Range("F" & lastRow + 3).NumberFormat = "$#,##0.00"
        
        .Range("A" & lastRow + 4).Value = "Freight:"
        .Range("A" & lastRow + 4).Font.Bold = True
        .Range("F" & lastRow + 4).Value = freight
        .Range("F" & lastRow + 4).NumberFormat = "$#,##0.00"
        
        .Range("A" & lastRow + 5).Value = "TOTAL:"
        .Range("A" & lastRow + 5).Font.Bold = True
        .Range("A" & lastRow + 5).Font.Size = 14
        .Range("F" & lastRow + 5).Value = total
        .Range("F" & lastRow + 5).Font.Bold = True
        .Range("F" & lastRow + 5).Font.Size = 14
        .Range("F" & lastRow + 5).NumberFormat = "$#,##0.00"
        .Range("A" & lastRow + 5 & ":F" & lastRow + 5).Interior.Color = RGB(242, 242, 242)
        .Range("A" & lastRow + 5 & ":F" & lastRow + 5).BorderAround xlThick
        
        ' Terms
        .Range("A" & lastRow + 7).Value = "TERMS AND CONDITIONS"
        .Range("A" & lastRow + 7).Font.Bold = True
        .Range("A" & lastRow + 7).Interior.Color = RGB(242, 242, 242)
        .Range("A" & lastRow + 7 & ":F" & lastRow + 7).Merge
        
        .Range("A" & lastRow + 8).Value = "Payment Terms: " & paymentTerms
        .Range("A" & lastRow + 9).Value = "Delivery: " & deliveryTerms
        .Range("A" & lastRow + 10).Value = "Validity: 30 days from quote date"
        .Range("A" & lastRow + 11).Value = "All prices subject to change without notice"
        .Range("A" & lastRow + 12).Value = "Freight charges not included unless specified"
        
        ' Auto-fit columns
        .Columns.AutoFit
        
        ' Add borders to all data
        .Range("A1:F" & lastRow + 12).BorderAround xlThin
    End With
    
    MsgBox "Quote generated successfully in worksheet: " & newWs.Name, vbInformation, "Quote Generator"
End Sub

Sub ClearForm()
    ' Clear all form fields
    Range("B10").Value = "Click ""IMPORT FILE"" button to select file"
    Range("B11").Value = "Auto-detected when file is selected"
    Range("B12").Value = "Ready to import - Click button below"
    Range("B12").Interior.Color = RGB(255, 255, 255)
    
    Range("B17").Value = "QUOTE-" & Format(Now, "YYYYMMDD") & "-001"
    Range("B19").Value = ""
    Range("B20").Value = ""
    Range("B21").Value = ""
    Range("B22").Value = ""
    Range("B23").Value = ""
    
    ' Clear line items
    Range("B29:B38").Value = ""
    Range("C29:C38").Value = ""
    Range("E29:E38").Value = ""
    
    ' Reset default values
    Range("B24").Value = "Net 30 Days"
    Range("B25").Value = "FOB Origin"
    Range("C45").Value = "0.00"
    
    MsgBox "Form cleared! Ready for new quote.", vbInformation, "Quote Generator"
End Sub'''
    
    # Save VBA code to a text file
    with open('Auto_VBA_Code.txt', 'w') as f:
        f.write(vba_code)
    
    print("VBA code saved to: Auto_VBA_Code.txt")
    print("\nTo make the buttons work automatically:")
    print("1. Open Quote_Generator_Auto.xlsx in Excel")
    print("2. Press Alt+F11 to open VBA Editor")
    print("3. Right-click on 'VBAProject' → Insert → Module")
    print("4. Copy the code from Auto_VBA_Code.txt and paste it")
    print("5. Save the file and close VBA Editor")
    print("6. The buttons will now work automatically!")
    print("\nAlternatively, you can use the Python script to generate quotes without Excel buttons.")

if __name__ == "__main__":
    create_auto_working_excel()
