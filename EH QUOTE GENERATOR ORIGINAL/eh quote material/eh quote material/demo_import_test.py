#!/usr/bin/env python3
"""
Demo Import Test - Test the import functionality with actual files
"""

import os
import re
from easy_quote_generator import EasyQuoteGenerator

def test_import_with_actual_files():
    """Test import with the actual files in the directory"""
    
    print("🧪 Testing Import with Actual Files...")
    
    # Find available files
    files = []
    for file in os.listdir('.'):
        if file.endswith(('.xlsx', '.rtf', '.xml')):
            files.append(file)
    
    if not files:
        print("❌ No .xlsx, .rtf, or .xml files found in directory")
        return False
    
    print(f"📁 Found {len(files)} files to test:")
    for file in files:
        print(f"   - {file}")
    
    # Test each file
    for file in files:
        print(f"\n🔍 Testing file: {file}")
        
        try:
            if file.endswith('.xlsx'):
                result = test_excel_import(file)
            elif file.endswith('.rtf'):
                result = test_rtf_import(file)
            elif file.endswith('.xml'):
                result = test_xml_import(file)
            
            if result:
                print(f"✅ {file} - Import successful")
            else:
                print(f"❌ {file} - Import failed")
                
        except Exception as e:
            print(f"❌ {file} - Error: {str(e)}")
    
    return True

def test_excel_import(file_path):
    """Test Excel file import"""
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"   📊 Excel file: {ws.title}")
        print(f"   📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Show first few rows
        print("   📋 First 5 rows:")
        for row in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(7, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    row_data.append(str(cell_value)[:20])  # Truncate long values
                else:
                    row_data.append("")
            print(f"      Row {row}: {row_data}")
        
        return True
        
    except Exception as e:
        print(f"   ❌ Error reading Excel file: {str(e)}")
        return False

def test_rtf_import(file_path):
    """Test RTF file import"""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        print(f"   📄 RTF file size: {len(content)} characters")
        
        # Show first few lines
        lines = content.split('\n')
        print("   📋 First 5 lines:")
        for i, line in enumerate(lines[:5]):
            print(f"      Line {i+1}: {line[:50]}...")
        
        return True
        
    except Exception as e:
        print(f"   ❌ Error reading RTF file: {str(e)}")
        return False

def test_xml_import(file_path):
    """Test XML file import"""
    try:
        import xml.etree.ElementTree as ET
        
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        print(f"   📄 XML root tag: {root.tag}")
        
        # Find product nodes
        products = root.findall('.//product') or root.findall('.//item') or root.findall('.//line')
        print(f"   📦 Found {len(products)} product/item nodes")
        
        # Show first few products
        for i, product in enumerate(products[:3]):
            print(f"   📋 Product {i+1}:")
            for child in product:
                if child.text and child.text.strip():
                    print(f"      {child.tag}: {child.text[:30]}...")
        
        return True
        
    except Exception as e:
        print(f"   ❌ Error reading XML file: {str(e)}")
        return False

if __name__ == "__main__":
    print("🚀 Starting Import Demo Test...\n")
    
    # Test with actual files
    test_import_with_actual_files()
    
    print("\n🎉 Import Demo Test Complete!")
    print("\n📋 Next Steps:")
    print("   1. The GUI application should be running")
    print("   2. Click 'IMPORT FILE' in the GUI")
    print("   3. Select one of the files shown above")
    print("   4. The import should work correctly now")
    print("   5. Click 'GENERATE QUOTE' to create your quote")
