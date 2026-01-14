#!/usr/bin/env python3
"""
Easy Quote Generator - No Excel VBA Needed!
Creates quotes directly using Python - much easier than Excel buttons
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import xml.etree.ElementTree as ET
import re

class EasyQuoteGenerator:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Easy Quote Generator")
        self.root.geometry("800x600")
        
        # Quote data
        self.quote_data = {
            'quote_number': f"QUOTE-{datetime.now().strftime('%Y%m%d')}-001",
            'quote_date': datetime.now().strftime('%m/%d/%Y'),
            'customer_company': '',
            'contact_person': '',
            'phone': '',
            'email': '',
            'customer_ref': '',
            'payment_terms': 'Net 30 Days',
            'delivery_terms': 'FOB Origin',
            'line_items': []
        }
        
        self.create_gui()
    
    def create_gui(self):
        """Create the user interface"""
        
        # Title
        title_label = tk.Label(self.root, text="EASY QUOTE GENERATOR", 
                              font=('Arial', 16, 'bold'), fg='white', bg='#366092')
        title_label.pack(fill='x', padx=10, pady=10)
        
        # File import section
        import_frame = tk.LabelFrame(self.root, text="FILE IMPORT", font=('Arial', 12, 'bold'))
        import_frame.pack(fill='x', padx=10, pady=5)
        
        tk.Button(import_frame, text="IMPORT FILE (.xlsx, .rtf, .xml)", 
                 command=self.import_file, bg='#2196F3', fg='white', 
                 font=('Arial', 11, 'bold')).pack(pady=10)
        
        self.status_label = tk.Label(import_frame, text="Ready to import files", 
                                   fg='blue')
        self.status_label.pack(pady=5)
        
        # Quote details section
        details_frame = tk.LabelFrame(self.root, text="QUOTE DETAILS", font=('Arial', 12, 'bold'))
        details_frame.pack(fill='x', padx=10, pady=5)
        
        # Create form fields
        fields = [
            ('Quote Number:', 'quote_number'),
            ('Quote Date:', 'quote_date'),
            ('Customer Company:', 'customer_company'),
            ('Contact Person:', 'contact_person'),
            ('Phone:', 'phone'),
            ('Email:', 'email'),
            ('Customer Reference:', 'customer_ref'),
            ('Payment Terms:', 'payment_terms'),
            ('Delivery Terms:', 'delivery_terms')
        ]
        
        self.entry_vars = {}
        for i, (label, key) in enumerate(fields):
            row = i // 2
            col = (i % 2) * 2
            
            tk.Label(details_frame, text=label, font=('Arial', 10, 'bold')).grid(
                row=row, column=col, sticky='w', padx=5, pady=2)
            
            var = tk.StringVar(value=self.quote_data[key])
            entry = tk.Entry(details_frame, textvariable=var, width=25)
            entry.grid(row=row, column=col+1, padx=5, pady=2, sticky='ew')
            
            self.entry_vars[key] = var
        
        # Configure grid weights
        details_frame.grid_columnconfigure(1, weight=1)
        details_frame.grid_columnconfigure(3, weight=1)
        
        # Line items section
        items_frame = tk.LabelFrame(self.root, text="LINE ITEMS", font=('Arial', 12, 'bold'))
        items_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Create treeview for line items
        columns = ('Description', 'Qty', 'Unit', 'Unit Price', 'Total')
        self.tree = ttk.Treeview(items_frame, columns=columns, show='headings', height=8)
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120)
        
        self.tree.pack(side='left', fill='both', expand=True)
        
        # Scrollbar for treeview
        scrollbar = ttk.Scrollbar(items_frame, orient='vertical', command=self.tree.yview)
        scrollbar.pack(side='right', fill='y')
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Buttons
        button_frame = tk.Frame(self.root)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Button(button_frame, text="ADD ITEM", command=self.add_item, 
                 bg='#4CAF50', fg='white', font=('Arial', 11, 'bold')).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="REMOVE ITEM", command=self.remove_item, 
                 bg='#FF5722', fg='white', font=('Arial', 11, 'bold')).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="CLEAR ALL", command=self.clear_all, 
                 bg='#FF9800', fg='white', font=('Arial', 11, 'bold')).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="GENERATE QUOTE", command=self.generate_quote, 
                 bg='#366092', fg='white', font=('Arial', 12, 'bold')).pack(side='right', padx=5)
        
        # Add some sample items
        self.add_sample_items()
    
    def add_sample_items(self):
        """Add sample line items"""
        sample_items = [
            ('Sample Product 1', '2', 'EA', '100.00'),
            ('Sample Product 2', '1', 'EA', '250.00')
        ]
        
        for item in sample_items:
            self.tree.insert('', 'end', values=item)
            self.quote_data['line_items'].append({
                'description': item[0],
                'quantity': item[1],
                'unit': item[2],
                'unit_price': item[3]
            })
    
    def import_file(self):
        """Import data from file"""
        file_path = filedialog.askopenfilename(
            title="Select File to Import",
            filetypes=[
                ("Excel Files", "*.xlsx"),
                ("RTF Files", "*.rtf"),
                ("XML Files", "*.xml"),
                ("All Files", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        try:
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.xlsx':
                self.import_excel_file(file_path)
            elif file_ext == '.rtf':
                self.import_rtf_file(file_path)
            elif file_ext == '.xml':
                self.import_xml_file(file_path)
            else:
                messagebox.showerror("Error", "Unsupported file type")
                
        except Exception as e:
            messagebox.showerror("Import Error", f"Error importing file: {str(e)}")
            self.status_label.config(text="Import failed", fg='red')
    
    def import_excel_file(self, file_path):
        """Import Excel file"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Clear existing items
        self.clear_items()
        
        # Find data
        item_count = 0
        for row in range(2, ws.max_row + 1):  # Skip header
            if item_count >= 10:  # Limit to 10 items
                break
                
            # Try to find data in common columns
            if ws.cell(row=row, column=1).value or ws.cell(row=row, column=2).value:
                description = ws.cell(row=row, column=2).value or ws.cell(row=row, column=3).value or ""
                quantity = ws.cell(row=row, column=3).value or ws.cell(row=row, column=4).value or "1"
                unit_price = ws.cell(row=row, column=4).value or ws.cell(row=row, column=5).value or ws.cell(row=row, column=6).value or "0.00"
                
                # Clean up the data - remove any non-numeric characters from quantity and price
                description = str(description).strip()
                quantity = str(quantity).strip()
                unit_price = str(unit_price).strip()
                
                # Skip if description contains header-like text
                if any(header in description.lower() for header in ['customer:', 'quantity:', 'unit:', 'price:', 'product:']):
                    continue
                
                # Clean quantity - extract only numbers
                try:
                    # Extract numbers from quantity
                    qty_match = re.search(r'(\d+\.?\d*)', quantity)
                    if qty_match:
                        quantity = qty_match.group(1)
                    else:
                        quantity = "1"
                except:
                    quantity = "1"
                
                # Clean unit price - extract only numbers
                try:
                    # Extract numbers from unit price
                    price_match = re.search(r'(\d+\.?\d*)', unit_price)
                    if price_match:
                        unit_price = price_match.group(1)
                    else:
                        unit_price = "0.00"
                except:
                    unit_price = "0.00"
                
                if description and len(description) > 3:  # Only add meaningful descriptions
                    self.tree.insert('', 'end', values=(description, quantity, 'EA', unit_price))
                    self.quote_data['line_items'].append({
                        'description': description,
                        'quantity': quantity,
                        'unit': 'EA',
                        'unit_price': unit_price
                    })
                    item_count += 1
        
        self.status_label.config(text=f"Excel file imported - {item_count} items loaded", fg='green')
        messagebox.showinfo("Import Complete", f"Excel file imported successfully!\n{item_count} items loaded.")
    
    def import_rtf_file(self, file_path):
        """Import RTF file"""
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        # Clear existing items
        self.clear_items()
        
        # Basic parsing for RTF
        lines = content.split('\n')
        item_count = 0
        
        for line in lines:
            line = line.strip()
            if len(line) > 10 and 'product' in line.lower() and item_count < 10:
                self.tree.insert('', 'end', values=(line, '1', 'EA', '0.00'))
                self.quote_data['line_items'].append({
                    'description': line,
                    'quantity': '1',
                    'unit': 'EA',
                    'unit_price': '0.00'
                })
                item_count += 1
        
        self.status_label.config(text=f"RTF file imported - {item_count} items loaded", fg='orange')
        messagebox.showinfo("Import Complete", f"RTF file imported with basic parsing!\n{item_count} items loaded.\nYou may need to adjust quantities and prices manually.")
    
    def import_xml_file(self, file_path):
        """Import XML file"""
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        # Clear existing items
        self.clear_items()
        
        # Look for product nodes
        products = root.findall('.//product') or root.findall('.//item') or root.findall('.//line')
        
        item_count = 0
        for product in products:
            if item_count >= 10:
                break
                
            description = ""
            quantity = "1"
            unit_price = "0.00"
            
            # Try to get description
            desc_elem = product.find('description') or product.find('name') or product.find('title')
            if desc_elem is not None:
                description = desc_elem.text or ""
            
            # Try to get quantity
            qty_elem = product.find('quantity') or product.find('qty')
            if qty_elem is not None:
                quantity = qty_elem.text or "1"
            
            # Try to get price
            price_elem = product.find('price') or product.find('unitprice') or product.find('cost')
            if price_elem is not None:
                unit_price = price_elem.text or "0.00"
            
            if description:
                self.tree.insert('', 'end', values=(description, quantity, 'EA', unit_price))
                self.quote_data['line_items'].append({
                    'description': description,
                    'quantity': quantity,
                    'unit': 'EA',
                    'unit_price': unit_price
                })
                item_count += 1
        
        self.status_label.config(text=f"XML file imported - {item_count} items loaded", fg='green')
        messagebox.showinfo("Import Complete", f"XML file imported successfully!\n{item_count} items loaded.")
    
    def add_item(self):
        """Add new line item"""
        dialog = ItemDialog(self.root)
        if dialog.result:
            self.tree.insert('', 'end', values=dialog.result)
            self.quote_data['line_items'].append({
                'description': dialog.result[0],
                'quantity': dialog.result[1],
                'unit': dialog.result[2],
                'unit_price': dialog.result[3]
            })
    
    def remove_item(self):
        """Remove selected line item"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            self.tree.delete(selection[0])
            
            # Remove from data
            for i, line_item in enumerate(self.quote_data['line_items']):
                if (line_item['description'] == item['values'][0] and 
                    line_item['quantity'] == item['values'][1]):
                    del self.quote_data['line_items'][i]
                    break
    
    def clear_items(self):
        """Clear all line items"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.quote_data['line_items'] = []
    
    def clear_all(self):
        """Clear all data"""
        self.clear_items()
        for key, var in self.entry_vars.items():
            if key == 'quote_number':
                var.set(f"QUOTE-{datetime.now().strftime('%Y%m%d')}-001")
            elif key == 'quote_date':
                var.set(datetime.now().strftime('%m/%d/%Y'))
            elif key in ['payment_terms', 'delivery_terms']:
                var.set(self.quote_data[key])
            else:
                var.set('')
        self.status_label.config(text="All data cleared", fg='blue')
    
    def generate_quote(self):
        """Generate quote Excel file"""
        try:
            # Update quote data from form
            for key, var in self.entry_vars.items():
                self.quote_data[key] = var.get()
            
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Quote"
            
            # Define styles
            header_font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
            subheader_font = Font(name='Arial', size=12, bold=True)
            normal_font = Font(name='Arial', size=10)
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
            thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), 
                                top=Side(style='thick'), bottom=Side(style='thick'))
            
            # Set column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 6
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            
            # Header
            ws['A1'] = 'QUOTE'
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:F1')
            ws.row_dimensions[1].height = 40
            
            # Company info
            ws['A2'] = 'KINDER MORGAN'
            ws['A2'].font = Font(name='Arial', size=12, bold=True)
            ws['A3'] = '1234 Energy Drive'
            ws['A4'] = 'Houston, TX 77002'
            ws['A5'] = 'Phone: (713) 369-9000'
            ws['A6'] = 'Email: quotes@kindermorgan.com'
            
            # Quote details
            ws['D2'] = 'Quote No.:'
            ws['D2'].font = subheader_font
            ws['E2'] = self.quote_data['quote_number']
            ws['D3'] = 'Date:'
            ws['D3'].font = subheader_font
            ws['E3'] = self.quote_data['quote_date']
            ws['D4'] = 'Customer Ref:'
            ws['D4'].font = subheader_font
            ws['E4'] = self.quote_data['customer_ref']
            
            # Customer info
            ws['A8'] = 'CUSTOMER INFORMATION'
            ws['A8'].font = subheader_font
            ws['A8'].fill = light_fill
            ws.merge_cells('A8:F8')
            
            ws['A9'] = 'Company:'
            ws['A9'].font = subheader_font
            ws['B9'] = self.quote_data['customer_company']
            ws['A10'] = 'Contact:'
            ws['A10'].font = subheader_font
            ws['B10'] = self.quote_data['contact_person']
            ws['A11'] = 'Phone:'
            ws['A11'].font = subheader_font
            ws['B11'] = self.quote_data['phone']
            ws['A12'] = 'Email:'
            ws['A12'].font = subheader_font
            ws['B12'] = self.quote_data['email']
            
            # Line items header
            ws['A14'] = 'Item #'
            ws['B14'] = 'Qty'
            ws['C14'] = 'Unit'
            ws['D14'] = 'Description'
            ws['E14'] = 'Unit Price'
            ws['F14'] = 'Total Price'
            
            # Format header
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}14'].font = subheader_font
                ws[f'{col}14'].fill = header_fill
                ws[f'{col}14'].alignment = Alignment(horizontal='center')
                ws[f'{col}14'].border = thick_border
            
            # Add line items
            last_row = 14
            subtotal = 0
            
            for i, item in enumerate(self.quote_data['line_items'], 1):
                last_row += 1
                
                # Safely convert quantity and unit price to float
                try:
                    quantity = float(str(item['quantity']).replace(',', '')) if item['quantity'] else 0
                except (ValueError, TypeError):
                    quantity = 0
                
                try:
                    unit_price = float(str(item['unit_price']).replace(',', '')) if item['unit_price'] else 0
                except (ValueError, TypeError):
                    unit_price = 0
                
                total_price = quantity * unit_price
                subtotal += total_price
                
                ws[f'A{last_row}'] = i
                ws[f'B{last_row}'] = item['quantity']
                ws[f'C{last_row}'] = item['unit']
                ws[f'D{last_row}'] = item['description']
                ws[f'E{last_row}'] = unit_price
                ws[f'F{last_row}'] = total_price
                
                # Format price columns
                ws[f'E{last_row}'].number_format = '$#,##0.00'
                ws[f'F{last_row}'].number_format = '$#,##0.00'
                
                # Add borders
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{last_row}'].border = thin_border
            
            # Totals
            tax_rate = 0.05
            tax_amount = subtotal * tax_rate
            freight = 0.00
            total = subtotal + tax_amount + freight
            
            ws[f'A{last_row + 2}'] = 'Subtotal:'
            ws[f'A{last_row + 2}'].font = subheader_font
            ws[f'F{last_row + 2}'] = subtotal
            ws[f'F{last_row + 2}'].number_format = '$#,##0.00'
            
            ws[f'A{last_row + 3}'] = 'Tax (5%):'
            ws[f'A{last_row + 3}'].font = subheader_font
            ws[f'F{last_row + 3}'] = tax_amount
            ws[f'F{last_row + 3}'].number_format = '$#,##0.00'
            
            ws[f'A{last_row + 4}'] = 'Freight:'
            ws[f'A{last_row + 4}'].font = subheader_font
            ws[f'F{last_row + 4}'] = freight
            ws[f'F{last_row + 4}'].number_format = '$#,##0.00'
            
            ws[f'A{last_row + 5}'] = 'TOTAL:'
            ws[f'A{last_row + 5}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'A{last_row + 5}'].fill = light_fill
            ws[f'F{last_row + 5}'] = total
            ws[f'F{last_row + 5}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'F{last_row + 5}'].fill = light_fill
            ws[f'F{last_row + 5}'].number_format = '$#,##0.00'
            
            # Add borders to totals
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{last_row + 5}'].border = thick_border
            
            # Terms
            ws[f'A{last_row + 7}'] = 'TERMS AND CONDITIONS'
            ws[f'A{last_row + 7}'].font = subheader_font
            ws[f'A{last_row + 7}'].fill = light_fill
            ws.merge_cells(f'A{last_row + 7}:F{last_row + 7}')
            
            ws[f'A{last_row + 8}'] = f"Payment Terms: {self.quote_data['payment_terms']}"
            ws[f'A{last_row + 9}'] = f"Delivery: {self.quote_data['delivery_terms']}"
            ws[f'A{last_row + 10}'] = "Validity: 30 days from quote date"
            ws[f'A{last_row + 11}'] = "All prices subject to change without notice"
            ws[f'A{last_row + 12}'] = "Freight charges not included unless specified"
            
            # Auto-fit columns
            for col in range(1, 7):  # Columns A through F
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col)
                for row in range(1, last_row + 15):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            filename = f"quote_{self.quote_data['quote_number'].replace('/', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            messagebox.showinfo("Quote Generated", f"Quote generated successfully!\nSaved as: {filename}")
            self.status_label.config(text=f"Quote generated: {filename}", fg='green')
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generating quote: {str(e)}")
    
    def run_command_line(self):
        """Run in command line mode"""
        print("\n" + "="*60)
        print("🚀 EASY QUOTE GENERATOR - COMMAND LINE MODE")
        print("="*60)
        
        while True:
            print("\n📋 MENU:")
            print("1. Import file (.xlsx, .rtf, .xml)")
            print("2. Add line item manually")
            print("3. View current quote data")
            print("4. Generate quote")
            print("5. Clear all data")
            print("6. Exit")
            
            choice = input("\nEnter your choice (1-6): ").strip()
            
            if choice == "1":
                self.import_file_cli()
            elif choice == "2":
                self.add_item_cli()
            elif choice == "3":
                self.view_data_cli()
            elif choice == "4":
                self.generate_quote_cli()
            elif choice == "5":
                self.clear_all_cli()
            elif choice == "6":
                print("👋 Goodbye!")
                break
            else:
                print("❌ Invalid choice. Please enter 1-6.")
    
    def import_file_cli(self):
        """Import file in command line mode"""
        print("\n📁 Available files:")
        files = [f for f in os.listdir('.') if f.endswith(('.xlsx', '.rtf', '.xml'))]
        for i, file in enumerate(files, 1):
            print(f"   {i}. {file}")
        
        if not files:
            print("❌ No .xlsx, .rtf, or .xml files found in current directory")
            return
        
        try:
            choice = int(input(f"\nSelect file (1-{len(files)}): ")) - 1
            if 0 <= choice < len(files):
                file_path = files[choice]
                print(f"📥 Importing {file_path}...")
                
                # Import the file
                if file_path.endswith('.xlsx'):
                    self.import_excel_file(file_path)
                elif file_path.endswith('.rtf'):
                    self.import_rtf_file(file_path)
                elif file_path.endswith('.xml'):
                    self.import_xml_file(file_path)
                
                print(f"✅ File imported successfully!")
            else:
                print("❌ Invalid file selection")
        except ValueError:
            print("❌ Please enter a valid number")
        except Exception as e:
            print(f"❌ Error importing file: {e}")
    
    def add_item_cli(self):
        """Add item in command line mode"""
        print("\n➕ ADD LINE ITEM:")
        description = input("Description: ").strip()
        if not description:
            print("❌ Description cannot be empty")
            return
        
        quantity = input("Quantity (default: 1): ").strip() or "1"
        unit = input("Unit (default: EA): ").strip() or "EA"
        unit_price = input("Unit Price (default: 0.00): ").strip() or "0.00"
        
        # Add to data
        self.quote_data['line_items'].append({
            'description': description,
            'quantity': quantity,
            'unit': unit,
            'unit_price': unit_price
        })
        
        print(f"✅ Added: {description} - Qty: {quantity} - Price: ${unit_price}")
    
    def view_data_cli(self):
        """View current data in command line mode"""
        print("\n📊 CURRENT QUOTE DATA:")
        print(f"Quote Number: {self.quote_data['quote_number']}")
        print(f"Quote Date: {self.quote_data['quote_date']}")
        print(f"Customer Company: {self.quote_data['customer_company']}")
        print(f"Contact Person: {self.quote_data['contact_person']}")
        print(f"Phone: {self.quote_data['phone']}")
        print(f"Email: {self.quote_data['email']}")
        print(f"Customer Reference: {self.quote_data['customer_ref']}")
        print(f"Payment Terms: {self.quote_data['payment_terms']}")
        print(f"Delivery Terms: {self.quote_data['delivery_terms']}")
        
        print(f"\n📦 LINE ITEMS ({len(self.quote_data['line_items'])}):")
        if self.quote_data['line_items']:
            for i, item in enumerate(self.quote_data['line_items'], 1):
                print(f"   {i}. {item['description']} - Qty: {item['quantity']} {item['unit']} - ${item['unit_price']}")
        else:
            print("   No items added yet")
    
    def generate_quote_cli(self):
        """Generate quote in command line mode"""
        if not self.quote_data['line_items']:
            print("❌ No line items to generate quote. Add some items first.")
            return
        
        print("\n🔄 Generating quote...")
        try:
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Quote"
            
            # Define styles
            header_font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
            subheader_font = Font(name='Arial', size=12, bold=True)
            normal_font = Font(name='Arial', size=10)
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
            thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), 
                                top=Side(style='thick'), bottom=Side(style='thick'))
            
            # Set column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 6
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            
            # Header
            ws['A1'] = 'QUOTE'
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:F1')
            ws.row_dimensions[1].height = 40
            
            # Company info
            ws['A2'] = 'KINDER MORGAN'
            ws['A2'].font = Font(name='Arial', size=12, bold=True)
            ws['A3'] = '1234 Energy Drive'
            ws['A4'] = 'Houston, TX 77002'
            ws['A5'] = 'Phone: (713) 369-9000'
            ws['A6'] = 'Email: quotes@kindermorgan.com'
            
            # Quote details
            ws['D2'] = 'Quote No.:'
            ws['D2'].font = subheader_font
            ws['E2'] = self.quote_data['quote_number']
            ws['D3'] = 'Date:'
            ws['D3'].font = subheader_font
            ws['E3'] = self.quote_data['quote_date']
            ws['D4'] = 'Customer Ref:'
            ws['D4'].font = subheader_font
            ws['E4'] = self.quote_data['customer_ref']
            
            # Customer info
            ws['A8'] = 'CUSTOMER INFORMATION'
            ws['A8'].font = subheader_font
            ws['A8'].fill = light_fill
            ws.merge_cells('A8:F8')
            
            ws['A9'] = 'Company:'
            ws['A9'].font = subheader_font
            ws['B9'] = self.quote_data['customer_company']
            ws['A10'] = 'Contact:'
            ws['A10'].font = subheader_font
            ws['B10'] = self.quote_data['contact_person']
            ws['A11'] = 'Phone:'
            ws['A11'].font = subheader_font
            ws['B11'] = self.quote_data['phone']
            ws['A12'] = 'Email:'
            ws['A12'].font = subheader_font
            ws['B12'] = self.quote_data['email']
            
            # Line items header
            ws['A14'] = 'Item #'
            ws['B14'] = 'Qty'
            ws['C14'] = 'Unit'
            ws['D14'] = 'Description'
            ws['E14'] = 'Unit Price'
            ws['F14'] = 'Total Price'
            
            # Format header
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}14'].font = subheader_font
                ws[f'{col}14'].fill = header_fill
                ws[f'{col}14'].alignment = Alignment(horizontal='center')
                ws[f'{col}14'].border = thick_border
            
            # Add line items
            last_row = 14
            subtotal = 0
            
            for i, item in enumerate(self.quote_data['line_items'], 1):
                last_row += 1
                
                # Safely convert quantity and unit price to float
                try:
                    quantity = float(str(item['quantity']).replace(',', '')) if item['quantity'] else 0
                except (ValueError, TypeError):
                    quantity = 0
                
                try:
                    unit_price = float(str(item['unit_price']).replace(',', '')) if item['unit_price'] else 0
                except (ValueError, TypeError):
                    unit_price = 0
                
                total_price = quantity * unit_price
                subtotal += total_price
                
                ws[f'A{last_row}'] = i
                ws[f'B{last_row}'] = item['quantity']
                ws[f'C{last_row}'] = item['unit']
                ws[f'D{last_row}'] = item['description']
                ws[f'E{last_row}'] = unit_price
                ws[f'F{last_row}'] = total_price
                
                # Format price columns
                ws[f'E{last_row}'].number_format = '$#,##0.00'
                ws[f'F{last_row}'].number_format = '$#,##0.00'
                
                # Add borders
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{last_row}'].border = thin_border
            
            # Totals
            tax_rate = 0.05
            tax_amount = subtotal * tax_rate
            freight = 0.00
            total = subtotal + tax_amount + freight
            
            ws[f'A{last_row + 2}'] = 'Subtotal:'
            ws[f'A{last_row + 2}'].font = subheader_font
            ws[f'F{last_row + 2}'] = subtotal
            ws[f'F{last_row + 2}'].number_format = '$#,##0.00'
            
            ws[f'A{last_row + 3}'] = 'Tax (5%):'
            ws[f'A{last_row + 3}'].font = subheader_font
            ws[f'F{last_row + 3}'] = tax_amount
            ws[f'F{last_row + 3}'].number_format = '$#,##0.00'
            
            ws[f'A{last_row + 4}'] = 'Freight:'
            ws[f'A{last_row + 4}'].font = subheader_font
            ws[f'F{last_row + 4}'] = freight
            ws[f'F{last_row + 4}'].number_format = '$#,##0.00'
            
            ws[f'A{last_row + 5}'] = 'TOTAL:'
            ws[f'A{last_row + 5}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'A{last_row + 5}'].fill = light_fill
            ws[f'F{last_row + 5}'] = total
            ws[f'F{last_row + 5}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'F{last_row + 5}'].fill = light_fill
            ws[f'F{last_row + 5}'].number_format = '$#,##0.00'
            
            # Add borders to totals
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{last_row + 5}'].border = thick_border
            
            # Terms
            ws[f'A{last_row + 7}'] = 'TERMS AND CONDITIONS'
            ws[f'A{last_row + 7}'].font = subheader_font
            ws[f'A{last_row + 7}'].fill = light_fill
            ws.merge_cells(f'A{last_row + 7}:F{last_row + 7}')
            
            ws[f'A{last_row + 8}'] = f"Payment Terms: {self.quote_data['payment_terms']}"
            ws[f'A{last_row + 9}'] = f"Delivery: {self.quote_data['delivery_terms']}"
            ws[f'A{last_row + 10}'] = "Validity: 30 days from quote date"
            ws[f'A{last_row + 11}'] = "All prices subject to change without notice"
            ws[f'A{last_row + 12}'] = "Freight charges not included unless specified"
            
            # Auto-fit columns
            for col in range(1, 7):  # Columns A through F
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col)
                for row in range(1, last_row + 15):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            filename = f"quote_{self.quote_data['quote_number'].replace('/', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            print(f"✅ Quote generated successfully!")
            print(f"📁 Saved as: {filename}")
            print(f"💰 Subtotal: ${subtotal:,.2f}")
            print(f"💰 Tax (5%): ${tax_amount:,.2f}")
            print(f"💰 Total: ${total:,.2f}")
            print(f"📦 Items: {len(self.quote_data['line_items'])}")
            
        except Exception as e:
            print(f"❌ Error generating quote: {e}")
    
    def clear_all_cli(self):
        """Clear all data in command line mode"""
        self.quote_data['line_items'] = []
        self.quote_data['quote_number'] = f"QUOTE-{datetime.now().strftime('%Y%m%d')}-001"
        self.quote_data['quote_date'] = datetime.now().strftime('%m/%d/%Y')
        self.quote_data['customer_company'] = ''
        self.quote_data['contact_person'] = ''
        self.quote_data['phone'] = ''
        self.quote_data['email'] = ''
        self.quote_data['customer_ref'] = ''
        self.quote_data['payment_terms'] = 'Net 30 Days'
        self.quote_data['delivery_terms'] = 'FOB Origin'
        
        print("✅ All data cleared!")

    def run(self):
        """Run the application"""
        try:
            self.root.mainloop()
        except Exception as e:
            print(f"Error running GUI: {e}")
            print("Falling back to command line mode...")
            self.run_command_line()

class ItemDialog:
    def __init__(self, parent):
        self.result = None
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Add Line Item")
        self.dialog.geometry("400x200")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Create form
        tk.Label(self.dialog, text="Description:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.desc_var = tk.StringVar()
        tk.Entry(self.dialog, textvariable=self.desc_var, width=40).grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(self.dialog, text="Quantity:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.qty_var = tk.StringVar(value="1")
        tk.Entry(self.dialog, textvariable=self.qty_var, width=10).grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        tk.Label(self.dialog, text="Unit:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.unit_var = tk.StringVar(value="EA")
        tk.Entry(self.dialog, textvariable=self.unit_var, width=10).grid(row=2, column=1, sticky='w', padx=5, pady=5)
        
        tk.Label(self.dialog, text="Unit Price:", font=('Arial', 10, 'bold')).grid(row=3, column=0, sticky='w', padx=5, pady=5)
        self.price_var = tk.StringVar(value="0.00")
        tk.Entry(self.dialog, textvariable=self.price_var, width=10).grid(row=3, column=1, sticky='w', padx=5, pady=5)
        
        # Buttons
        button_frame = tk.Frame(self.dialog)
        button_frame.grid(row=4, column=0, columnspan=2, pady=20)
        
        tk.Button(button_frame, text="Add", command=self.add_item, 
                 bg='#4CAF50', fg='white', font=('Arial', 10, 'bold')).pack(side='left', padx=5)
        tk.Button(button_frame, text="Cancel", command=self.cancel, 
                 bg='#FF5722', fg='white', font=('Arial', 10, 'bold')).pack(side='left', padx=5)
        
        # Center the dialog
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")
    
    def add_item(self):
        """Add the item"""
        if self.desc_var.get().strip():
            self.result = (
                self.desc_var.get().strip(),
                self.qty_var.get().strip(),
                self.unit_var.get().strip(),
                self.price_var.get().strip()
            )
            self.dialog.destroy()
    
    def cancel(self):
        """Cancel adding item"""
        self.dialog.destroy()

if __name__ == "__main__":
    app = EasyQuoteGenerator()
    app.run()
